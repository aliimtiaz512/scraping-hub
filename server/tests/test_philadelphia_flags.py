"""Excluded niches: which PHLContracts bids come back red, and which do not.

The client keeps twenty service categories out of scope. A run does not drop
those bids — the city's Open Bids list is the deliverable — but the sheet marks
them, so a reviewer can skip past them instead of reading every title.

    server/.venv/bin/python -m pytest server/tests/test_philadelphia_flags.py
"""

import os
import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.scrapers.philadelphia import export, flags  # noqa: E402
from app.scrapers.sam.engine.evaluator import RULE_B  # noqa: E402


# =============================================================================
# The list
# =============================================================================


def test_the_niches_are_the_clients_own_list_not_a_second_copy():
    """`RULE_B` is the same client's same twenty categories, already in this
    codebase. A copy here would be a second place to edit when they add a
    twenty-first, and the two would drift without anyone noticing."""
    assert len(RULE_B) == 20
    for name in RULE_B.values():
        assert " ".join(name.lower().split()) in flags.EXCLUDED_NICHES


@pytest.mark.parametrize("written,also", [
    ("construction & demolition services", "construction and demolition services"),
    ("religious & education coordinator", "religious and education coordinator"),
    ("research & development", "research and development"),
])
def test_both_the_ampersand_and_the_word_are_matched(written, also):
    """The client writes "&" and a bid says "and"; they are the same category,
    so the pair is generated rather than listed twice by hand."""
    assert written in flags.EXCLUDED_NICHES
    assert also in flags.EXCLUDED_NICHES


def test_the_longest_term_is_reported_when_several_match():
    """A "waste management services" bid also contains "management services";
    naming the shorter one would send a reviewer to the wrong category."""
    assert flags.check({"description": "Waste Management Services citywide"})[1] == (
        "waste management services"
    )


# =============================================================================
# Matching
# =============================================================================


@pytest.mark.parametrize("description,niche", [
    ("Annual Custodial Services for City Facilities", "custodial services"),
    ("Citywide Waste Management Services", "waste management services"),
    ("Research and Development partnership", "research and development"),
    ("Annual audit of financial statements", "audit"),
    ("Rental of Equipment for the Streets Department", "rental of equipment"),
    ("Hotel Room Booking and Lodging for conferences", "hotel room booking and lodging"),
    ("Aircraft Lavatory Services at PHL", "aircraft lavatory services"),
])
def test_a_bid_in_an_excluded_niche_is_flagged(description, niche):
    flagged, matched = flags.check({"description": description})

    assert flagged is True
    assert matched == niche


@pytest.mark.parametrize("description", [
    "Auditorium lighting replacement",        # 'audit' inside a longer word
    "Auditory testing equipment",
    "Supply of laptops and docking stations",
    "Jackhammers and related hand tools",
    "License tag stickers",
])
def test_a_bid_that_merely_contains_the_letters_is_not_flagged(description):
    """Substring matching would flag an auditorium as an audit service — the
    same trap the GSA screen fell into. Every term is matched on whole words."""
    assert flags.check({"description": description}) == (False, "")


def test_a_niche_broken_across_a_line_break_still_matches():
    """The portal's own descriptions are full of them."""
    flagged, matched = flags.check({"description": "Annual Custodial\n  Services required"})

    assert flagged is True
    assert matched == "custodial services"


def test_the_match_is_case_insensitive():
    assert flags.check({"description": "CUSTODIAL SERVICES"})[0] is True
    assert flags.check({"description": "custodial services"})[0] is True


def test_the_title_and_category_are_read_as_well_as_the_description():
    """A niche named in one field and not another is still that niche."""
    assert flags.check({"title": "Custodial Services"})[0] is True
    assert flags.check({"category": "Waste Management Services"})[0] is True


def test_a_bid_with_nothing_to_read_is_not_flagged():
    assert flags.check({}) == (False, "")
    assert flags.check({"description": "", "title": None}) == (False, "")


# =============================================================================
# The sheet
# =============================================================================


def _sheet(records):
    from openpyxl import load_workbook

    out = Path(tempfile.mkdtemp()) / "Philadelphia_Bids_Summary.xlsx"
    export.generate_excel_from_records(records, out)
    return load_workbook(out).active


def _is_red(cell) -> bool:
    return str(cell.fill.fgColor.rgb or "").endswith("FFC7CE")


def test_a_flagged_bid_is_red_across_the_whole_row():
    sheet = _sheet([{"bid_number": "23-10495",
                     "description": "Annual Custodial Services for City Facilities"}])

    assert all(_is_red(cell) for cell in sheet[2]), "the row is not fully highlighted"


def test_a_bid_in_scope_is_left_alone():
    sheet = _sheet([{"bid_number": "23-10498", "description": "Supply of laptops"}])

    assert not any(_is_red(cell) for cell in sheet[2])


def test_the_status_and_reason_columns_say_why():
    """A row shaded red with nothing naming the niche makes a reviewer re-derive
    it by reading the description — the work this is meant to save."""
    sheet = _sheet([{"bid_number": "23-10495", "description": "Annual Custodial Services"}])
    # Located by name, not by position: a column added ahead of these is an
    # ordinary change to the sheet and must not read as a lost flag.
    headers = [c.value for c in sheet[1]]

    assert "Niche Flag" in headers and "Niche Flag Reason" in headers
    assert sheet.cell(2, headers.index("Niche Flag") + 1).value == "FLAGGED"
    assert "custodial services" in sheet.cell(2, headers.index("Niche Flag Reason") + 1).value


def test_the_header_row_keeps_its_own_colour_when_every_bid_is_flagged():
    """Highlighting runs before the header is styled, so a predicate matching
    every row cannot leave the header red too."""
    sheet = _sheet([{"bid_number": "1", "description": "Custodial Services"},
                    {"bid_number": "2", "description": "Waste Management Services"}])

    assert not _is_red(sheet.cell(1, 1)), "the header was highlighted"
    assert all(_is_red(sheet.cell(r, 1)) for r in (2, 3))


def test_a_sheet_rebuilt_from_the_database_marks_the_same_rows():
    """The flag is derived from columns the row already carries, so a sheet
    rebuilt months later cannot disagree with the one that shipped."""
    class Row:
        bid_number = "23-10495"
        description = "Annual Custodial Services"
        title = None
        category = None
        buyer_description = None
        items = []
        extra_header_data = {}
        additional_header = None

    assert export._cell(Row(), "flag_status") == "FLAGGED"
    assert "custodial services" in export._cell(Row(), "flag_reason")


def test_the_shared_writer_highlights_nothing_when_no_predicate_is_given():
    """Every other portal calls write_table without one and must be unchanged."""
    from app.core import excel_style

    workbook, sheet = excel_style.new_workbook("t")
    excel_style.write_table(sheet, ["A"], [["Custodial Services"], ["Laptops"]])

    assert not any(_is_red(sheet.cell(r, 1)) for r in (2, 3))
