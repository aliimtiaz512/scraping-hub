"""The shared evaluation matrix, run over PHLContracts bids.

The matrix is not reimplemented for Philadelphia — `app.scrapers.sam.evaluation`
is the one engine and this checks the mapping into it, the hint that makes it
usable on a one-line description, and that running it changed none of the three
things a PHL run delivers.

    server/.venv/bin/python -m pytest server/tests/test_philadelphia_evaluation.py
"""

import os
import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.scrapers.philadelphia import evaluation, export, storage  # noqa: E402
from app.scrapers.philadelphia.models import EXCEL_COLUMNS  # noqa: E402


# =============================================================================
# It is the same engine, not a copy
# =============================================================================


def test_the_matrix_is_sams_engine_not_a_reimplementation():
    """One set of rules across three portals is the only way a PURSUE on this
    sheet means what a PURSUE means on SAM's."""
    from app.scrapers.sam.evaluation import evaluate as sam_evaluate

    assert evaluation.evaluate_bid is sam_evaluate


def test_the_hardware_rule_is_unisons_not_a_second_copy():
    from app.scrapers.unison.evaluation import requirement_hint

    assert evaluation._line_item_hint is requirement_hint


# =============================================================================
# The requirement-type problem, and the line-item hint that solves it
# =============================================================================


def test_a_quantified_item_table_proves_hardware():
    """A Philadelphia description is one line, so without this the funnel has
    only prose to go on and sends a box of jackhammers to MANUAL_REVIEW."""
    hint, why = evaluation.hint_for({
        "description": "Jackhammers",
        "items": [{"name": "Jackhammer 60lb", "quantity": "4", "unit": "EA"}],
    })

    assert hint == "HARDWARE"
    assert "line item" in why


def test_a_bid_with_no_items_gets_no_hint():
    hint, why = evaluation.hint_for({"description": "Consulting", "items": []})

    assert hint is None
    assert why == "no line items to judge from"


def test_a_service_led_description_overrides_the_item_table():
    """The hint proves *what is being procured*; a description that names a
    service settles it regardless of how the rows look."""
    hint, _ = evaluation.hint_for({
        "description": "Custodial services for various City agencies",
        "items": [{"name": "Mop", "quantity": "10", "unit": "EA"}],
    })

    assert hint is None


@pytest.mark.parametrize("description,items,expected", [
    ("Jackhammers", [{"name": "Jackhammer", "quantity": "4", "unit": "EA"}], "PURSUE"),
    ("6x4 CNG Truck with Compactor Body",
     [{"name": "FREIGHTLINER CHASSIS", "quantity": "2", "unit": "EA"}], "PURSUE"),
    # The hint can only promote to HARDWARE — it never rescues a Rule B bid.
    ("Custodial Service and Maintenance Supplies",
     [{"name": "Mop", "quantity": "10", "unit": "EA"}], "REJECT"),
])
def test_the_verdict_for_a_real_philadelphia_bid(description, items, expected):
    verdict = evaluation.evaluate({
        "bid_number": "B1", "description": description, "items": items,
    })

    assert verdict["decision"] == expected


def test_an_excluded_service_is_rejected_by_the_matrixs_own_rule():
    verdict = evaluation.evaluate({
        "bid_number": "B2624630",
        "description": "Custodial Service and Maintenance Supplies for City agencies",
    })

    assert verdict["decision"] == "REJECT"
    assert verdict["rule"] == "B10"
    assert "Custodial" in verdict["reason"]


# =============================================================================
# What the engine is given
# =============================================================================


def test_the_item_lines_reach_the_engine():
    """A Philadelphia bid says what it is buying in its item blocks, not in its
    one-line description."""
    text = evaluation.build_full_text({
        "description": "Annual requirement",
        "items": [{"name": "FREIGHTLINER 114SD CHASSIS", "quantity": "2", "unit": "EA"}],
    })

    assert "FREIGHTLINER 114SD CHASSIS" in text
    assert "EA" in text


def test_the_administrative_header_rows_are_left_out():
    """Addresses and phone numbers add words without adding signal."""
    text = evaluation.build_full_text({
        "description": "Pumps",
        "extra_header_data": {
            "Type Code": "MI - Micro Purchase",
            "Ship-to Address": "29th & Cambria Streets, Philadelphia PA",
            "Info Contact": "Michael Perce 215-459-4753",
        },
    })

    assert "Micro Purchase" in text
    assert "Cambria" not in text
    assert "215-459-4753" not in text


def test_an_evaluation_failure_leaves_the_bid_in_the_report(monkeypatch):
    """A bid without a verdict is still a bid the city published."""
    monkeypatch.setattr(
        evaluation, "evaluate_bid",
        lambda *a, **k: (_ for _ in ()).throw(RuntimeError("engine down")))

    verdict = evaluation.evaluate({"bid_number": "B1", "description": "Pumps"})

    assert verdict["decision"] == "PENDING"
    assert "RuntimeError" in verdict["reason"]


# =============================================================================
# The output package — the strict rule
# =============================================================================


def test_the_sheet_carries_the_matrix_columns_without_losing_the_old_ones():
    headers = [header for _, header in EXCEL_COLUMNS]

    for added in ("Evaluation Status", "Matrix Rule", "Evaluation Reason",
                  "Requirement Type"):
        assert added in headers
    # Everything the sheet carried before this change.
    for kept in ("Bid #", "Organization", "Buyer", "Description",
                 "Bid Opening Date", "Fiscal Year",
                 "Procurement / Solicitation Type", "Total Document Count",
                 "Detail URL"):
        assert kept in headers, f"{kept} was dropped from the summary sheet"


def test_the_verdict_reaches_the_sheet():
    from openpyxl import load_workbook

    record = {"bid_number": "B1", "description": "Jackhammers",
              "items": [{"name": "Jackhammer", "quantity": "4", "unit": "EA"}]}
    record.update(evaluation.evaluate(record))

    out = Path(tempfile.mkdtemp())
    export.generate_excel_from_records([record], storage.summary_path(out))
    sheet = load_workbook(storage.summary_path(out)).active
    headers = [c.value for c in sheet[1]]

    assert sheet.cell(2, headers.index("Evaluation Status") + 1).value == "PURSUE"
    assert sheet.cell(2, headers.index("Requirement Type") + 1).value == "HARDWARE"


def test_a_rejected_bid_keeps_its_row():
    """The verdict is a column, not a filter — a REJECT still ships its row and
    its document count, because a reader disagreeing with the matrix needs to
    see the bid to say so."""
    record = {"bid_number": "B2624630",
              "description": "Custodial Service and Maintenance Supplies",
              "documents_downloaded": 11}
    record.update(evaluation.evaluate(record))

    assert record["decision"] == "REJECT"
    assert record["documents_downloaded"] == 11


# =============================================================================
# Resolving the borderline — a sheet of MANUAL_REVIEW has evaluated nothing
# =============================================================================


def test_a_bid_with_quantified_goods_is_pursued_rather_than_queued():
    verdict = evaluation.evaluate({
        "bid_number": "B1", "description": "Office furniture for the Water Department",
        "items": [{"name": "Task chair", "quantity": "120", "unit": "EA"}],
    })

    assert verdict["decision"] == "PURSUE"
    assert verdict["requirement_type"] == "HARDWARE"


def test_a_service_the_allowed_list_does_not_name_is_rejected_rather_than_queued():
    """The funnel leaves this one undecided — it matches neither the excluded
    list nor the allowed one. Settling it is the point: every verdict left in
    the queue is a bid somebody has to read."""
    verdict = evaluation.evaluate({
        "bid_number": "B1", "description": "License Tag Stickers",
    })

    assert verdict["decision"] == "REJECT"


def test_a_settled_verdict_says_it_was_settled():
    """A run that decided forty bids without a person must be able to say which
    forty and why."""
    verdict = evaluation.evaluate({
        "bid_number": "B1", "description": "License Tag Stickers",
    })

    assert "settled by the matrix rather than sent to review" in verdict["reason"]


def test_a_bid_the_funnel_already_rejected_is_not_relabelled_as_settled():
    """Rule B #1 decides a maintenance bid outright; it never reached the
    resolver, so it must not claim to have been settled by it."""
    verdict = evaluation.evaluate({
        "bid_number": "B1",
        "description": "Swimming Pool Maintenance Services at recreation centers",
    })

    assert verdict["decision"] == "REJECT"
    assert verdict["rule"] == "B1"
    assert "settled by the matrix" not in verdict["reason"]


@pytest.mark.parametrize("description,judgeable", [
    ("License Tag Stickers", True),                    # short, but names a thing
    ("Jackhammers and hand tools", True),
    ("Swimming Pool Maintenance Services", True),
    ("Annual requirement", False),                     # names nothing
    ("See attached", False),
    ("Various items", False),
    ("", False),
])
def test_what_counts_as_enough_to_judge_on(description, judgeable):
    """Length is deliberately not the test — "License Tag Stickers" is twenty
    characters and perfectly clear, "Annual requirement" is eighteen and means
    nothing. What separates them is whether any word names a thing."""
    assert evaluation.judgeable({"description": description})[0] is judgeable


def test_a_bid_with_items_is_judgeable_however_thin_its_description():
    assert evaluation.judgeable({
        "description": "See attached",
        "items": [{"name": "Chevrolet Tahoe", "quantity": "48", "unit": "EA"}],
    })[0] is True


@pytest.mark.parametrize("description", ["Annual requirement", "See attached"])
def test_the_few_with_nothing_to_judge_on_still_go_to_a_person(description):
    """Inventing a verdict from nothing is worse than the review it saves."""
    verdict = evaluation.evaluate({"bid_number": "B1", "description": description})

    assert verdict["decision"] == "MANUAL_REVIEW"
    assert "names nothing to judge on" in verdict["reason"]


def test_the_resolver_never_touches_a_decided_verdict():
    """A REJECT from Rule B is not borderline and must come through unchanged."""
    decided = {"decision": "REJECT", "reason": "Excluded service category (Rule B #10)",
               "rule": "B10", "requirement_type": "SERVICE"}

    assert evaluation.resolve({"description": "Custodial Services"}, decided) == decided


def test_most_of_a_realistic_run_is_decided_without_a_person():
    """The measure that matters: a client who has to review most of the sheet
    has not been given an evaluation."""
    bids = [
        ("Jackhammers", [{"name": "Jackhammer", "quantity": "4", "unit": "EA"}]),
        ("6x4 CNG Truck with Compactor Body",
         [{"name": "FREIGHTLINER CHASSIS", "quantity": "2", "unit": "EA"}]),
        ("RPC Marked Police Vehicles",
         [{"name": "CHEVROLET TAHOE", "quantity": "48", "unit": "EA"}]),
        ("Custodial Service and Maintenance Supplies", []),
        ("Health AMS Tablecloths and Bubbles with Logos", []),
        ("License Tag Stickers", []),
        ("Swimming Pool Maintenance Services", []),
        ("Elevator inspection and repair services citywide", []),
        ("Annual requirement", []),
        ("See attached", []),
    ]
    verdicts = [
        evaluation.evaluate({"bid_number": f"B{i}", "description": d, "items": it})["decision"]
        for i, (d, it) in enumerate(bids)
    ]
    queued = verdicts.count("MANUAL_REVIEW")

    assert queued <= 2, f"{queued} of {len(bids)} still need a person"
    assert set(verdicts) <= {"PURSUE", "REJECT", "MANUAL_REVIEW"}


# =============================================================================
# The sheet reads the way a person reads a bid
# =============================================================================


def test_the_bid_is_identified_before_anything_judges_it():
    headers = [header for _, header in EXCEL_COLUMNS]

    assert headers[0] == "Bid #"
    assert headers.index("Description") < headers.index("Evaluation Status")
    assert headers.index("Bid Opening Date") < headers.index("Evaluation Status")
    assert headers.index("Evaluation Status") < headers.index("Detail URL")


def test_the_two_kinds_of_status_are_not_both_called_status():
    """One column from the matrix and one from the niche list, both named
    Status, is the kind of thing a reader has to stop and decode."""
    headers = [header for _, header in EXCEL_COLUMNS]

    assert "Evaluation Status" in headers
    assert "Niche Flag" in headers
    assert "Status" not in headers


# =============================================================================
# The sheet's colours: red is out, yellow needs a person, clean is worth reading
# =============================================================================


def _styled_sheet(records):
    from openpyxl import load_workbook

    for record in records:
        record.update(evaluation.evaluate(record))
    out = Path(tempfile.mkdtemp())
    export.generate_excel_from_records(records, storage.summary_path(out))
    return load_workbook(storage.summary_path(out)).active


def _fill(cell) -> str:
    return str(getattr(cell.fill.fgColor, "rgb", "") or "")[-6:]


def _font(cell) -> str:
    colour = getattr(cell.font.color, "rgb", None)
    return str(colour)[-6:] if isinstance(colour, str) else ""


@pytest.mark.parametrize("description,items,decision,fill,font", [
    ("Custodial Service and Maintenance Supplies", [], "REJECT", "FADBD8", "78281F"),
    ("Annual requirement", [], "MANUAL_REVIEW", "FCF3CF", "7D6608"),
    ("Jackhammers", [{"name": "Jackhammer", "quantity": "4", "unit": "EA"}], "PURSUE", "", ""),
])
def test_a_rows_colour_says_what_its_verdict_means(description, items, decision, fill, font):
    sheet = _styled_sheet([{"bid_number": "B1", "description": description, "items": items}])
    headers = [c.value for c in sheet[1]]

    assert sheet.cell(2, headers.index("Evaluation Status") + 1).value == decision
    assert _fill(sheet.cell(2, 1)) == (fill or "000000")
    if font:
        assert _font(sheet.cell(2, 1)) == font
        assert sheet.cell(2, 1).font.bold is True


def test_a_pursued_bid_is_left_completely_alone():
    """The rows worth reading are the ones without a colour — so nothing is
    applied to them at all, not even a white fill."""
    sheet = _styled_sheet([
        {"bid_number": "B1", "description": "Jackhammers",
         "items": [{"name": "Jackhammer", "quantity": "4", "unit": "EA"}]},
    ])

    assert all(_fill(cell) in ("", "000000") for cell in sheet[2])
    assert not any(cell.font.bold for cell in sheet[2])


def test_the_whole_row_is_tinted_not_just_the_status_cell():
    sheet = _styled_sheet([
        {"bid_number": "B1", "description": "Custodial Service and Maintenance Supplies"},
    ])

    assert {_fill(cell) for cell in sheet[2]} == {"FADBD8"}


def test_the_header_keeps_its_own_styling_whatever_the_rows_do():
    """Row tinting runs before the header is styled, so a sheet where every bid
    is rejected cannot leave the header red."""
    sheet = _styled_sheet([
        {"bid_number": "B1", "description": "Custodial Service and Maintenance Supplies"},
        {"bid_number": "B2", "description": "Elevator inspection and repair services"},
    ])

    assert _fill(sheet.cell(1, 1)) == "1F4E78"
    assert {_fill(sheet.cell(r, 1)) for r in (2, 3)} == {"FADBD8"}


def test_an_excluded_niche_is_red_even_without_a_matrix_verdict():
    """Red means out of scope and the niche list is one of the two things that
    can say so. One colour for one meaning — the Niche Flag column says which
    list it was."""
    sheet = _styled_sheet([{"bid_number": "B1", "description": "Annual Custodial Services"}])
    headers = [c.value for c in sheet[1]]

    assert sheet.cell(2, headers.index("Niche Flag") + 1).value == "FLAGGED"
    assert _fill(sheet.cell(2, 1)) == "FADBD8"


def test_the_colour_is_read_off_the_written_cells_not_re_derived():
    """A second pass over the same record could disagree with the first, and a
    row tinted red whose Evaluation Status reads PURSUE is worse than either
    answer alone."""
    from app.scrapers.philadelphia.export import _COLUMN_INDEX, _row_style

    values = [""] * len(EXCEL_COLUMNS)
    values[_COLUMN_INDEX["decision"]] = "REJECT"
    assert _row_style(values) == "reject"

    values[_COLUMN_INDEX["decision"]] = "MANUAL_REVIEW"
    assert _row_style(values) == "review"

    values[_COLUMN_INDEX["decision"]] = "PURSUE"
    assert _row_style(values) is None


def test_an_unevaluated_bid_is_not_silently_treated_as_pursued():
    """PENDING is an evaluation that failed, not a verdict — it needs a person,
    so it gets the yellow rather than passing as clean."""
    from app.scrapers.philadelphia.export import _COLUMN_INDEX, _row_style

    values = [""] * len(EXCEL_COLUMNS)
    values[_COLUMN_INDEX["decision"]] = "PENDING"

    assert _row_style(values) == "review"


def test_the_portal_delivers_a_bare_sheet_not_a_zip():
    from app.core import exports

    assert "philadelphia" in exports.EXCEL_ONLY_PORTALS
    assert "philadelphia" not in exports.DOC_PORTALS
