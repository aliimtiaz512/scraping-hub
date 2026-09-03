"""RideMetro's Euna Supplier Network sweep: agency filtering, row extraction,
and the agency-grouped report.

Pure logic and stubbed elements — no browser, no portal, no DB. The HTML the
stubs mimic is the live markup captured from the account (a MUI card list on My
Network, a Bootstrap tab pane per agency portal).

    server/.venv/bin/python -m pytest server/tests/test_ridemetro_network.py
"""

import os
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook  # noqa: E402

from selenium.common.exceptions import NoSuchElementException  # noqa: E402
from selenium.webdriver.common.by import By  # noqa: E402

from app.scrapers.ridemetro import export, opportunities, workbook  # noqa: E402
from app.scrapers.ridemetro.models import SHEET_COLUMNS  # noqa: E402
from app.scrapers.ridemetro.network import (  # noqa: E402
    Agency,
    go_to_selector,
    read_agencies,
)

HEADERS_WITH_DEPARTMENT = ["Status", "Ref. #", "Project", "Department", "Close Date", "Days Left", "Action"]
HEADERS_WITHOUT_DEPARTMENT = ["Status", "Ref. #", "Project", "Close Date", "Days Left", "Action"]


# -- stubs -------------------------------------------------------------------


class FakeLink:
    def __init__(self, href):
        self._href = href

    def get_attribute(self, name):
        return self._href if name == "href" else None


class FakeCell:
    """A <td>: text lives in textContent (the pane's cells often render as
    invisible, which is exactly why the scraper never reads `.text`)."""

    def __init__(self, text="", links=()):
        self._text = text
        self._links = list(links)

    def get_attribute(self, name):
        return self._text if name == "textContent" else None

    @property
    def text(self):  # what Selenium returns for a non-displayed element
        return ""

    def find_elements(self, by, value):
        return self._links


class FakeRow:
    def __init__(self, cells):
        self._cells = cells

    def find_elements(self, by, value):
        if value == "td":
            return self._cells
        # A row-level anchor search: flatten the cells' links.
        return [link for cell in self._cells for link in cell._links]


class FakeDriver:
    def __init__(self, headers, rows):
        self._headers = [FakeCell(h) for h in headers]
        self._rows = rows

    def find_elements(self, by, value):
        if value == opportunities.SEL["header_cells"][1]:
            return self._headers
        if value == opportunities.SEL["rows"][1]:
            return self._rows
        return []


def _row(values, url=None):
    cells = [FakeCell(v) for v in values]
    cells.append(FakeCell("View Opportunity", links=[FakeLink(url)] if url else []))
    return FakeRow(cells)


# -- agency filtering --------------------------------------------------------


def test_incomplete_is_not_mistaken_for_complete():
    """The whole filter turns on this: "Incomplete" contains "complete", so a
    substring test would sweep exactly the agencies that must be skipped."""
    assert Agency("A", "Complete", "https://a.bonfirehub.com").is_complete
    assert Agency("A", "complete", "https://a.bonfirehub.com").is_complete
    assert not Agency("A", "Incomplete", "https://a.bonfirehub.com").is_complete
    assert not Agency("A", "INCOMPLETE", "https://a.bonfirehub.com").is_complete
    assert not Agency("A", "", "https://a.bonfirehub.com").is_complete


def test_an_agency_with_no_portal_is_not_reported_as_a_failure():
    """My Network lists every organisation the account has touched, and some
    publish no public portal at all. That is a fact about the agency: it reads
    as a note, not as "could not be read"."""
    agency = Agency("Bonfire", "Complete", "", agency_id="1080662-ca")
    agency.note = export.NOTE_NO_PORTAL
    record = agency.as_record()
    assert record["error"] is None
    assert record["skipped"] is False
    assert export._note_for(record) == export.NOTE_NO_PORTAL


def test_an_agency_with_nothing_open_is_a_success_not_a_failure():
    """An empty portal is a result, not a crash. The agency was reached, its
    list was read, and it held nothing — so nothing anywhere in the record marks
    it failed, and the report says so in words rather than leaving a gap."""
    agency = Agency("Ada County Highway District", "Complete",
                    "https://achdidaho.bonfirehub.com", agency_id="3042794-us")
    record = agency.as_record()

    assert record["opportunities"] == 0
    assert record["error"] is None      # not a failure
    assert record["note"] is None       # nor a portal that could never have one
    assert record["skipped"] is False   # it was visited
    assert export._note_for(record) == export.NOTE_EMPTY


def test_a_real_failure_still_outranks_a_note():
    record = Agency("A", "Complete", "https://a.bonfirehub.com").as_record()
    record["error"] = "timed out"
    record["note"] = export.NOTE_NO_PORTAL
    assert export._note_for(record).startswith("Could not be read:")


def test_agency_record_reports_the_skip():
    record = Agency("Ada County", "Incomplete", "https://adacounty.bonfirehub.com/registration").as_record()
    assert record["skipped"] is True
    assert record["name"] == "Ada County"


def test_go_to_selector_quotes_names_with_parentheses_and_quotes():
    name = 'Metropolitan Transit Authority of Harris County (METRO)'
    assert go_to_selector(name)[1] == (
        f'a[aria-label="Go to {name}"], button[aria-label="Go to {name}"]'
    )
    assert go_to_selector('The "Big" County')[1] == (
        'a[aria-label="Go to The \\"Big\\" County"], '
        'button[aria-label="Go to The \\"Big\\" County"]'
    )


def test_go_to_selector_matches_the_button_form_too():
    """Half the reason agencies went missing: the roster renders "Go to Agency"
    as a <button> for some agencies, and a locator naming only <a> finds
    neither the control nor, through it, the agency."""
    assert "button[aria-label=" in go_to_selector("Bonfire")[1]


# -- reading the roster ------------------------------------------------------


class FakeControl:
    """A "Go to Agency" control. `href` is None on the <button> form, which is
    what Selenium returns for an attribute an element does not carry."""

    def __init__(self, name, href=None):
        self._attrs = {"aria-label": f"Go to {name}", "href": href}

    def get_attribute(self, name):
        return self._attrs.get(name)


class FakeText:
    def __init__(self, text):
        self._text = text

    def get_attribute(self, name):
        return self._text if name == "textContent" else None

    @property
    def text(self):
        return self._text


class FakeCard:
    """One <li data-testid="agency-…-list-item"> of the My Network list."""

    def __init__(self, agency_id, name, status, control=None, heading=True):
        self._attrs = {"data-testid": f"agency-{agency_id}-list-item"}
        self._name = name
        self._status = status
        self._control = control
        self._heading = heading

    def get_attribute(self, name):
        return self._attrs.get(name)

    def find_element(self, by, value):
        if by == By.XPATH:  # the Status box
            return FakeText(f"Status{self._status}")
        if "aria-label^=" in value:  # the "Go to Agency" control
            if self._control is None:
                raise NoSuchElementException(value)
            return self._control
        if value == "p":  # the card heading
            if not self._heading:
                raise NoSuchElementException(value)
            return FakeText(self._name)
        if value.startswith("input"):  # the select checkbox
            return _SelectBox(self._name)
        raise NoSuchElementException(value)


class _SelectBox:
    """<input type="checkbox" aria-label="Select Bonfire"> — the name's last
    hiding place when a card has neither a control nor a heading."""

    def __init__(self, name):
        self._name = name

    def get_attribute(self, name):
        return f"Select {self._name}" if name == "aria-label" else None


class FakeRoster:
    def __init__(self, cards):
        self._cards = cards

    def find_elements(self, by, value):
        return self._cards


# The four shapes the live page actually serves, in the order it lists them.
ROSTER_CARDS = [
    # Incomplete, linked at /registration — read, then skipped by status.
    FakeCard("2974208-us", "Metropolitan Transit Authority of Harris County (METRO)",
             "Incomplete",
             FakeControl("Metropolitan Transit Authority of Harris County (METRO)",
                         "https://ridemetro.bonfirehub.com/registration")),
    # Incomplete AND button-rendered.
    FakeCard("1086816-ca", "Agriculture Financial Services Corporation", "Incomplete",
             FakeControl("Agriculture Financial Services Corporation")),
    # Complete and linked — the case that always worked.
    FakeCard("2169295-us", "Harris County", "Complete",
             FakeControl("Harris County", "https://harriscountytx.bonfirehub.com")),
    # Complete and button-rendered — the case that was silently dropped.
    FakeCard("1080662-ca", "Bonfire", "Complete", FakeControl("Bonfire")),
]


def test_button_rendered_agencies_are_in_the_roster():
    """The bug: "Go to Agency" is a <button> with no href on some cards, and a
    locator that named only <a> dropped those agencies from the sweep entirely
    rather than merely losing their URL."""
    agencies = read_agencies(FakeRoster(ROSTER_CARDS))
    assert [a.name for a in agencies] == [
        "Metropolitan Transit Authority of Harris County (METRO)",
        "Agriculture Financial Services Corporation",
        "Harris County",
        "Bonfire",
    ]
    bonfire = agencies[3]
    assert bonfire.url == ""          # no href to fall back to
    assert bonfire.is_complete        # …but it is scraped all the same
    assert bonfire.agency_id == "1080662-ca"


def test_only_incomplete_agencies_are_left_out_of_the_sweep():
    agencies = read_agencies(FakeRoster(ROSTER_CARDS))
    assert [a.name for a in agencies if a.is_complete] == ["Harris County", "Bonfire"]
    assert [a.agency_id for a in agencies if a.is_incomplete] == ["2974208-us", "1086816-ca"]


def test_a_card_with_no_control_is_still_listed():
    """It cannot be opened, but leaving it out would hide the gap from the
    Active-count reconciliation that is meant to catch exactly this."""
    agencies = read_agencies(FakeRoster([FakeCard("999-us", "Ghost County", "Complete")]))
    assert [(a.name, a.url) for a in agencies] == [("Ghost County", "")]


def test_a_card_with_no_readable_name_is_dropped():
    card = FakeCard("999-us", "", "Complete", heading=False)
    assert read_agencies(FakeRoster([card])) == []


# -- reading one agency's open list ------------------------------------------


def test_rows_are_read_by_header_not_position():
    """Only some agencies publish a Department column, so the same extractor has
    to handle both widths without shifting Close Date into it."""
    with_dept = FakeDriver(HEADERS_WITH_DEPARTMENT, [_row(
        ["Open", "IFB 2026000025", "Auxiliary Power Supply Parts", "Purchasing",
         "Aug 19th 2026, 2:00 PM CDT", "14"],
        "https://ridemetro.bonfirehub.com/opportunities/246231",
    )])
    without_dept = FakeDriver(HEADERS_WITHOUT_DEPARTMENT, [_row(
        ["Open", "BCZ26-00030328", "Advanced Metering Infrastructure",
         "Aug 7th 2026, 1:00 PM CDT", "2"],
        "https://dallascityhall.bonfirehub.com/opportunities/224829",
    )])

    a = opportunities.read_rows(with_dept)[0]
    assert a["department"] == "Purchasing"
    assert a["close_date"] == "Aug 19th 2026, 2:00 PM CDT"
    assert a["days_left"] == "14"
    assert a["opportunity_url"].endswith("/opportunities/246231")

    b = opportunities.read_rows(without_dept)[0]
    assert "department" not in b
    assert b["close_date"] == "Aug 7th 2026, 1:00 PM CDT"
    assert b["days_left"] == "2"


def test_placeholder_row_is_not_an_opportunity():
    """An agency with nothing open still renders one row of prose."""
    driver = FakeDriver(HEADERS_WITHOUT_DEPARTMENT,
                        [FakeRow([FakeCell("There are no open projects at this time.")])])
    assert opportunities.read_rows(driver) == []


def test_raw_data_keeps_the_portal_headers():
    driver = FakeDriver(HEADERS_WITH_DEPARTMENT, [_row(
        ["Open", "2026000018", "Precast Concrete Waste Receptacles", "Purchasing",
         "Aug 26th 2026, 2:00 PM CDT", "21"],
        "https://ridemetro.bonfirehub.com/opportunities/245158",
    )])
    raw = opportunities.read_rows(driver)[0]["raw_data"]
    assert raw["Ref. #"] == "2026000018"
    assert raw["Days Left"] == "21"


# -- the report --------------------------------------------------------------


ROSTER = [
    {"name": "Ada County", "url": "https://adacounty.bonfirehub.com/registration",
     "status": "Incomplete", "skipped": True, "opportunities": 0, "error": None},
    {"name": "Ada County Highway District", "url": "https://achdidaho.bonfirehub.com/",
     "status": "Complete", "skipped": False, "opportunities": 0, "error": None},
    {"name": "City of Dallas", "url": "https://dallascityhall.bonfirehub.com/",
     "status": "Complete", "skipped": False, "opportunities": 2, "error": None},
    {"name": "Houston City College", "url": "https://hccs.bonfirehub.com/",
     "status": "Complete", "skipped": False, "opportunities": 0,
     "error": "Timed out waiting for the opportunities list"},
]

RECORDS = [
    {"agency": "City of Dallas", "status": "Open", "ref_number": "BCZ26-00030328",
     "project": "Advanced Metering Infrastructure", "department": None,
     "close_date": "Aug 7th 2026, 1:00 PM CDT", "days_left": "2",
     "opportunity_url": "https://dallascityhall.bonfirehub.com/opportunities/224829"},
    {"agency": "City of Dallas", "status": "Open", "ref_number": "IFS AVI B700005",
     "project": "Purchase of Public Opinion Surveys", "department": None,
     "close_date": "Aug 10th 2026, 3:00 PM CDT", "days_left": "6",
     "opportunity_url": "https://dallascityhall.bonfirehub.com/opportunities/215486"},
]


def _report():
    tmp = TemporaryDirectory()
    out = Path(tmp.name) / "report.xlsx"
    written = export.generate_excel_from_records(RECORDS, out, ROSTER)
    return tmp, load_workbook(out).active, written


def test_report_is_grouped_by_agency_in_roster_order():
    tmp, sheet, written = _report()
    with tmp:
        assert written == 2
        banners = [
            sheet.cell(row=r, column=1).value
            for r in range(1, sheet.max_row + 1)
            if sheet.cell(row=r, column=1).font.size == 14
        ]
        assert banners == [a["name"] for a in ROSTER]


def test_each_block_is_banner_then_headers_then_rows_then_a_blank():
    tmp, sheet, _ = _report()
    with tmp:
        headers = [header for _, header in SHEET_COLUMNS]
        # Ada County: skipped, so a note stands in for its rows.
        assert sheet["A1"].value == "Ada County"
        assert [c.value for c in sheet[2]] == headers
        assert sheet["A3"].value == export.NOTE_SKIPPED
        assert all(c.value is None for c in sheet[4])
        # Ada County Highway District: visited, nothing open.
        assert sheet["A7"].value == export.NOTE_EMPTY
        # City of Dallas: two real rows under its own header row.
        assert sheet["A9"].value == "City of Dallas"
        assert [c.value for c in sheet[10]] == headers
        assert sheet["B11"].value == "BCZ26-00030328"
        assert sheet["B12"].value == "IFS AVI B700005"
        assert all(c.value is None for c in sheet[13])
        # Houston City College: reached, but the read failed — said so, not dropped.
        assert sheet["A14"].value == "Houston City College"
        assert sheet["A16"].value.startswith("Could not be read:")


def test_banner_spans_every_data_column():
    tmp, sheet, _ = _report()
    with tmp:
        merged = {str(r) for r in sheet.merged_cells.ranges}
        last = chr(ord("A") + len(SHEET_COLUMNS) - 1)
        assert f"A1:{last}1" in merged


def test_bid_urls_are_clickable():
    tmp, sheet, _ = _report()
    with tmp:
        assert sheet["G11"].hyperlink.target.endswith("/opportunities/224829")


def test_an_agency_with_rows_but_no_roster_entry_is_still_reported():
    """A roster read that came up short must not silently drop scraped rows."""
    with TemporaryDirectory() as tmp:
        out = Path(tmp) / "report.xlsx"
        assert export.generate_excel_from_records(RECORDS, out, []) == 2
        sheet = load_workbook(out).active
        assert sheet["A1"].value == "City of Dallas"


def test_records_without_a_reference_or_url_are_not_written():
    with TemporaryDirectory() as tmp:
        out = Path(tmp) / "report.xlsx"
        junk = [{"agency": "City of Dallas", "status": "Open", "project": "no id at all"}]
        assert export.generate_excel_from_records(junk, out, ROSTER[:3]) == 0


def test_illegal_characters_do_not_break_the_workbook():
    with TemporaryDirectory() as tmp:
        out = Path(tmp) / "report.xlsx"
        rows = [{"agency": "A", "ref_number": "R1", "project": "bell\x07char",
                 "opportunity_url": "https://a.bonfirehub.com/opportunities/1"}]
        workbook.build([("A", rows)], out)
        assert load_workbook(out).active["C3"].value == "bellchar"
