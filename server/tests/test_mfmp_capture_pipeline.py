"""MFMP after the matrix: human OTP, every bid kept, one organised archive.

Three changes, three things worth holding:

* **Login waits for a person.** The portal answers a correct password with a
  one-time code. Leaving `/login` is not being signed in — the OTP challenge is
  its own route — so the old check handed a half-authenticated session to the
  search step, which failed later with a message about a missing button.
* **Nothing is scored, categorised or dropped.** Every advertisement the portal
  returns reaches the summary sheet with its documents beside it. What is worth
  pursuing is the reviewer's call, and the scraper no longer pre-empts it.
* **The archive has a shape.** `MyFlorida_Export/` with the summary at its root
  and one folder per bid under `Bids_Data/`, so a row in the sheet names a
  folder the reader can actually open.

    server/.venv/bin/python -m pytest server/tests/test_mfmp_capture_pipeline.py
"""

import logging
import os
import sys
import zipfile
from pathlib import Path

import pytest
from selenium.common.exceptions import TimeoutException

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.core import run_manager  # noqa: E402
from app.scrapers.myflorida import storage, workbook  # noqa: E402
from app.scrapers.myflorida import scraper as mfmp  # noqa: E402
from app.scrapers.myflorida.scraper import LoginTimeout, MFMPScraper  # noqa: E402


# =============================================================================
# The OTP window
# =============================================================================


class FakePortal:
    """A browser that reaches the dashboard after `steps` polls, if ever."""

    def __init__(self, steps: int, url: str = "https://vendor.myfloridamarketplace.com/otp"):
        self.polls = 0
        self.steps = steps
        self.current_url = url

    def find_elements(self, _by, _value):
        self.polls += 1
        if self.polls >= self.steps:
            self.current_url = "https://vendor.myfloridamarketplace.com/vendor/ads"
            return [object()]
        return []


def _scraper(tmp_path, monkeypatch, driver):
    run = run_manager.create_run("myflorida", tmp_path, {"category": "x"})
    instance = MFMPScraper(run["run_id"], codes=["10-11"])
    instance.driver = driver
    monkeypatch.setattr(instance, "screenshot", lambda name: None)
    return instance


class ImmediateWait:
    """Polls the condition without sleeping, then times out like Selenium's."""

    def __init__(self, driver, timeout, attempts=5):
        self.driver, self.timeout, self.attempts = driver, timeout, attempts

    def until(self, condition):
        for _ in range(self.attempts):
            if condition(self.driver):
                return True
        raise TimeoutException(f"gave up after {self.timeout}s")


def test_the_run_waits_for_the_one_time_password(tmp_path, monkeypatch, caplog):
    caplog.set_level(logging.INFO)
    monkeypatch.setattr(mfmp.settings, "mfmp_manual_otp", True)
    monkeypatch.setattr(mfmp.settings, "mfmp_otp_wait_seconds", 120)
    driver = FakePortal(steps=3)
    instance = _scraper(tmp_path, monkeypatch, driver)
    seen: list[int] = []
    monkeypatch.setattr(instance, "wait", lambda t=None: (seen.append(t), ImmediateWait(driver, t))[1])

    instance._await_authenticated()

    assert seen == [120], "the OTP window, not the default element wait"
    run = run_manager.get_run(instance.run_id)
    assert run["awaiting_otp"] is False, "the flag is cleared once through"
    assert any("one-time password" in w for w in run["warnings"]), (
        "the prompt must reach the run — nobody is reading the server's stdout"
    )


def test_leaving_the_login_page_is_not_being_signed_in(tmp_path, monkeypatch):
    """The OTP challenge is its own route, so a URL check alone passes the moment
    the code is *demanded* — and hands a half-authenticated session onwards."""
    driver = FakePortal(steps=99, url="https://vendor.myfloridamarketplace.com/otp")
    instance = _scraper(tmp_path, monkeypatch, driver)

    assert instance._authenticated(driver) is False, "off /login, but not inside"


def test_the_dashboard_is_what_counts_as_signed_in(tmp_path, monkeypatch):
    driver = FakePortal(steps=1, url="https://vendor.myfloridamarketplace.com/vendor/ads")
    instance = _scraper(tmp_path, monkeypatch, driver)

    assert instance._authenticated(driver) is True


def test_a_code_that_never_arrives_fails_with_something_actionable(tmp_path, monkeypatch):
    monkeypatch.setattr(mfmp.settings, "mfmp_manual_otp", True)
    monkeypatch.setattr(mfmp.settings, "mfmp_otp_wait_seconds", 90)
    driver = FakePortal(steps=999)
    instance = _scraper(tmp_path, monkeypatch, driver)
    monkeypatch.setattr(instance, "wait", lambda t=None: ImmediateWait(driver, t))

    with pytest.raises(LoginTimeout) as excinfo:
        instance._await_authenticated()

    message = str(excinfo.value)
    assert "90s" in message
    assert "Chrome window" in message, "say what a human should have done"
    assert run_manager.get_run(instance.run_id)["awaiting_otp"] is False


def test_with_manual_otp_off_the_wait_is_the_ordinary_one(tmp_path, monkeypatch):
    """An account without OTP should not sit for two minutes on every login."""
    monkeypatch.setattr(mfmp.settings, "mfmp_manual_otp", False)
    driver = FakePortal(steps=1)
    instance = _scraper(tmp_path, monkeypatch, driver)
    seen: list[int] = []
    monkeypatch.setattr(instance, "wait", lambda t=None: (seen.append(t), ImmediateWait(driver, t))[1])

    instance._await_authenticated()

    assert seen == [mfmp.LOGIN_FORM_TIMEOUT]
    assert "one-time password" not in " ".join(
        run_manager.get_run(instance.run_id)["warnings"]
    )


def test_the_browser_is_visible_while_manual_otp_is_on(tmp_path, monkeypatch):
    """There is nothing to type a code into in a headless window, and the run's
    own Live-preview flag must not decide it."""
    monkeypatch.setattr(mfmp.settings, "mfmp_manual_otp", True)
    run = run_manager.create_run("myflorida", tmp_path, {})
    instance = MFMPScraper(run["run_id"], codes=[])
    calls: list[dict] = []
    monkeypatch.setattr(instance, "start_driver", lambda **kw: calls.append(kw))
    monkeypatch.setattr(instance, "login", _stop)
    monkeypatch.setattr(instance, "cleanup", lambda: None)
    monkeypatch.setattr(mfmp.run_manager, "remove_empty_folder", lambda run_id: None)

    instance.run()

    assert calls == [{"headless": False}]


def _stop(*_args, **_kwargs):
    from app.core.base_scraper import StopRequested

    raise StopRequested("far enough — the driver call is what this test is about")


# =============================================================================
# Nothing is dropped
# =============================================================================


def test_the_close_date_filter_is_gone_from_the_pipeline():
    """It pruned the workbook to ads at least N days from closing, which removed
    bids a reviewer might still have wanted. Judging what is worth pursuing is
    the reviewer's job now, so the call and the function are both gone."""
    from app.scrapers.myflorida import ingest

    assert not hasattr(ingest, "filter_workbook_by_close_date")
    source = Path(mfmp.__file__).read_text()
    assert "from app.core.closing_filter import" not in source, (
        "the flow should no longer import the close-date rule at all"
    )


def test_the_sweep_no_longer_imports_the_classifier():
    """The matrix is out of the pipeline. The modules still exist — they serve
    the lane views of runs recorded before this change — but nothing in the
    scraping path reaches for them."""
    source = Path(
        Path(mfmp.__file__).parent / "sweep" / "scraper.py"
    ).read_text()

    for banned in ("from app.scrapers.myflorida.sweep.routing import",
                   "from app.scrapers.myflorida.sweep.scoring import",
                   "classify(", "get_config()"):
        assert banned not in source, f"the sweep still reaches for {banned!r}"


def test_the_sweep_delivers_a_zip_not_a_bare_sheet():
    """It kept nothing but a workbook while its documents were deleted after
    scoring. It keeps them now, so the deliverable is the archive."""
    from app.core import exports

    assert "myflorida_sweep" not in exports.EXCEL_ONLY_PORTALS
    assert "myflorida_sweep" in exports.DOC_PORTALS


# =============================================================================
# The archive's shape
# =============================================================================


def test_the_layout_is_one_root_with_the_summary_at_its_top(tmp_path):
    root = storage.export_root(tmp_path)

    assert root.name == "MyFlorida_Export"
    assert storage.summary_path(tmp_path).name == "MyFlorida_Bids_Summary.xlsx"
    assert storage.summary_path(tmp_path).parent == root


def test_each_bid_gets_a_folder_named_for_its_ad_number(tmp_path):
    folder = storage.bid_folder(tmp_path, "DMS-21/22-001", "Statewide signage services")

    assert folder.is_dir()
    assert folder.parent.name == "Bids_Data"
    assert "DMS" in folder.name and "signage" not in folder.name, (
        "the ad number addresses the folder; the title is in the sheet"
    )


def test_a_bid_with_no_ad_number_still_gets_somewhere_to_live(tmp_path):
    folder = storage.bid_folder(tmp_path, "", "An advertisement with no number")

    assert folder.is_dir()
    assert folder.name


def test_the_folder_column_points_inside_the_archive():
    """An absolute server path would be useless to whoever unpacks the ZIP."""
    reference = storage.folder_reference("RFP-2026-114", "Anything")

    assert reference == "Bids_Data/RFP-2026-114"
    assert not reference.startswith("/")


def test_the_summary_holds_every_captured_bid(tmp_path):
    from openpyxl import load_workbook

    records = [
        {"ad_number": f"AD-{n}", "title": f"Bid {n}", "agency": "DMS",
         "close_date": "2026-09-01", "documents": ["a.pdf", "b.pdf"],
         "folder": f"Bids_Data/AD-{n}"}
        for n in range(5)
    ]
    path = workbook.build_from_records(records, tmp_path)
    sheet = load_workbook(path).active
    rows = list(sheet.iter_rows(values_only=True))

    assert path == storage.summary_path(tmp_path)
    assert len(rows) == 6, "a header and five bids — none filtered out"
    headers = [str(h) for h in rows[0]]
    for expected in ("Title", "Advertisement Number", "Agency", "Closing Date", "Documents"):
        assert expected in headers
    assert rows[1][headers.index("Documents")] == 2


def test_the_summary_columns_are_the_same_for_every_run(tmp_path):
    """The sheet used to be the portal's own export passed through, so its shape
    depended on the search. A reviewer comparing two runs is comparing one
    spreadsheet now, and an ad that rendered nothing still fills every column."""
    from openpyxl import load_workbook

    expected = [header for _, header in workbook.RECORD_COLUMNS]

    full = workbook.build_from_records(
        [{key: "x" for key, _ in workbook.RECORD_COLUMNS}], tmp_path / "full"
    )
    bare = workbook.build_from_records([{"ad_number": "AD-1"}], tmp_path / "bare")

    for path in (full, bare):
        headers = next(load_workbook(path).active.iter_rows(values_only=True))
        assert [str(h) for h in headers] == expected
    assert len(expected) == 17


def test_a_bid_whose_detail_page_failed_still_reaches_the_sheet(tmp_path):
    """Its status and commodity codes are gone, but the row names it, dates it
    and says who posted it. A row that silently vanished would be worse."""
    from openpyxl import load_workbook

    path = workbook.build_from_records(
        [{"ad_number": "AD-9", "title": "Grid only", "agency": "DMS",
          "open_date": "08/26/2026", "close_date": "08/29/2026"}],
        tmp_path,
    )
    sheet = load_workbook(path).active
    header, row = list(sheet.iter_rows(values_only=True))[:2]
    cells = dict(zip((str(h) for h in header), row))

    assert cells["Advertisement Number"] == "AD-9"
    assert cells["Closing Date"] == "08/29/2026"
    assert cells["Status"] is None
    assert cells["Documents"] == 0


def test_the_zip_is_the_export_root_with_documents_in_place(tmp_path, monkeypatch):
    """End to end on the packaging: what a reviewer downloads unpacks to one
    folder, with the sheet at the top and each bid's files under its own name."""
    from app.core import exports

    run_dir = tmp_path / "run"
    run_dir.mkdir()
    storage.summary_path(run_dir).write_text("summary")
    (storage.bid_folder(run_dir, "AD-1") / "Solicitation.pdf").write_text("one")
    (storage.bid_folder(run_dir, "AD-2") / "Specs.docx").write_text("two")
    # Staging that must never reach the archive.
    (run_dir / "_exports").mkdir()
    (run_dir / "_exports" / "page1.xlsx").write_text("raw")
    (run_dir / "_downloads").mkdir()
    (run_dir / "_downloads" / "part.crdownload").write_text("mid-flight")

    run = {
        "run_id": "abc123", "scraper": "myflorida_sweep", "folder": str(run_dir),
        "excel_path": str(storage.summary_path(run_dir)), "search": "open",
    }
    monkeypatch.setattr(exports, "excel_bytes", lambda r: (b"regenerated", "MyFlorida_Sweep_(open).xlsx"))
    out = tmp_path / "out.zip"
    exports.build_zip(run, out)

    with zipfile.ZipFile(out) as zf:
        names = sorted(zf.namelist())

    assert names == [
        "MyFlorida_Export/Bids_Data/AD-1/Solicitation.pdf",
        "MyFlorida_Export/Bids_Data/AD-2/Specs.docx",
        "MyFlorida_Export/MyFlorida_Bids_Summary.xlsx",
    ], names


def test_the_summary_is_not_shipped_twice(tmp_path, monkeypatch):
    """The regenerated sheet comes back under a different name than the copy in
    the tree, so a name check alone would add a second one at the ZIP root."""
    from app.core import exports

    run_dir = tmp_path / "run"
    run_dir.mkdir()
    storage.summary_path(run_dir).write_bytes(b"the sheet")
    run = {
        "run_id": "abc123", "scraper": "myflorida", "folder": str(run_dir),
        "excel_path": str(storage.summary_path(run_dir)),
    }
    monkeypatch.setattr(exports, "excel_bytes", lambda r: (b"same rows", "MyFlorida_(x).xlsx"))
    out = tmp_path / "out.zip"
    exports.build_zip(run, out)

    with zipfile.ZipFile(out) as zf:
        assert zf.namelist() == ["MyFlorida_Export/MyFlorida_Bids_Summary.xlsx"]


# =============================================================================
# The results grid and the detail page, as the summary sheet's two sources
# =============================================================================


class FakeCell:
    def __init__(self, text, link=None):
        self.text = text
        self._link = link

    def find_elements(self, _by, _value):
        return [FakeCell(self._link)] if self._link is not None else []


class FakeRow:
    def __init__(self, cells):
        self._cells = cells

    def find_elements(self, _by, _value):
        return self._cells


class FakeGrid:
    """A driver whose results table is the portal's eight columns."""

    def __init__(self, rows):
        self.rows = rows

    def find_elements(self, _by, _value):
        return self.rows


def _grid_row(number="AD-16589"):
    """One results row, in the portal's column order."""
    return FakeRow([
        FakeCell("Intent to Award Single Source provider to Cochlear Americas"),
        FakeCell(f"  {number} ", link=number),
        FakeCell("IA-27-038"),
        FakeCell("1"),
        FakeCell("Florida School for the Deaf and the Blind (FSDB)"),
        FakeCell("Agency Decision"),
        FakeCell("08/26/2026 01:30 AM"),
        FakeCell("08/29/2026 01:30 AM"),
    ])


def test_the_grid_gives_the_row_everything_it_alone_carries(tmp_path, monkeypatch):
    """The posting window has no other source — it is on the grid and not on the
    detail page's topBar under those names, so the sheet's Open and Closing
    dates are only ever as good as this read."""
    instance = _scraper(tmp_path, monkeypatch, FakeGrid([_grid_row()]))

    bid = instance.collect_bids()[0]

    assert bid["ad_number"] == "AD-16589"
    assert bid["agency_ad_number"] == "IA-27-038"
    assert bid["version"] == "1"
    assert bid["agency"] == "Florida School for the Deaf and the Blind (FSDB)"
    assert bid["ad_type"] == "Agency Decision"
    assert bid["open_date"] == "08/26/2026 01:30 AM"
    assert bid["close_date"] == "08/29/2026 01:30 AM"


def test_the_ad_number_is_the_link_text_not_the_cell_text(tmp_path, monkeypatch):
    """It is the dedup key, the documents folder name and the database's unique
    constraint. The cell's rendered text has padding the link's does not, and a
    key that differs by a space forks every one of those."""
    instance = _scraper(tmp_path, monkeypatch, FakeGrid([_grid_row()]))

    bid = instance.collect_bids()[0]

    assert bid["ad_number"] == "AD-16589"
    assert bid["number"] == "AD-16589"


def test_a_row_with_no_link_is_not_collected(tmp_path, monkeypatch):
    """Without an ad number there is nothing to open, nowhere to put its files
    and no key to record it under."""
    row = FakeRow([FakeCell("A title"), FakeCell("AD-1")])  # no anchor
    instance = _scraper(tmp_path, monkeypatch, FakeGrid([row]))

    assert instance.collect_bids() == []


def test_the_detail_page_fills_what_the_grid_cannot(tmp_path, monkeypatch):
    """Status, commodity codes and the contact exist on no grid column."""
    instance = _scraper(tmp_path, monkeypatch, FakeGrid([]))
    parsed = {"status": "OPEN", "commodity_codes": "42211705 — Hearing aid",
              "contact_name": "Kim Whitwam", "published_date": "08/05/2026 12:32 AM"}

    row = instance._summary_row({"ad_number": "AD-16589", "title": "T"}, parsed, ["a.pdf"])

    assert row["status"] == "OPEN"
    assert row["commodity_codes"] == "42211705 — Hearing aid"
    assert row["contact_name"] == "Kim Whitwam"
    assert row["documents"] == ["a.pdf"]


def test_the_grid_wins_on_identity_and_the_posting_window(tmp_path, monkeypatch):
    """The detail page's topBar carries Start/End dates of its own, and its
    Advertisement Number is a second reading of the run's key. Both defer to the
    grid: the key has to be the string everything else is filed under, and the
    two dates have to mean the same thing on every row of the sheet."""
    instance = _scraper(tmp_path, monkeypatch, FakeGrid([]))
    bid = {"ad_number": "AD-16589", "open_date": "08/26/2026", "close_date": "08/29/2026"}
    parsed = {"ad_number": "AD-99999", "open_date": "elsewhere", "close_date": "elsewhere"}

    row = instance._summary_row(bid, parsed, [])

    assert row["ad_number"] == "AD-16589"
    assert row["open_date"] == "08/26/2026"
    assert row["close_date"] == "08/29/2026"


def test_a_detail_page_that_will_not_parse_does_not_fail_the_bid(tmp_path, monkeypatch):
    """A hundred-bid run must not end because one page rendered badly."""
    class Broken:
        @property
        def page_source(self):
            raise RuntimeError("no page")

        current_url = "https://x/detail/1"

    instance = _scraper(tmp_path, monkeypatch, Broken())

    assert instance.read_detail() == {}
    run = run_manager.get_run(instance.run_id)
    assert any("detail page" in w for w in run["warnings"])


def test_a_bid_never_opened_still_reaches_the_sheet(tmp_path, monkeypatch):
    """Stop landed, or the detail route timed out twice. The grid row is what is
    left and it still names the ad, dates it and says who posted it."""
    instance = _scraper(tmp_path, monkeypatch, FakeGrid([]))
    found = {"AD-1": {"ad_number": "AD-1", "title": "Opened", "agency": "DMS"},
             "AD-2": {"ad_number": "AD-2", "title": "Never opened", "agency": "FSDB"}}
    instance._records["AD-1"] = {"ad_number": "AD-1", "title": "Opened", "status": "OPEN"}

    records = instance._summary_records(found)

    assert [r["ad_number"] for r in records] == ["AD-1", "AD-2"]
    assert records[0]["status"] == "OPEN"
    assert records[1]["agency"] == "FSDB", "the grid's fields, with no status"
    assert not records[1].get("status")
