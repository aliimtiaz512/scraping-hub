"""City of Philadelphia (PHLContracts): the shape of what a run produces.

The portal is Periscope's BSO platform — 2000s table markup with no ids on rows
and no data attributes — so the parsing is positional-after-naming: find the
header cells by their labels, then read the cells beneath them. These tests run
that logic against the portal's real markup (the snippets the scraper was built
from) rather than against a convenient shape, because a fake that is tidier than
the page it stands for is how a parser passes its tests and fails its portal.

    server/.venv/bin/python -m pytest server/tests/test_philadelphia.py
"""

import os
import sys
import zipfile
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.scrapers.philadelphia import storage  # noqa: E402
from app.scrapers.philadelphia.models import EXCEL_COLUMNS, PhiladelphiaBid  # noqa: E402
from app.scrapers.philadelphia.scraper import PhiladelphiaScraper  # noqa: E402


# =============================================================================
# The Open Bids table, as the portal renders it
# =============================================================================

# Two rows of the real dashboard markup: the section heading, the results table
# with its nine columns, and the "View More…" row that closes it. Rows carry no
# id — the bid link is the only handle — and the blank Alternate Id cell is the
# portal's, not a simplification.
OPEN_BIDS_HTML = """
<table><tr><td class="sectionheader-02">Bids / Bid Amendments</td></tr></table>
<table id="resultsTable"><tbody>
  <tr><th>Bid #</th><th>Organization</th><th>Description</th></tr>
  <tr><td><a href="/bso/seller/bidAck.sda?bidId=B999">B999</a></td>
      <td>Somewhere Else</td><td>A different section entirely</td></tr>
</tbody></table>

<table><tr><td class="sectionheader-02">Open Bids</td></tr>
<tr><td>
  <table id="resultsTable" name="resultsTable" class="table-02">
    <tbody>
      <tr>
        <th scope="col" class="listheading">Bid #</th>
        <th scope="col" class="listheading">Organization</th>
        <th scope="col" class="listheading">Alternate Id</th>
        <th scope="col" class="listheading">Buyer</th>
        <th scope="col" class="listheading">Description</th>
        <th scope="col" class="listheading">Bid Opening Date</th>
        <th scope="col" class="listheading">Bid Q &amp; A</th>
        <th scope="col" class="listheading">Create Quote<br>/View Auction</th>
        <th scope="col" class="listheading">Bid Holder</th>
      </tr>
      <tr class="tableStripe-01">
        <td class="tableText-01" align="center">
          <a href="/bso/seller/bidAck.sda?status=Open&amp;destination=detail&amp;bidId=B2727750"
             class="link-01">B2727750</a></td>
        <td class="tableText-01">City of Philadelphia</td>
        <td class="tableText-01"></td>
        <td class="tableText-01">Purchaser MP</td>
        <td class="tableText-01">Jackhammers</td>
        <td class="tableText-01">08/20/2026 11:59:00 PM</td>
        <td class="tableText-01"><a href="/bso/bid/vendorQandA.sda?docId=B2727750">View</a></td>
        <td class="tableText-01"><a href="javascript:createQuote2('B2727750','V1');">Create Quote</a></td>
        <td class="tableText-01"></td>
      </tr>
      <tr class="tableStripe-02">
        <td class="tableText-01" align="center">
          <a href="/bso/seller/bidAck.sda?status=Open&amp;destination=detail&amp;bidId=B2727732"
             class="link-01">B2727732</a></td>
        <td class="tableText-01">City of Philadelphia</td>
        <td class="tableText-01">Theresa.baker@phila.gov</td>
        <td class="tableText-01">Purchaser MP</td>
        <td class="tableText-01">License Tag Stickers</td>
        <td class="tableText-01">08/24/2026 11:59:59 PM</td>
        <td class="tableText-01"><a href="/bso/bid/vendorQandA.sda?docId=B2727732">View</a></td>
        <td class="tableText-01"></td>
        <td class="tableText-01"></td>
      </tr>
      <tr class="tableStripe-02">
        <td class="inputs-01" colspan="10" align="center">
          <a href="/bso/bid/bidList.sda?status=Open&amp;category=all" class="link-01">View More...</a>
        </td>
      </tr>
    </tbody>
  </table>
</td></tr></table>
"""

# The detail page's Header Information: label cells (`t-head-01`) each followed
# by their value, several pairs to a row, with the portal's own &nbsp; padding.
HEADER_HTML = """
<table>
  <tr><td class="sectionHeader-02" colspan="6"><h2>Header Information</h2></td></tr>
  <tr class="tableStripe-01">
    <td class="t-head-01">Bid Number:</td><td class="tableText-01">B2727750</td>
    <td class="t-head-01">Description:</td><td class="tableText-01">Jackhammers</td>
    <td class="t-head-01">Bid Opening Date:</td><td class="tableText-01">08/20/2026 11:59:00 PM</td>
  </tr>
  <tr>
    <td class="t-head-01">Purchaser:</td><td class="tableText-01">Purchaser MP</td>
    <td class="t-head-01">Organization:</td><td class="tableText-01">City of Philadelphia</td>
  </tr>
  <tr>
    <td class="t-head-01">Department:</td><td class="tableText-01">28 - Water Department</td>
    <td class="t-head-01">Location:</td><td class="tableText-01">1247 - Facilities</td>
  </tr>
  <tr>
    <td class="t-head-01">Alternate Id:</td><td class="tableText-01">&nbsp;</td>
    <td class="t-head-01">Required Date:</td><td class="tableText-01">09/01/2026</td>
  </tr>
  <tr>
    <td class="t-head-01">Info Contact:</td><td class="tableText-01">Michael Perce 215-459-4753</td>
    <td class="t-head-01">Bid Type:</td><td class="tableText-01">OPEN</td>
  </tr>
  <tr>
    <td class="t-head-01">Ship-to Address:</td>
    <td class="tableText-01">A. Candelora<br>29th &amp; Cambria Streets<br>Philadelphia, PA 19132</td>
  </tr>
  <tr>
    <td valign="top" class="t-head-01">File Attachments:</td>
    <td colspan="5">
      <a href="javascript:editFile('391619');" class="link-01">MP_Terms_and_Conditions_B2727750.pdf</a><br>
      <a href="javascript:editFile('391620');" class="link-01">Consent_Authorization_B2727750.pdf</a>
    </td>
  </tr>
</table>
"""


@pytest.fixture(scope="module")
def browser():
    """A real browser, because these are DOM-walking scripts.

    Skipped rather than faked when Chrome is unavailable: a hand-rolled stand-in
    for `querySelectorAll` would be testing the stand-in.
    """
    selenium = pytest.importorskip("selenium")
    from selenium import webdriver
    from selenium.common.exceptions import WebDriverException

    options = webdriver.ChromeOptions()
    for flag in ("--headless=new", "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"):
        options.add_argument(flag)
    try:
        driver = webdriver.Chrome(options=options)
    except WebDriverException as exc:  # pragma: no cover - environment dependent
        pytest.skip(f"Chrome unavailable: {exc.__class__.__name__}")
    yield driver
    driver.quit()


def _render(browser, html: str):
    from urllib.parse import quote

    browser.get(
        "data:text/html;charset=utf-8,"
        + quote(f"<html><body>{html}</body></html>")
    )
    return browser


def _scraper_on(browser, tmp_path):
    from app.core import run_manager

    run = run_manager.create_run("philadelphia", tmp_path)
    scraper = PhiladelphiaScraper(run["run_id"])
    scraper.driver = browser
    return scraper


def test_the_open_bids_table_is_read_not_the_first_table_on_the_page(browser, tmp_path):
    """The dashboard renders four `table#resultsTable`s — Request for Revision,
    Bids / Bid Amendments, Open Bids, Closed Bids. Taking the first one would
    scrape a different section every time the portal reorders them."""
    scraper = _scraper_on(_render(browser, OPEN_BIDS_HTML), tmp_path)
    rows = scraper.collect_rows()

    assert [r["bid_number"] for r in rows] == ["B2727750", "B2727732"]
    assert not any(r["bid_number"] == "B999" for r in rows), "read the wrong section"


def test_every_summary_column_is_captured(browser, tmp_path):
    scraper = _scraper_on(_render(browser, OPEN_BIDS_HTML), tmp_path)
    first = scraper.collect_rows()[0]

    assert first["organization"] == "City of Philadelphia"
    assert first["buyer"] == "Purchaser MP"
    assert first["description"] == "Jackhammers"
    assert first["bid_opening_date"] == "08/20/2026 11:59:00 PM"
    assert first["alternate_id"] == "", "the portal leaves this blank on most bids"
    assert first["detail_url"].endswith("bidId=B2727750")


def test_a_populated_alternate_id_lands_in_its_own_column(browser, tmp_path):
    """Columns are located by header name and then read positionally, so a value
    appearing in only some rows must not shift the ones after it."""
    scraper = _scraper_on(_render(browser, OPEN_BIDS_HTML), tmp_path)
    second = scraper.collect_rows()[1]

    assert second["alternate_id"] == "Theresa.baker@phila.gov"
    assert second["description"] == "License Tag Stickers"


def test_the_view_more_row_is_not_a_bid(browser, tmp_path):
    """It sits in the same table with a colspan and a link — and no bid id."""
    scraper = _scraper_on(_render(browser, OPEN_BIDS_HTML), tmp_path)

    assert all("View More" not in r["description"] for r in scraper.collect_rows())


def test_a_page_without_the_table_reads_as_empty_not_as_broken(browser, tmp_path):
    scraper = _scraper_on(_render(browser, "<p>No open bids at this time.</p>"), tmp_path)

    assert scraper.collect_rows() == []


# =============================================================================
# The detail page's Header Information
# =============================================================================


def test_every_header_pair_is_captured_under_its_published_label(browser, tmp_path):
    """The labels differ per bid type, so they are kept as published rather than
    mapped onto columns that would be empty for half the portal."""
    scraper = _scraper_on(_render(browser, HEADER_HTML), tmp_path)
    header = scraper.driver.execute_script(scraper._JS_HEADER)

    assert header["Bid Number"] == "B2727750"
    assert header["Department"] == "28 - Water Department"
    assert header["Location"] == "1247 - Facilities"
    assert header["Info Contact"] == "Michael Perce 215-459-4753"
    assert header["Bid Type"] == "OPEN"
    assert header["Required Date"] == "09/01/2026"


def test_a_multi_line_address_keeps_its_lines(browser, tmp_path):
    scraper = _scraper_on(_render(browser, HEADER_HTML), tmp_path)
    header = scraper.driver.execute_script(scraper._JS_HEADER)

    assert "Cambria" in header["Ship-to Address"]
    assert "Philadelphia" in header["Ship-to Address"]


def test_a_label_whose_value_is_only_a_nbsp_is_not_recorded_as_text(browser, tmp_path):
    """`&nbsp;` is the portal's way of writing "empty" — storing it as a value
    would put a non-breaking space in the record where nothing was published."""
    scraper = _scraper_on(_render(browser, HEADER_HTML), tmp_path)
    header = scraper.driver.execute_script(scraper._JS_HEADER)

    assert header.get("Alternate Id", "") == ""


def test_the_attachment_links_are_found_by_their_javascript_href(browser, tmp_path):
    """They carry no URL — `javascript:editFile('391619')` — which is why they
    are clicked rather than fetched."""
    from selenium.webdriver.common.by import By

    scraper = _scraper_on(_render(browser, HEADER_HTML), tmp_path)
    links = scraper.driver.find_elements(By.XPATH, "//a[contains(@href, 'editFile')]")

    assert [(link.text or "").strip() for link in links] == [
        "MP_Terms_and_Conditions_B2727750.pdf",
        "Consent_Authorization_B2727750.pdf",
    ]


# =============================================================================
# Storage layout and packaging
# =============================================================================


def test_the_layout_is_one_root_with_the_summary_at_its_top(tmp_path):
    root = storage.export_root(tmp_path)

    assert root.name == "CityOfPhiladelphia_Export"
    assert storage.summary_path(tmp_path).name == "Philadelphia_Bids_Summary.xlsx"
    assert storage.summary_path(tmp_path).parent == root


def test_each_bid_gets_a_folder_named_for_its_bid_number(tmp_path):
    folder = storage.bid_folder(tmp_path, "B2727750")

    assert folder.is_dir()
    assert folder.parent.name == "Bids_Data"
    assert folder.name == "B2727750"


def test_the_folder_column_points_inside_the_archive():
    assert storage.folder_reference("B2727750") == "Bids_Data/B2727750"
    assert not storage.folder_reference("B2727750").startswith("/")


def test_the_summary_carries_every_bid_and_names_its_folder(tmp_path):
    from openpyxl import load_workbook

    from app.scrapers.philadelphia import export

    records = [
        {"bid_number": f"B{n}", "description": f"Bid {n}", "organization": "City of Philadelphia",
         "buyer": "Purchaser MP", "bid_opening_date": "08/20/2026", "documents_downloaded": 2}
        for n in range(4)
    ]
    path = storage.summary_path(tmp_path)
    export.generate_excel_from_records(records, path)
    sheet = load_workbook(path).active
    rows = list(sheet.iter_rows(values_only=True))

    assert len(rows) == 5, "a header and four bids — none filtered out"
    headers = [str(h) for h in rows[0]]
    assert headers == [header for _, header in EXCEL_COLUMNS]
    assert rows[1][headers.index("Folder")] == "Bids_Data/B0"
    assert rows[1][headers.index("Total Document Count")] == 2


def test_the_zip_is_the_export_root_with_each_bids_files_in_place(tmp_path, monkeypatch):
    """What a reviewer downloads unpacks to one folder: the sheet at the top and
    each bid's documents under its own number."""
    from app.core import exports

    run_dir = tmp_path / "run"
    run_dir.mkdir()
    storage.summary_path(run_dir).write_text("summary")
    (storage.bid_folder(run_dir, "B2727750") / "Terms.pdf").write_text("one")
    storage.items_path(run_dir, "B2727750").write_text("BID ITEM SPECIFICATIONS")
    (storage.bid_folder(run_dir, "B2727732") / "Stickers.pdf").write_text("two")

    run = {
        "run_id": "abc123", "scraper": "philadelphia", "folder": str(run_dir),
        "excel_path": str(storage.summary_path(run_dir)), "search": "all open bids",
    }
    monkeypatch.setattr(exports, "excel_bytes", lambda r: (b"regenerated", "Philadelphia_(x).xlsx"))
    out = tmp_path / "out.zip"
    exports.build_zip(run, out)

    with zipfile.ZipFile(out) as zf:
        names = sorted(zf.namelist())

    assert names == [
        "CityOfPhiladelphia_Export/Bids_Data/B2727732/Stickers.pdf",
        "CityOfPhiladelphia_Export/Bids_Data/B2727750/Terms.pdf",
        "CityOfPhiladelphia_Export/Bids_Data/B2727750/bid_items_details.txt",
        "CityOfPhiladelphia_Export/Philadelphia_Bids_Summary.xlsx",
    ], names
    assert not any(name.endswith(".json") for name in names), (
        "the archive must contain nothing a non-technical reader cannot open"
    )


def test_the_portal_delivers_a_zip_not_a_bare_sheet():
    from app.core import exports

    assert "philadelphia" in exports.DOC_PORTALS
    assert "philadelphia" not in exports.EXCEL_ONLY_PORTALS


# =============================================================================
# The table contract
# =============================================================================


def test_a_bid_is_keyed_on_its_number_so_a_rerun_updates_it():
    """The Open Bids list is a live set — the same bid is in it every day until
    it closes. One row per bid, not one per sighting."""
    assert PhiladelphiaBid.__tablename__ == "city_of_philadelphia_bids"
    assert PhiladelphiaBid.__table__.primary_key.columns.keys() == ["bid_number"]


def test_every_column_the_brief_named_exists():
    columns = PhiladelphiaBid.__table__.columns
    for name in (
        "bid_number", "organization", "alternate_id", "buyer", "description",
        "bid_opening_date", "extra_header_data", "file_paths", "scraped_at",
    ):
        assert name in columns, f"{name} is missing from city_of_philadelphia_bids"


def test_clearing_a_run_does_not_delete_bids_that_are_still_open():
    """run_id is ON DELETE SET NULL — an old run's history going away must not
    take the current bids with it."""
    fk = next(iter(PhiladelphiaBid.__table__.c.run_id.foreign_keys))

    assert fk.ondelete == "SET NULL"


# =============================================================================
# What a live run found: the list page is not the dashboard
# =============================================================================

# The full Open Bids list (`bidList.sda`) reached through "View More…". A live
# run proved it does NOT carry `id="resultsTable"` — the wait timed out, the
# extractor found nothing, and a portal full of bids was reported as empty. The
# rows are the same shape; only the table's identity differs.
LIST_PAGE_HTML = """
<h1>Open Bids</h1>
<table class="table-02" cellspacing="1">
  <tbody>
    <tr>
      <th class="listheading">Bid #</th>
      <th class="listheading">Organization</th>
      <th class="listheading">Alternate Id</th>
      <th class="listheading">Buyer</th>
      <th class="listheading">Description</th>
      <th class="listheading">Bid Opening Date</th>
    </tr>
    <tr class="tableStripe-01">
      <td class="tableText-01"><a href="/bso/seller/bidAck.sda?bidId=B2727750">B2727750</a></td>
      <td class="tableText-01">City of Philadelphia</td>
      <td class="tableText-01"></td>
      <td class="tableText-01">Purchaser MP</td>
      <td class="tableText-01">Jackhammers</td>
      <td class="tableText-01">08/20/2026 11:59:00 PM</td>
    </tr>
    <tr class="tableStripe-02">
      <td class="tableText-01"><a href="/bso/seller/bidAck.sda?bidId=B2727746">B2727746</a></td>
      <td class="tableText-01">City of Philadelphia</td>
      <td class="tableText-01"></td>
      <td class="tableText-01">Purchaser MP</td>
      <td class="tableText-01">ANNUAL PREVENTIVE MAINT. SERVICE</td>
      <td class="tableText-01">08/19/2026 11:59:00 PM</td>
    </tr>
  </tbody>
</table>
"""

# The same list with no <th> at all — headers rendered as styled <td>s, which
# this platform does on some pages.
NO_HEADERS_HTML = """
<table>
  <tr><td class="listheading">Bid #</td><td class="listheading">Organization</td>
      <td class="listheading">Alternate Id</td><td class="listheading">Buyer</td>
      <td class="listheading">Description</td><td class="listheading">Bid Opening Date</td></tr>
  <tr>
    <td><a href="/bso/seller/bidAck.sda?bidId=B2727733">B2727733</a></td>
    <td>City of Philadelphia</td><td></td><td>Purchaser OIT</td>
    <td>10 Macbook Pro with AppleCare</td><td>08/17/2026 12:00:00 PM</td>
  </tr>
</table>
"""


def test_the_full_list_page_is_read_without_a_table_id(browser, tmp_path):
    """The regression, from a live run: `#resultsTable` is the *dashboard's* id.
    Keying on it made the whole Open Bids list unreadable."""
    scraper = _scraper_on(_render(browser, LIST_PAGE_HTML), tmp_path)
    rows = scraper.collect_rows()

    assert [r["bid_number"] for r in rows] == ["B2727750", "B2727746"]
    assert rows[0]["description"] == "Jackhammers"
    assert rows[1]["buyer"] == "Purchaser MP"


def test_a_list_rendered_without_header_cells_is_still_read(browser, tmp_path):
    """No <th> means no names to locate columns by, so the portal's own column
    order carries it — better than returning nothing from a page full of bids."""
    scraper = _scraper_on(_render(browser, NO_HEADERS_HTML), tmp_path)
    rows = scraper.collect_rows()

    assert len(rows) == 1
    assert rows[0]["bid_number"] == "B2727733"
    assert rows[0]["description"] == "10 Macbook Pro with AppleCare"


def test_an_empty_page_is_reported_with_what_it_actually_contained(browser, tmp_path, caplog):
    """A zero that cannot tell "no open bids today" from "the list is somewhere
    this did not look" is what made the first live run silent."""
    import logging

    caplog.set_level(logging.WARNING)
    scraper = _scraper_on(
        _render(browser, "<title>Session Expired</title><p>Please sign in again.</p>"),
        tmp_path,
    )
    scraper.screenshot = lambda name: None
    assert scraper.collect_rows() == []

    reported = " ".join(r.getMessage() for r in caplog.records)
    assert "no bid rows found" in reported
    assert "0 bidId links" in reported
    assert "Please sign in again" in reported, "say what the page said"


def test_the_dashboard_table_still_wins_when_both_are_present(browser, tmp_path):
    """The heading strategy runs first, so the dashboard keeps reading the Open
    Bids section rather than whichever table happens to have the most links."""
    scraper = _scraper_on(_render(browser, OPEN_BIDS_HTML), tmp_path)

    assert [r["bid_number"] for r in scraper.collect_rows()] == ["B2727750", "B2727732"]


# =============================================================================
# Pagination: the 41 records behind a 25-row page
# =============================================================================

# The pager, exactly as PHLContracts renders it under the Open Bids list. There
# is no Next button: a span for the page you are on, an anchor per other page,
# and the hrefs are script calls rather than URLs — which is why a walk that
# followed hrefs read 25 of 41 and stopped.
PAGER_HTML = """
<table id="resultsTable"><tbody>
  <tr><th>Bid #</th><th>Organization</th><th>Alternate Id</th><th>Buyer</th>
      <th>Description</th><th>Bid Opening Date</th></tr>
  <tr><td><a href="/bso/seller/bidAck.sda?bidId=B2727750" class="link-01">B2727750</a></td>
      <td>City of Philadelphia</td><td></td><td>Purchaser MP</td>
      <td>Jackhammers</td><td>08/20/2026 11:59:00 PM</td></tr>
  <tr><td><a href="/bso/seller/bidAck.sda?bidId=B2727732" class="link-01">B2727732</a></td>
      <td>City of Philadelphia</td><td></td><td>Purchaser MP</td>
      <td>License Tag Stickers</td><td>08/21/2026 11:59:00 PM</td></tr>
</tbody></table>
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
    <tbody>
    <tr>
        <td class="inputs-01" align="center" width="80%" aria-live="polite">1-25 of 41</td>
    </tr>
    <tr>
        <td align="center" class="inputs-01" valign="bottom">
            <nav aria-label="Pagination">
            <span aria-current="page">1</span>&nbsp;
            <a href="javascript:viewPage(2)" class="link-01">2</a>&nbsp;
            </nav>
        </td>
    </tr>
    </tbody>
</table>
"""

# The same pager once the turn has happened: page 2 is current, 1 links back,
# and the counter has moved to the last sixteen records.
PAGER_HTML_PAGE_2 = (
    PAGER_HTML.replace("1-25 of 41", "26-41 of 41")
    .replace('<span aria-current="page">1</span>', '<a href="javascript:viewPage(1)">1</a>')
    .replace('<a href="javascript:viewPage(2)" class="link-01">2</a>',
             '<span aria-current="page">2</span>')
    .replace("B2727750", "B2730001").replace("B2727732", "B2730002")
)


def _rows(start: int, count: int) -> list[dict]:
    return [{"bid_number": f"B{start + i}", "detail_url": f"/bso/x?bidId=B{start + i}"}
            for i in range(count)]


def _walking_scraper(tmp_path, pages, pagers, turns=True):
    """A scraper wired to a fake portal: `pages[i]` rows behind `pagers[i]`.

    Everything above the DOM is real — the accumulator, the de-duplication, the
    termination rules and the console stream. Only the browser is stood in for.
    """
    from app.core import run_manager

    run = run_manager.create_run("philadelphia", tmp_path)
    scraper = PhiladelphiaScraper(run["run_id"])
    at = {"page": 0}

    scraper.collect_rows = lambda: pages[at["page"]]
    scraper._pager_state = lambda: pagers[at["page"]]
    scraper._dashboard_preview_rows = lambda: []
    scraper._next_page_url = lambda seen: None
    scraper.screenshot = lambda name: None

    def click(target):
        if not turns:
            return False          # the click landed and the table never redrew
        at["page"] = target - 1
        return True

    scraper._click_to_page = click
    return scraper


def test_the_pager_names_the_page_after_this_one(browser, tmp_path):
    """The portal renders no Next control, so the next page is the anchor
    labelled one past the `aria-current` span — `javascript:viewPage(2)`."""
    scraper = _scraper_on(_render(browser, PAGER_HTML), tmp_path)
    pager = scraper._pager_state()

    assert pager["current"] == 1
    assert pager["next"] == 2, "page 2 is a script-call anchor, not an href"
    assert pager["total"] == 41, "the total comes from the '1-25 of 41' counter"


def test_the_last_page_has_no_page_after_it(browser, tmp_path):
    """On page 2 the only anchor points backwards — which is what ends the walk
    rather than a Next button going missing."""
    scraper = _scraper_on(_render(browser, PAGER_HTML_PAGE_2), tmp_path)
    pager = scraper._pager_state()

    assert pager["current"] == 2
    assert pager["next"] is None
    assert (pager["from"], pager["to"], pager["total"]) == (26, 41, 41)


def test_a_list_that_fits_on_one_page_has_no_pager(browser, tmp_path):
    """No pager is a complete read, not a failure to find one."""
    scraper = _scraper_on(_render(browser, OPEN_BIDS_HTML), tmp_path)

    assert scraper._pager_state() == {}


def test_the_page_is_derived_from_the_counter_when_nothing_is_marked_current(
    browser, tmp_path
):
    """A pager that styles the current page instead of marking it: rows 26-41 of
    a 25-row page is page 2, so the walk still knows what to click next."""
    html = PAGER_HTML_PAGE_2.replace('<span aria-current="page">2</span>', "<b>2</b>")
    scraper = _scraper_on(_render(browser, html), tmp_path)

    assert scraper._pager_state()["current"] == 2


def test_the_fingerprint_tells_the_two_pages_apart(browser, tmp_path):
    """`viewPage()` redraws the table in place: same selectors, same table id,
    different rows. Waiting for "a table with rows" would be satisfied by the
    page being left, so the wait compares content."""
    scraper = _scraper_on(_render(browser, PAGER_HTML), tmp_path)
    page_1 = scraper._list_signature()
    _render(browser, PAGER_HTML_PAGE_2)
    page_2 = scraper._list_signature()

    assert page_1 and page_2
    assert page_1 != page_2


def test_a_page_with_no_rows_has_no_fingerprint(browser, tmp_path):
    """"Cannot say", not "no rows" — an empty document must never read as a
    completed page turn."""
    scraper = _scraper_on(_render(browser, "<p>Signed out.</p>"), tmp_path)

    assert scraper._list_signature() == ""


def test_every_record_is_kept_across_the_page_turn(tmp_path):
    """The 41-of-41 case: page 2's sixteen append to page 1's twenty-five
    instead of replacing them."""
    scraper = _walking_scraper(
        tmp_path,
        [_rows(1, 25), _rows(26, 16)],
        [{"total": 41, "current": 1, "next": 2}, {"total": 41, "current": 2, "next": None}],
    )
    rows = scraper.collect_all_pages()

    assert len(rows) == 41
    assert rows[0]["bid_number"] == "B1" and rows[-1]["bid_number"] == "B41"


def test_the_console_stream_names_each_page_and_the_total(tmp_path, caplog):
    import logging

    caplog.set_level(logging.INFO)
    _walking_scraper(
        tmp_path,
        [_rows(1, 25), _rows(26, 16)],
        [{"total": 41, "current": 1, "next": 2}, {"total": 41, "current": 2, "next": None}],
    ).collect_all_pages()

    stream = "\n".join(r.getMessage() for r in caplog.records)
    assert "[PAGINATION INITIALIZED]" in stream
    assert "[PAGE 1]: 25 records extracted successfully." in stream
    assert "[PAGINATING]: Navigating to Page 2..." in stream
    assert "[PAGE 2]: 16 records extracted successfully." in stream
    assert "Total 41 records accumulated across 2 pages." in stream


def test_the_walk_stops_once_the_reported_total_is_in_hand(tmp_path):
    """A pager offering a page 3 that redraws page 1 would otherwise walk until
    the cap. The counter's total is believed over the pager's links."""
    scraper = _walking_scraper(
        tmp_path,
        [_rows(1, 25), _rows(26, 16), _rows(1, 25)],
        [{"total": 41, "current": 1, "next": 2},
         {"total": 41, "current": 2, "next": 3},
         {"total": 41, "current": 3, "next": None}],
    )

    assert len(scraper.collect_all_pages()) == 41


def test_a_page_turn_that_never_redraws_ends_the_walk(tmp_path):
    """Better twenty-five bids and a warning than the same twenty-five a hundred
    times over."""
    scraper = _walking_scraper(
        tmp_path,
        [_rows(1, 25), _rows(26, 16)],
        [{"total": 41, "current": 1, "next": 2}, {"total": 41, "current": 2, "next": None}],
        turns=False,
    )

    assert len(scraper.collect_all_pages()) == 25


def test_a_bid_printed_twice_is_carried_once(tmp_path):
    """The city lists a bid once per commodity it is classified under, so the
    same number can appear on one page or on two."""
    scraper = _walking_scraper(
        tmp_path,
        [_rows(1, 25) + _rows(1, 1), _rows(26, 16)],
        [{"total": 42, "current": 1, "next": 2}, {"total": 42, "current": 2, "next": None}],
    )
    rows = scraper.collect_all_pages()

    assert len(rows) == 41
    assert len({r["bid_number"] for r in rows}) == 41


# =============================================================================
# Attachments: finding the same file again after a download
# =============================================================================

# A bid's File Attachments row, as the detail page renders it. Five anchors, no
# URLs — the file id inside editFile() is all that identifies them.
ATTACHMENTS_HTML = """
<table>
  <tr>
    <td valign="top" class="t-head-01">File Attachments:</td>
    <td colspan="5">
      <a href="javascript:editFile('500001');" class="link-01">Seller_SP_Checklist_Bid B2727714.pdf</a><br>
      <a href="javascript:editFile('500002');" class="link-01">Micro Purchase Terms and Conditions B2727714.pdf</a><br>
      <a href="javascript:editFile('500003');" class="link-01">PHL-Contracts-Consent_Authorization BID B2727714.pdf</a><br>
      <a href="javascript:editFile('500004');" class="link-01">LGBTQ_Applicant_Data_Form Bid B2727714.pdf</a><br>
      <a href="javascript:editFile('500005');" class="link-01">80-247B FORM_Rev 11-2023~15.pdf</a>
    </td>
  </tr>
</table>
"""

# The same row after a click has taken the browser somewhere else: the page the
# download left behind carries no attachment anchors at all.
AFTER_DOWNLOAD_HTML = "<p>Your file is being prepared.</p>"

# Some bids print an attachment anchor twice — the same file id in two places in
# the layout. Counted twice, the second copy is downloaded again, and on a page
# that redraws it becomes an off-by-one against the list read at the start.
DUPLICATED_ANCHOR_HTML = """
<table><tr><td class="t-head-01">File Attachments:</td><td>
  <a href="javascript:editFile('600001');" class="link-01">80-247B FORM_Rev 11-2023~15.pdf</a><br>
  <a href="javascript:editFile('600002');" class="link-01">Insurance_Requirements.pdf</a>
</td></tr></table>
<table><tr><td>
  <a href="javascript:editFile('600001');" class="link-01">80-247B FORM_Rev 11-2023~15.pdf</a>
</td></tr></table>
"""


class _FakeClickDriver:
    """The real browser with one script call intercepted.

    Assigning over `driver.execute_script` is not possible on a Selenium driver,
    and replacing it wholesale would stub out the attachment read as well as the
    click — so the substitution is made here and everything else passed through.
    """

    def __init__(self, driver, execute_script):
        self._driver = driver
        self.execute_script = execute_script

    def __getattr__(self, name):
        return getattr(self._driver, name)


def test_each_attachment_is_identified_by_its_file_id(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ATTACHMENTS_HTML), tmp_path)
    attachments = scraper._attachment_list()

    assert [a["file_id"] for a in attachments] == [
        "500001", "500002", "500003", "500004", "500005",
    ]
    assert attachments[0]["name"] == "Seller_SP_Checklist_Bid B2727714.pdf"


def test_an_attachment_rendered_twice_is_listed_once(browser, tmp_path):
    """Same file id, two places in the layout — one file to download."""
    scraper = _scraper_on(_render(browser, DUPLICATED_ANCHOR_HTML), tmp_path)
    attachments = scraper._attachment_list()

    assert [a["file_id"] for a in attachments] == ["600001", "600002"]


def test_an_attachment_is_found_by_id_not_by_position(browser, tmp_path):
    """The fourth file is the fourth file wherever it sits in the anchor list."""
    scraper = _scraper_on(_render(browser, ATTACHMENTS_HTML), tmp_path)
    link = scraper._attachment_anchor("500004", None)

    assert link is not None
    assert link.text.strip() == "LGBTQ_Applicant_Data_Form Bid B2727714.pdf"


def test_a_download_that_leaves_the_page_reopens_the_bid(browser, tmp_path):
    """B2727714 lost four of five attachments this way: the first download took
    the browser off the detail page, and every file after it was reported as no
    longer on the page. Reopening the bid is what makes the rest reachable."""
    from urllib.parse import quote

    scraper = _scraper_on(_render(browser, AFTER_DOWNLOAD_HTML), tmp_path)
    detail_url = ("data:text/html;charset=utf-8,"
                  + quote(f"<html><body>{ATTACHMENTS_HTML}</body></html>"))

    link = scraper._attachment_anchor("500002", detail_url)

    assert link is not None, "the file was still there — the page had moved on"
    assert link.text.strip() == "Micro Purchase Terms and Conditions B2727714.pdf"


def test_a_missing_attachment_is_still_reported_when_reopening_does_not_help(
    browser, tmp_path
):
    """The reload is a second opinion, not a guarantee — a file that is genuinely
    gone must still be named rather than quietly dropped."""
    from urllib.parse import quote

    scraper = _scraper_on(_render(browser, AFTER_DOWNLOAD_HTML), tmp_path)
    detail_url = ("data:text/html;charset=utf-8,"
                  + quote(f"<html><body>{ATTACHMENTS_HTML}</body></html>"))

    assert scraper._attachment_anchor("999999", detail_url) is None


def test_every_attachment_is_saved_even_when_each_click_leaves_the_page(
    browser, tmp_path
):
    """The whole loop, against a page that empties itself after every download:
    five listed, five saved, none reported missing."""
    from urllib.parse import quote

    scraper = _scraper_on(_render(browser, ATTACHMENTS_HTML), tmp_path)
    detail_url = ("data:text/html;charset=utf-8,"
                  + quote(f"<html><body>{ATTACHMENTS_HTML}</body></html>"))
    folder = tmp_path / "bid"
    folder.mkdir()
    clicked: list[str] = []

    # Stand in for the servlet: the click lands, a file arrives, and the page it
    # was clicked from is gone. Everything else the scraper runs — reading the
    # attachment list, scrolling — goes through to the real browser.
    real_execute = browser.execute_script

    def fake_execute(script, *args):
        if "click()" not in script:
            return real_execute(script, *args)
        clicked.append(args[0].text.strip())
        _render(browser, AFTER_DOWNLOAD_HTML)

    def fake_download(timeout=None, ignore=None):
        landing = tmp_path / f"{len(clicked)}.pdf"
        landing.write_bytes(b"%PDF-1.4")
        return landing

    scraper.driver = _FakeClickDriver(browser, fake_execute)
    scraper.wait_for_download = fake_download

    saved, errors = scraper.download_attachments("B2727714", folder, detail_url)

    assert errors == []
    assert len(saved) == 5
    assert clicked[1] == "Micro Purchase Terms and Conditions B2727714.pdf"
    assert clicked[4] == "80-247B FORM_Rev 11-2023~15.pdf"


# =============================================================================
# The layout table that wraps the bid table
# =============================================================================

# The full Open Bids list as `bidList.sda` serves it: no id="resultsTable", the
# bid table nested inside a layout table, and a search panel above it carrying a
# header cell of its own.
#
# Every strategy in the extractor matches the *wrapper* here, because document
# order returns a table before the table it wraps and the wrapper carries the
# same bid links by descent. Read that way the page yields a phantom first row
# — the layout row holding the whole bid table, whose cells are every cell in it
# flattened into one — and a one-place column shift on the real rows from the
# search panel's stray <th>. That is one bug behind three symptoms: a duplicated
# bid number, a first record with the right Bid # and the wrong everything, and
# fields shifted by one.
WRAPPED_LIST_HTML = """
<table class="layout-01">
  <tr><td>
    <table><tr><th>Search Criteria</th></tr>
           <tr><td>Status: Open</td></tr></table>
  </td></tr>
  <tr><td>
    <table class="table-02">
      <tr>
        <th class="listheading">Bid #</th><th class="listheading">Organization</th>
        <th class="listheading">Alternate Id</th><th class="listheading">Buyer</th>
        <th class="listheading">Description</th><th class="listheading">Bid Opening Date</th>
      </tr>
      <tr>
        <td><a href="/bso/seller/bidAck.sda?bidId=B2727750">B2727750</a></td>
        <td>City of Philadelphia</td><td></td><td>Purchaser MP</td>
        <td>Jackhammers</td><td>08/20/2026 11:59:00 PM</td>
      </tr>
      <tr>
        <td><a href="/bso/seller/bidAck.sda?bidId=B2727732">B2727732</a></td>
        <td>Water Department</td><td></td><td>Theresa Baker</td>
        <td>License Tag Stickers</td><td>08/21/2026 11:59:00 PM</td>
      </tr>
    </table>
  </td></tr>
</table>
"""


def test_the_first_row_is_a_bid_not_the_layout_row_around_the_table(browser, tmp_path):
    """The reported bug: Bid # right, Organization / Buyer / Description /
    Opening Date wrong. The row it read was the layout row."""
    scraper = _scraper_on(_render(browser, WRAPPED_LIST_HTML), tmp_path)
    first = scraper.collect_rows()[0]

    assert first["bid_number"] == "B2727750"
    assert first["organization"] == "City of Philadelphia"
    assert first["buyer"] == "Purchaser MP"
    assert first["description"] == "Jackhammers"
    assert first["bid_opening_date"] == "08/20/2026 11:59:00 PM"


def test_a_wrapped_list_yields_one_row_per_bid(browser, tmp_path):
    """The phantom row carried the first bid's link, so the same bid number came
    back twice — which is what the console's duplicate-key warning was."""
    scraper = _scraper_on(_render(browser, WRAPPED_LIST_HTML), tmp_path)
    rows = scraper.collect_rows()

    assert [r["bid_number"] for r in rows] == ["B2727750", "B2727732"]


def test_a_header_from_a_neighbouring_table_does_not_shift_the_columns(
    browser, tmp_path
):
    """The search panel's "Search Criteria" <th> sat in front of the bid table's
    own headers, moving every column index along by one."""
    scraper = _scraper_on(_render(browser, WRAPPED_LIST_HTML), tmp_path)
    second = scraper.collect_rows()[1]

    assert second["organization"] == "Water Department", "columns shifted by one"
    assert second["buyer"] == "Theresa Baker"
    assert second["alternate_id"] == ""


def test_every_row_of_a_wrapped_list_is_read(browser, tmp_path):
    """A long page, to be sure the fix is not "the first row happens to work"."""
    body = "".join(
        f'<tr><td><a href="/bso/seller/bidAck.sda?bidId=B{n}">B{n}</a></td>'
        f"<td>Org {n}</td><td></td><td>Buyer {n}</td><td>Desc {n}</td>"
        f"<td>08/{n}/2026</td></tr>"
        for n in range(10, 35)
    )
    html = WRAPPED_LIST_HTML.replace("</table>\n  </td></tr>\n</table>", body + "</table></td></tr></table>")
    scraper = _scraper_on(_render(browser, html), tmp_path)
    rows = scraper.collect_rows()

    assert len(rows) == 27, "two original rows plus twenty-five"
    assert rows[5]["description"] == "Desc 13"
    assert rows[5]["organization"] == "Org 13"


def test_cell_text_is_normalised_before_it_is_recorded(browser, tmp_path):
    """Nested markup and the layout's non-breaking spaces are the portal's, not
    the value's — a Description must not reach Excel carrying either."""
    html = WRAPPED_LIST_HTML.replace(
        "<td>Jackhammers</td>",
        "<td>\n  <b>Jackhammers</b>&nbsp;&mdash;\n  <span>Rental</span>\n</td>",
    )
    scraper = _scraper_on(_render(browser, html), tmp_path)

    assert scraper.collect_rows()[0]["description"] == "Jackhammers — Rental"


# =============================================================================
# Document counts: what the sheet reports vs what is on disk
# =============================================================================


def test_the_document_count_is_taken_from_the_bids_folder(tmp_path):
    folder = storage.bid_folder(tmp_path, "B2727714")
    for name in ("Specs_Part1.pdf", "Pricing_Sheet.xlsx", "Terms.pdf"):
        (folder / name).write_bytes(b"%PDF-1.4")

    assert storage.saved_documents(tmp_path, "B2727714") == [
        "Pricing_Sheet.xlsx", "Specs_Part1.pdf", "Terms.pdf",
    ]


@pytest.mark.parametrize("generated", sorted(storage.GENERATED_FILENAMES))
def test_the_runs_own_files_are_not_counted_as_documents(tmp_path, generated):
    """`bid_items_details.txt` is written by the run, not published by the city —
    counting it would report one document for a bid that has none. The retired
    JSON name is here too: a workspace left over from before it was dropped must
    not start counting as a document now."""
    folder = storage.bid_folder(tmp_path, "B2626875")
    (folder / generated).write_text("written by the run")

    assert storage.saved_documents(tmp_path, "B2626875") == []


def test_a_bid_whose_folder_was_never_made_counts_zero(tmp_path):
    assert storage.saved_documents(tmp_path, "B0000000") == []


def test_a_download_wait_does_not_return_a_file_that_was_already_there(tmp_path):
    """The partial-download bug: between the click and Chrome opening its
    `.crdownload`, the newest file in the staging folder is still the *previous*
    download. Returned then, the same file is claimed twice and the bid ends up
    short — so the wait is told what was already staged."""
    import threading
    import time as _time

    from app.core.base_scraper import BaseScraper

    staging = tmp_path / "_downloads"
    staging.mkdir()
    previous = staging / "already_here.pdf"
    previous.write_bytes(b"%PDF-1.4")

    waiter = object.__new__(BaseScraper)
    waiter.download_dir = staging
    waiter.raise_if_stopped = lambda: None

    def land_the_real_one():
        _time.sleep(1.0)
        (staging / "the_new_one.pdf").write_bytes(b"%PDF-1.4")

    threading.Thread(target=land_the_real_one, daemon=True).start()
    got = waiter.wait_for_download(timeout=15, ignore={previous})

    assert got.name == "the_new_one.pdf"


def test_a_download_wait_without_the_staged_set_still_returns_the_newest(tmp_path):
    """The other scrapers call this with one download in flight at a time, and
    must keep working exactly as before."""
    from app.core.base_scraper import BaseScraper

    staging = tmp_path / "_downloads"
    staging.mkdir()
    (staging / "only_one.pdf").write_bytes(b"%PDF-1.4")

    waiter = object.__new__(BaseScraper)
    waiter.download_dir = staging
    waiter.raise_if_stopped = lambda: None

    assert waiter.wait_for_download(timeout=5).name == "only_one.pdf"


# =============================================================================
# Deliverables a client can open: header metadata in the sheet, items in .txt
# =============================================================================

# A detail page's item table, in the shape BSO renders it: nested in a layout
# table, headers naming the columns, and a full-width continuation row carrying
# the long specification for the item above it.
ITEMS_HTML = """
<table class="layout-01"><tr><td>
  <table class="table-02">
    <tr>
      <th class="listheading">Item #</th><th class="listheading">Description</th>
      <th class="listheading">Quantity</th><th class="listheading">UOM</th>
      <th class="listheading">Unit Price</th><th class="listheading">NIGP Code</th>
    </tr>
    <tr>
      <td class="tableText-01">1</td>
      <td class="tableText-01">Submersible Water Pump (50HP)</td>
      <td class="tableText-01">2</td><td class="tableText-01">Each</td>
      <td class="tableText-01">$4,200.00</td><td class="tableText-01">670-45</td>
    </tr>
    <tr>
      <td class="tableText-01" colspan="6">
        High-efficiency industrial pump with cast iron casing.
      </td>
    </tr>
    <tr>
      <td class="tableText-01">2</td>
      <td class="tableText-01">Installation &amp; Piping</td>
      <td class="tableText-01">1</td><td class="tableText-01">Lot</td>
      <td class="tableText-01">$8,000.00</td><td class="tableText-01">910-32</td>
    </tr>
  </table>
</td></tr></table>
"""

# The header table as the portal publishes it for a bid that carries the three
# promoted fields, under labels that differ from the column names.
HEADER_WITH_FISCAL_YEAR = {
    "Bid Number": "23-10492",
    "Description": "Pump Replacement Services",
    "Fiscal Year": "FY2027",
    "Bid Type": "Formal Solicitation",
    "Pre-Bid Conference": "08/18/2026 10:00 AM, 1515 Arch St, 17th Floor",
    "Department": "28 - Water Department",
    "Info Contact": "Michael Perce 215-459-4753",
    "File Attachments": "Specs.pdf",
}


def test_the_promoted_header_fields_reach_their_own_columns():
    from app.scrapers.philadelphia import details

    promoted = details.promote_header(HEADER_WITH_FISCAL_YEAR)

    assert promoted["fiscal_year"] == "FY2027"
    assert promoted["solicitation_type"] == "Formal Solicitation"
    assert promoted["pre_bid_conference"].startswith("08/18/2026 10:00 AM")


def test_a_label_the_city_words_differently_still_lands_in_its_column():
    """"FY" and "Pre Bid Meeting" are the same fields under other names — a
    column that empties itself when the portal rewords a row is not a column."""
    from app.scrapers.philadelphia import details

    promoted = details.promote_header(
        {"FY": "2027", "Pre Bid Meeting": "None scheduled", "Type Code": "MP"}
    )

    assert promoted["fiscal_year"] == "2027"
    assert promoted["pre_bid_conference"] == "None scheduled"
    assert promoted["solicitation_type"] == "MP"


def test_a_bid_missing_those_fields_gets_blank_cells_not_missing_keys():
    from app.scrapers.philadelphia import details

    assert details.promote_header({}) == {
        "fiscal_year": "", "solicitation_type": "", "pre_bid_conference": "",
    }


def test_the_rest_of_the_header_table_reaches_the_sheet_as_one_readable_cell():
    """This is what replaces the JSON file: nothing published is lost, and no
    label needs a column of its own."""
    from app.scrapers.philadelphia import details

    cell = details.additional_header(HEADER_WITH_FISCAL_YEAR)

    assert "Department: 28 - Water Department" in cell
    assert "Info Contact: Michael Perce 215-459-4753" in cell
    assert "Fiscal Year" not in cell, "it has a column of its own"
    assert "Description" not in cell, "it has a column of its own"
    assert "File Attachments" not in cell, "the bid's folder is the file list"


def test_the_summary_sheet_carries_the_expanded_schema(tmp_path):
    from openpyxl import load_workbook

    from app.scrapers.philadelphia import details, export

    record = {
        "bid_number": "23-10492", "description": "Pump Replacement Services",
        "organization": "Water Department", "buyer": "John Doe",
        "bid_opening_date": "08/25/2026", "documents_downloaded": 5,
        "extra_header_data": HEADER_WITH_FISCAL_YEAR,
        "items": [{"name": "Submersible Water Pump (50HP)"}, {"name": "Installation"}],
        **details.promote_header(HEADER_WITH_FISCAL_YEAR),
        "additional_header": details.additional_header(HEADER_WITH_FISCAL_YEAR),
    }
    path = storage.summary_path(tmp_path)
    export.generate_excel_from_records([record], path)
    rows = list(load_workbook(path).active.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    row = rows[1]

    assert "Fiscal Year" in headers
    assert "Procurement / Solicitation Type" in headers
    assert "Pre-Bid Conference Date / Details" in headers
    assert row[headers.index("Fiscal Year")] == "FY2027"
    assert row[headers.index("Procurement / Solicitation Type")] == "Formal Solicitation"
    assert row[headers.index("Total Document Count")] == 5
    assert row[headers.index("Line Items")] == 2
    assert "Department: 28 - Water Department" in row[headers.index(
        "Additional Header Information")]


def test_the_sheet_rebuilt_from_the_database_reads_the_same(tmp_path):
    """`generate_excel` rebuilds months later from stored rows, which carry the
    header table but not the rendered cell — it has to render the same one."""
    from app.scrapers.philadelphia import details, export

    class Row:
        bid_number = "23-10492"
        description = "Pump Replacement Services"
        extra_header_data = HEADER_WITH_FISCAL_YEAR
        items = [{"name": "a"}]
        additional_header = None

    assert export._cell(Row(), "additional_header") == details.additional_header(
        HEADER_WITH_FISCAL_YEAR
    )
    assert export._cell(Row(), "item_count") == 1


def test_the_item_text_file_is_readable_without_software():
    from app.scrapers.philadelphia import details

    text = details.render_items_text({
        "bid_number": "23-10492",
        "description": "Pump Replacement Services",
        "fiscal_year": "FY2027",
        "items": [
            {"item_number": "1", "name": "Submersible Water Pump (50HP)",
             "quantity": "2", "unit": "Each",
             "specification": "High-efficiency industrial pump with cast iron casing."},
            {"item_number": "2", "name": "Installation & Piping",
             "quantity": "1", "unit": "Lot",
             "specification": "Complete on-site installation and pressure testing."},
        ],
    })

    assert "BID ITEM SPECIFICATIONS" in text
    assert "Bid Number: 23-10492" in text
    assert "Title: Pump Replacement Services" in text
    assert "Fiscal Year: FY2027" in text
    assert "Item #1:" in text and "Item #2:" in text
    assert "  - Item Name: Submersible Water Pump (50HP)" in text
    assert "  - Quantity: 2" in text
    assert "  - Unit of Measure: Each" in text
    assert "  - Specification Details: Complete on-site installation and pressure testing." in text
    assert "{" not in text and "}" not in text, "this is prose, not a data format"


def test_a_bid_with_no_items_still_gets_a_file_that_names_it():
    """A folder of PDFs with nothing saying which bid they belong to is what
    this file exists to prevent."""
    from app.scrapers.philadelphia import details

    text = details.render_items_text({"bid_number": "B2727750", "description": "Jackhammers"})

    assert "Bid Number: B2727750" in text
    assert "no line-item breakdown" in text


def test_an_item_column_the_city_adds_is_printed_rather_than_dropped():
    from app.scrapers.philadelphia import details

    text = details.render_items_text({
        "bid_number": "B1", "items": [{"name": "Pump", "manufacturer": "Grundfos"}],
    })

    assert "  - Manufacturer: Grundfos" in text


def test_the_item_table_is_read_with_its_columns_named(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ITEMS_HTML), tmp_path)
    items = scraper.scrape_items("23-10492")

    assert len(items) == 2, "two items, not the continuation row as a third"
    assert items[0]["name"] == "Submersible Water Pump (50HP)"
    assert items[0]["quantity"] == "2"
    assert items[0]["unit"] == "Each"
    assert items[0]["unit_price"] == "$4,200.00"
    assert items[0]["nigp_code"] == "670-45"
    assert items[1]["name"] == "Installation & Piping"


def test_a_full_width_row_is_the_specification_for_the_item_above_it(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ITEMS_HTML), tmp_path)
    items = scraper.scrape_items("23-10492")

    assert items[0]["specification"] == (
        "High-efficiency industrial pump with cast iron casing."
    )
    assert items[1]["specification"] == "", "it belongs to item 1, not to both"


def test_a_detail_page_with_no_item_table_reads_as_no_items(browser, tmp_path):
    """Plenty of bids describe the work in their attachments. That is an empty
    list, not a failure."""
    scraper = _scraper_on(_render(browser, HEADER_HTML), tmp_path)

    assert scraper.scrape_items("B2727750") == []


def test_the_item_file_lands_in_the_bids_own_folder(tmp_path):
    from app.core import run_manager

    run = run_manager.create_run("philadelphia", tmp_path)
    scraper = PhiladelphiaScraper(run["run_id"])
    scraper._write_items_file({
        "bid_number": "B2727750", "description": "Jackhammers",
        "items": [{"name": "Jackhammer", "quantity": "4"}],
    })

    written = storage.items_path(scraper.run_dir, "B2727750")
    assert written.name == "bid_items_details.txt"
    assert "Bids_Data/B2727750" in str(written)
    assert "  - Item Name: Jackhammer" in written.read_text()
    assert storage.saved_documents(scraper.run_dir, "B2727750") == [], (
        "the file the run wrote is not one of the city's documents"
    )


# =============================================================================
# Advanced Search: the portal's own form, driven from the dashboard
# =============================================================================

# The Advanced Search page, reduced to the controls a run touches: the Document
# Type dropdown that reveals the bid form, the criteria inputs (every id
# colon-joined, because the page is JSF), the two selects the page fills in for
# itself, and the Search button.
ADVANCED_SEARCH_HTML = """
<div class="hidden-xs">
  <a id="advancedSearchTopNav"
     href="javascript:gotoBsoURL('view/search/supplier/advancedSearch.xhtml');"
     class="advanced-search-link">Advanced</a>
</div>
<form id="advancedSearchForm">
  <select id="advancedSearchForm:documentTypeSelect" name="advancedSearchForm:documentTypeSelect">
    <option value="">Select Document Type...</option>
    <option value="BID_SOLICITATIONS">Bid Solicitations</option>
    <option value="CONTRACT_BLANKETS">Blankets</option>
    <option value="VENDORS">Vendors</option>
  </select>
</form>
<form id="bidSearchForm">
  <input id="bidSearchForm:bidNbr" type="text">
  <input id="bidSearchForm:alternateId" type="text">
  <input id="bidSearchForm:desc" type="text">
  <input id="bidSearchForm:itemDesc" type="text">
  <select id="bidSearchForm:organization">
    <option value="">Select Organization...</option>
    <option value="AGENCY">City of Philadelphia</option>
  </select>
  <select id="bidSearchForm:departmentPrefix" disabled="disabled">
    <option value="">Select Department...</option>
  </select>
  <select id="bidSearchForm:buyer">
    <option value="">Select Buyer...</option>
    <option value="C.BELL">Bell, Carla</option>
    <option value="MP_PROD">MP, Purchaser</option>
  </select>
  <select id="bidSearchForm:classId">
    <option value="">Select NIGP Class...</option>
    <option value="720">720 - PUMPING EQUIPMENT AND ACCESSORIES</option>
  </select>
  <select id="bidSearchForm:classItemId" disabled="disabled">
    <option value="">Select NIGP Class Item...</option>
  </select>
  <select id="bidSearchForm:typeCode">
    <option value="">Select Type Code...</option>
    <option value="MI">Micro Purchase</option>
    <option value="RQ">Request for Proposal</option>
  </select>
  <input id="bidSearchForm:openingDateFrom_input" type="text">
  <input id="bidSearchForm:openingDateTo_input" type="text">
  <select id="bidSearchForm:status">
    <option value="">Select Status...</option>
    <option value="2BS">Sent</option>
  </select>
  <select id="bidSearchForm:categoryCode">
    <option value="">Select Category...</option>
    <option value="30">Water and Sewer Treatment Equipment, Supplies, and Services</option>
  </select>
  <div id="bidSearchForm:searchScopeType">
    <input id="bidSearchForm:searchScopeType_input" type="checkbox">
  </div>
  <button id="bidSearchForm:btnBidSearch" type="button">Search</button>
</form>
"""


def _value_of(browser, element_id):
    from selenium.webdriver.common.by import By

    return browser.find_element(By.ID, element_id).get_attribute("value")


def test_a_blank_field_is_not_a_filter():
    """The portal treats an empty input as no criterion, so carrying one through
    would only make a run claim to be narrower than it is."""
    from app.scrapers.philadelphia import search

    assert search.clean_filters(
        {"description": "  pumps  ", "buyer": "", "status": "   ", "category": None}
    ) == {"description": "pumps"}


def test_a_key_that_is_not_on_the_form_is_dropped():
    """The only thing to do with an id the page does not have is fail
    confusingly later."""
    from app.scrapers.philadelphia import search

    assert search.clean_filters({"description": "pumps", "sql": "drop table"}) == {
        "description": "pumps"
    }
    assert search.clean_filters("not a payload") == {}
    assert search.clean_filters(None) == {}


def test_a_run_with_no_criteria_describes_itself_as_the_whole_list():
    from app.scrapers.philadelphia import search

    assert search.describe({}) == "all open bids"
    assert search.describe(search.clean_filters({"description": ""})) == "all open bids"


def test_a_search_describes_itself_in_the_forms_own_words():
    """This is what a reader sees next to a stored run months later."""
    from app.scrapers.philadelphia import search

    summary = search.describe(
        search.clean_filters(
            {"description": "pumps", "type_code": "MI", "match_any": True}
        )
    )

    assert summary == "Type code: MI · Description: pumps · Match: any criterion"


def test_a_dependent_control_is_never_filled_before_its_parent():
    """Organization fills Department over AJAX and NIGP Class fills its Item.
    Filling a dependant first puts a value into a disabled, empty select — a
    filter silently dropped."""
    from app.scrapers.philadelphia import search

    order = list(search.ORDERED_FIELDS)
    for child, parent in search.DEPENDS_ON.items():
        assert order.index(parent) < order.index(child), f"{child} before {parent}"


def test_every_field_the_form_offers_has_a_label():
    from app.scrapers.philadelphia import search

    assert set(search.FILTER_KEYS) <= set(search.LABELS)


def test_the_document_type_dropdown_is_set_to_bid_solicitations(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    # Prove the selection happens rather than being skipped as already-present.
    scraper._search_form_present = lambda: False
    scraper._choose_bid_solicitations()

    assert _value_of(browser, "advancedSearchForm:documentTypeSelect") == "BID_SOLICITATIONS"


def test_every_criterion_reaches_its_own_input(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._fill_search_form({
        "description": "pump replacement",
        "item_description": "submersible",
        "bid_number": "B2727750",
        "opening_date_from": "08/01/2026",
        "opening_date_to": "09/30/2026",
    })

    assert _value_of(browser, "bidSearchForm:desc") == "pump replacement"
    assert _value_of(browser, "bidSearchForm:itemDesc") == "submersible"
    assert _value_of(browser, "bidSearchForm:bidNbr") == "B2727750"
    assert _value_of(browser, "bidSearchForm:openingDateFrom_input") == "08/01/2026"
    assert _value_of(browser, "bidSearchForm:openingDateTo_input") == "09/30/2026"


def test_a_dropdown_takes_the_portals_code(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._fill_search_form({"type_code": "MI", "status": "2BS"})

    assert _value_of(browser, "bidSearchForm:typeCode") == "MI"
    assert _value_of(browser, "bidSearchForm:status") == "2BS"


def test_a_dropdown_also_takes_what_it_says_on_screen(browser, tmp_path):
    """The dashboard's form stays in the words of the person filling it in:
    nobody should have to know that Carla Bell is `C.BELL`."""
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._fill_search_form({
        "buyer": "Bell, Carla",
        "type_code": "Micro Purchase",
        "category": "Water and Sewer Treatment Equipment, Supplies, and Services",
    })

    assert _value_of(browser, "bidSearchForm:buyer") == "C.BELL"
    assert _value_of(browser, "bidSearchForm:typeCode") == "MI"
    assert _value_of(browser, "bidSearchForm:categoryCode") == "30"


def test_a_partial_name_finds_its_option(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._fill_search_form({"buyer": "carla", "nigp_class": "pumping"})

    assert _value_of(browser, "bidSearchForm:buyer") == "C.BELL"
    assert _value_of(browser, "bidSearchForm:classId") == "720"


def test_a_partial_never_settles_for_the_placeholder(browser, tmp_path):
    """"Select Buyer..." matches far too readily — a filter that lands on the
    placeholder is a search that quietly did not filter."""
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._fill_search_form({"buyer": "Select"})

    assert _value_of(browser, "bidSearchForm:buyer") == ""


def test_a_criterion_the_form_will_not_take_is_reported_not_dropped(
    browser, tmp_path, caplog
):
    """A run that silently searched on less than it was asked for returns the
    wrong bids and looks like it worked."""
    import logging

    caplog.set_level(logging.WARNING)
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._fill_search_form({"buyer": "Nobody By That Name", "description": "pumps"})

    reported = " ".join(r.getMessage() for r in caplog.records)
    assert "could not be applied" in reported
    assert "Nobody By That Name" in reported
    assert _value_of(browser, "bidSearchForm:desc") == "pumps", "the rest still applied"

    from app.core import run_manager

    warnings = (run_manager.get_run(scraper.run_id) or {}).get("warnings") or []
    assert any("broader than asked for" in str(w) for w in warnings)


def test_a_disabled_dependent_select_is_not_filled_in(browser, tmp_path):
    """Department starts disabled with one empty option and is filled in by the
    page after an organization is chosen. This fixture never fires that AJAX, so
    the criterion has to be reported rather than written into a dead control."""
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._await_dependent_select = lambda element_id, timeout=0: False

    assert scraper._fill_one_criterion("department", "Water") is False


def test_match_any_flips_the_criteria_switch(browser, tmp_path):
    from selenium.webdriver.common.by import By

    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    scraper._fill_search_form({"match_any": True})

    checkbox = browser.find_element(By.ID, "bidSearchForm:searchScopeType_input")
    assert checkbox.is_selected(), "Match Criteria stayed on All"


def test_the_search_button_is_clicked_and_the_wait_is_for_rows(browser, tmp_path):
    """The button runs a PrimeFaces AJAX update rather than navigating, so what
    is waited on is content — and an empty panel is an answer, not a failure."""
    scraper = _scraper_on(_render(browser, ADVANCED_SEARCH_HTML), tmp_path)
    clicks: list[str] = []
    real_execute = browser.execute_script
    scraper.driver = _FakeClickDriver(
        browser,
        lambda script, *args: (
            clicks.append(args[0].get_attribute("id")) if "click()" in script
            else real_execute(script, *args)
        ),
    )

    scraper._submit_search()

    assert clicks == ["bidSearchForm:btnBidSearch"]


# =============================================================================
# Line items: the block layout the portal actually uses
# =============================================================================

# B2626551, reconstructed from a live run's own log. PHLContracts does not put
# line items in a header-and-columns grid — it prints a block per item: a cell
# naming the item and its NIGP class-item, beside a cell holding the commodity
# code and the description, with quantity and unit as labelled text beneath.
ITEM_BLOCKS_HTML = """
<table><tr><td>
<table>
  <tr><td class="t-head-01">Item: 1&nbsp;&nbsp;072-08</td>
      <td class="tableText-01">42831-002-156  FREIGHTLINER 114SD CHASSIS AS PER DFS SPEC 25026CNGb.25</td></tr>
  <tr><td colspan="2">Quantity: 2   UOM: EA   Unit Cost: $284,000.00</td></tr>
  <tr><td colspan="2">Chassis to be delivered to the Streets Department.</td></tr>
  <tr><td class="t-head-01">Item: 2&nbsp;&nbsp;072-09</td>
      <td class="tableText-01">42831-002-157  20HD CNG COMPACTOR BODY</td></tr>
  <tr><td colspan="2">Quantity: 1   UOM: EA</td></tr>
</table>
</td></tr></table>
"""

# A results list whose description happens to carry the words the old detector
# keyed on. "Item: 1 072-08" begins with "item"; "AS PER DFS SPEC" contains
# "spec". Matching those loosely is what made a data row look like a header row
# and every bid report zero items.
NOT_AN_ITEM_TABLE_HTML = """
<table>
  <tr><th>Bid #</th><th>Organization</th><th>Description</th></tr>
  <tr><td><a href="/bso/seller/bidAck.sda?bidId=B1">B1</a></td><td>City</td>
      <td>AS PER DFS SPEC 25026CNGb.25 line item replacement</td></tr>
</table>
"""


def test_the_block_layout_is_read_as_line_items(browser, tmp_path):
    """The shape a live run met: no header row, one block per item."""
    scraper = _scraper_on(_render(browser, ITEM_BLOCKS_HTML), tmp_path)
    items = scraper.scrape_items("B2626551")

    assert len(items) == 2
    assert items[0]["item_number"] == "1"
    assert items[0]["name"] == "FREIGHTLINER 114SD CHASSIS AS PER DFS SPEC 25026CNGb.25"
    assert items[0]["nigp_code"] == "072-08"
    assert items[0]["commodity_code"] == "42831-002-156"
    assert items[1]["item_number"] == "2"
    assert items[1]["name"] == "20HD CNG COMPACTOR BODY"


def test_a_blocks_quantity_and_unit_are_read_from_the_text_beneath_it(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ITEM_BLOCKS_HTML), tmp_path)
    items = scraper.scrape_items("B2626551")

    assert items[0]["quantity"] == "2"
    assert items[0]["unit"] == "EA"
    assert items[0]["unit_price"] == "$284,000.00"
    assert items[0]["specification"] == "Chassis to be delivered to the Streets Department."
    assert items[1]["quantity"] == "1"
    assert items[1]["specification"] == "", "it belongs to item 1, not to both"


def test_a_data_row_is_not_mistaken_for_a_header_row(browser, tmp_path):
    """The regression: `(item|line)` anywhere plus `spec` anywhere matched a data
    row, the first row was eaten as a header, and 45 bids reported zero items."""
    scraper = _scraper_on(_render(browser, NOT_AN_ITEM_TABLE_HTML), tmp_path)

    assert scraper.scrape_items("B1") == []


def test_a_header_cell_has_to_be_a_label_not_contain_one(browser, tmp_path):
    """Sixty characters of description is not a column heading, however many
    heading words it happens to contain."""
    html = """
    <table>
      <tr><td>Item: 1 072-08</td>
          <td>A LONG DESCRIPTION MENTIONING QUANTITY AND UNIT COST AND SPEC IN PASSING</td></tr>
    </table>
    """
    scraper = _scraper_on(_render(browser, html), tmp_path)
    items = scraper.scrape_items("B1")

    # Read as a block — which it is — rather than as a grid with that row eaten.
    assert len(items) == 1
    assert items[0]["item_number"] == "1"


def test_the_grid_layout_still_wins_when_the_page_has_one(browser, tmp_path):
    scraper = _scraper_on(_render(browser, ITEMS_HTML), tmp_path)
    items = scraper.scrape_items("23-10492")

    assert len(items) == 2
    assert items[0]["name"] == "Submersible Water Pump (50HP)"
    assert items[0]["unit"] == "Each"


def test_the_block_items_reach_the_text_file_readably():
    from app.scrapers.philadelphia import details

    text = details.render_items_text({
        "bid_number": "B2626551",
        "description": "6x4 CNG Truck with 20HD Compactor Body",
        "items": [{
            "item_number": "1",
            "name": "FREIGHTLINER 114SD CHASSIS AS PER DFS SPEC 25026CNGb.25",
            "commodity_code": "42831-002-156", "nigp_code": "072-08",
            "quantity": "2", "unit": "EA", "unit_price": "$284,000.00",
            "specification": "Chassis to be delivered to the Streets Department.",
        }],
    })

    assert "Item #1:" in text
    assert "  - Item Name: FREIGHTLINER 114SD CHASSIS AS PER DFS SPEC 25026CNGb.25" in text
    assert "  - Quantity: 2" in text
    assert "  - Commodity Code: 42831-002-156" in text
    assert "  - NIGP Code: 072-08" in text
