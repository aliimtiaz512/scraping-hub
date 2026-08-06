"""Unison's detail page: parsing it, and turning it into evaluator inputs.

The fixture is a real buy detail page (Buy #1210780_01) with its section
structure intact. What these pin down is the parsing of each section the report
needs, and — more importantly — the two pieces of the page that must *not* reach
the classifier, both of which were found by running the real evaluator over
them.

    server/.venv/bin/python -m pytest server/tests/test_unison_detail.py
"""

import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.scrapers.unison import detail, evaluation  # noqa: E402

FIXTURE = Path(__file__).parent / "fixtures" / "unison_buy_details.html"
PAGE_URL = "https://marketplace.unisonglobal.com/fbweb/buyDetails.do?buy_id=1210979"


@pytest.fixture(scope="module")
def parsed():
    return detail.parse(FIXTURE.read_text(encoding="utf-8"), PAGE_URL)


@pytest.fixture(scope="module")
def record(parsed):
    """The flat shape the runner hands the evaluator."""
    return {**parsed["general_info"], **parsed}


# -- General Buy Information --------------------------------------------------


def test_every_general_information_field_is_captured(parsed):
    info = parsed["general_info"]
    assert info["buy_number"] == "1210780_01"
    assert info["solicitation_number"] == "PR16173540"
    assert info["buy_description"] == "IT Items"
    assert info["category"] == "7B -- IT AND TELECOM - COMPUTE"
    assert info["subcategory"].startswith("7B20 -- IT AND TELECOM")
    assert info["naics"].startswith("541519 -- EXCEPTION")
    assert info["naics_size_standard"] == "150 Number of Employees"
    assert info["sam_contract_opportunity"] == "No"
    assert info["set_aside"] == "Small Business"
    assert info["buyer"] == "U.S. Embassy Buenos Aires, Argentina"
    assert info["end_date"] == "08/06/2026"
    assert info["end_time"] == "15:00 ET"
    assert info["seller_question_deadline"] == "No Seller Question Deadline Set"
    assert info["delivery"] == "Special Delivery Instructions - See Buy Terms"


def test_the_portals_generous_whitespace_is_collapsed(parsed):
    """Several values arrive wrapped over four indented lines."""
    assert parsed["general_info"]["naics_size_standard"] == "150 Number of Employees"


def test_the_repost_reason_keeps_the_reason_and_drops_the_standing_note(parsed):
    reason = parsed["general_info"]["repost_reason"]
    assert reason == "No Repost Reason provided."
    assert "courtesy only" not in reason


def test_an_unmapped_field_is_kept_rather_than_dropped():
    html = """
    <div class="title"><h2>General Buy Information</h2></div>
    <table><tbody>
      <tr><th scope="row">Buy #:</th><td>1</td></tr>
      <tr><th scope="row">Something New:</th><td>a value</td></tr>
    </tbody></table>
    """
    info = detail.parse(html)["general_info"]
    assert info["extra"] == {"Something New": "a value"}


# -- the other sections -------------------------------------------------------


def test_line_items_come_from_the_items_table_not_the_template_section(parsed):
    """"Line Item(s) Template - Optional" sits immediately before "Line Item(s)"
    and holds a Download Template button — a contains-match takes the wrong one."""
    items = parsed["line_items"]
    assert len(items) == 5
    assert items[0] == {
        "no": "001",
        "description": "4K HDMI Cable 15m (50ft) — for long-distance display connections",
        "qty": "4",
        "unit": "each",
    }
    assert not any("Download Template" in item["description"] for item in items)


def test_bidding_requirements_split_into_label_and_instruction(parsed):
    requirements = parsed["bidding_requirements"]
    assert [r["name"] for r in requirements][:2] == [
        "Open Market",
        "Contracting Method - Lowest Price Technically Acceptable (LPTA)",
    ]
    assert requirements[0]["text"].startswith("Open Market bids are accepted")


def test_buy_terms_come_back_as_name_and_description(parsed):
    terms = {t["name"].strip(): t["text"] for t in parsed["buy_terms"]}
    assert "Product Certification" in terms
    assert "Argentinean Vendors" in terms
    assert terms["Argentinean Vendors"].startswith("Vendors MUST be located in Argentina")


def test_shipping_information_is_the_place_of_performance(parsed):
    """A foreign buy commonly fills the city alone — empty state/zip is normal."""
    assert parsed["shipping"] == {"city": "Buenos Aires", "state": "", "zip": ""}


def test_seller_attachment_requirements_drop_the_standing_paragraph(parsed):
    required = parsed["seller_attachments_required"]
    assert required == "Specification sheet of quoted items is required"
    assert "20 MB" not in required


def test_buy_attachments_carry_an_absolute_download_link(parsed):
    attachments = parsed["attachments"]
    assert len(attachments) == 1
    assert attachments[0]["name"] == "ATTACHMENT_1_EXPRESS_NDAA_CERTIFICATION.pdf"
    assert attachments[0]["size"] == "522 KB"
    assert attachments[0]["url"].startswith(
        "https://marketplace.unisonglobal.com/fbweb/viewAtt.do?token="
    )


def test_a_page_missing_a_section_parses_to_empty_rather_than_failing():
    parsed = detail.parse("<html><body><p>nothing here</p></body></html>")
    assert parsed["general_info"] == {}
    assert parsed["line_items"] == []
    assert parsed["attachments"] == []
    assert parsed["shipping"] == {}


# -- Buy # --------------------------------------------------------------------


def test_the_buy_number_keeps_its_full_value_and_yields_the_suffix():
    assert detail.split_buy_number("1210780_01") == ("1210780_01", "01")
    assert detail.split_buy_number(" 1210344_03 ") == ("1210344_03", "03")


def test_a_buy_number_without_a_suffix_counts_as_zero_not_blank():
    """The column is a tally, so every row carries a number."""
    assert detail.split_buy_number("1210980") == ("1210980", "0")
    assert detail.split_buy_number("1210980_") == ("1210980_", "0")
    assert detail.split_buy_number("") == ("", "0")


# -- evaluator inputs ---------------------------------------------------------


def test_the_line_items_prove_a_supply_and_hint_hardware(record):
    hint, evidence = evaluation.requirement_hint(record)
    assert hint == "HARDWARE"
    assert "line items are quantified products" in evidence


def test_a_shipping_only_row_is_not_evidence_of_goods():
    """Most buys carry a "Shipping / 1 / LOT" row; on its own it proves nothing."""
    hint, _ = evaluation.requirement_hint(
        {"buy_description": "Something", "line_items": [
            {"description": "Shipping", "qty": "1", "unit": "LOT"},
        ]}
    )
    assert hint is None


def test_a_service_buy_is_never_hinted_into_hardware():
    hint, evidence = evaluation.requirement_hint({
        "buy_description": "Grease Trap Cleaning FY27 JBLM",
        "line_items": [{"description": "Grease trap cleaning", "qty": "12", "unit": "each"}],
    })
    assert hint is None
    assert "names a service" in evidence


def test_service_line_items_outvote_a_neutral_description():
    hint, _ = evaluation.requirement_hint({
        "buy_description": "FY27 Requirement",
        "line_items": [
            {"description": "Annual maintenance of chillers", "qty": "1", "unit": "each"},
            {"description": "Quarterly inspection visits", "qty": "4", "unit": "each"},
            {"description": "Replacement filters", "qty": "8", "unit": "each"},
        ],
    })
    assert hint is None


def test_the_naics_code_is_taken_bare_from_its_label(record):
    assert evaluation.naics_code(record) == "541519"
    assert evaluation.naics_code({"naics": ""}) == ""


def test_the_full_text_carries_every_section_for_the_body(record):
    full_text = evaluation.build_full_text(record, document_text="NDAA certification")
    for heading in ("Buy Description", "General Information", "Line Items",
                    "Bidding Requirements", "Buy Terms", "Place of Performance",
                    "Attached Documents"):
        assert f"{heading}:" in full_text
    assert "Buenos Aires" in full_text
    assert "NDAA certification" in full_text


# -- the two traps ------------------------------------------------------------


def test_the_classifier_reads_the_buy_description_alone(record, monkeypatch):
    """Category, Buy Terms and the rest reach the evaluator as body text only.

    Both exclusions are load-bearing: the subcategory contains the word SOFTWARE
    (Rule B #3) and the Buy Terms mention research and development (Rule B #20),
    so either one classifying would reject this hardware buy.
    """
    captured: dict = {}

    def fake_evaluate(bid_id, full_text, naics_code="", title="", requirement_hint=None):
        captured.update(title=title, full_text=full_text)
        return {"decision": "PURSUE", "reason": "", "requirement_type": "HARDWARE",
                "rule": "A", "location": None, "hinted": True}

    monkeypatch.setattr(evaluation, "evaluate_bid", fake_evaluate)
    evaluation.evaluate({**record, "buy_description": "IT Items"})

    assert captured["title"] == "IT Items"
    for trap in ("SOFTWARE", "research and development", "7B20"):
        assert trap not in captured["title"]
        assert trap.lower() in captured["full_text"].lower()  # …but the body has it


def test_the_real_funnel_pursues_this_buy(record):
    """End to end against the actual evaluator: a reseller-NAICS supply bound for
    Argentina is Rule A — pursued regardless of location."""
    verdict = evaluation.evaluate({**record, "buy_description": "IT Items",
                                   "buy_number": "1210780_01"})
    assert verdict["decision"] == "PURSUE"
    assert verdict["rule"] == "A"
    assert verdict["requirement_type"] == "HARDWARE"
    assert verdict["requirement_hinted"] is True


def test_the_verdict_carries_no_evaluator_payload(record):
    """Only the decision and its standard phrase are meant to travel onward."""
    verdict = evaluation.evaluate({**record, "buy_description": "IT Items"})
    assert "full_text" not in verdict
    assert set(verdict) == {
        "decision", "reason", "requirement_type", "rule", "location",
        "requirement_hinted", "hint_evidence",
    }


# -- the report's row colours -------------------------------------------------


def test_rejected_bids_are_tinted_light_red_in_the_export(tmp_path):
    """The verdict in the row colour, matching the SAM export's palette."""
    from openpyxl import load_workbook

    from app.scrapers.unison import export

    records = [
        {"buyer_number": "1", "decision": "PURSUE", "buyer_description": "Laptops"},
        {"buyer_number": "2", "decision": "REJECT", "buyer_description": "Kosher meals"},
        {"buyer_number": "3", "decision": "MANUAL_REVIEW", "buyer_description": "Grease trap"},
    ]
    out = tmp_path / "unison.xlsx"
    assert export.generate_excel_from_records(records, out) == 3

    sheet = load_workbook(out).active
    fills = [sheet.cell(row=r, column=1).fill.fgColor.rgb for r in (2, 3, 4)]
    assert fills[0] in ("00000000", None)   # pursued rows stay plain
    assert fills[1] == "00FFCCCC"           # light red — skip these
    assert fills[2] == "00FFD966"           # amber — the rows needing a person
    # The whole row is tinted, not just its first cell.
    assert {c.fill.fgColor.rgb for c in sheet[3] if c.value is not None} == {"00FFCCCC"}
