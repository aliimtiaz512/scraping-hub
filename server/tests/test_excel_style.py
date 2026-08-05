"""The shared Excel look (app/core/excel_style) that every portal export uses.

Pure openpyxl — no DB, no browser. These pin the standard itself; the portal
tests that build real sheets (SEPTA, RideMetro, the MyFlorida sweep) then only
have to check their own extras.

    server/.venv/bin/python -m pytest server/tests/test_excel_style.py
"""

import os
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook  # noqa: E402

from app.core import excel_style  # noqa: E402

HEADERS = ["Ref #", "Title", "Closing Date"]
ROWS = [
    ["R-1", "A reasonably long solicitation title", "2026-09-01"],
    ["R-2", "Short", "2026-09-02"],
]


def _sheet(headers=HEADERS, rows=ROWS):
    """A written-and-reloaded sheet, so what is asserted is what Excel gets."""
    workbook, sheet = excel_style.new_workbook("Test")
    written = excel_style.write_table(sheet, headers, rows)
    with TemporaryDirectory() as tmp:
        out = Path(tmp) / "sheet.xlsx"
        workbook.save(str(out))
        return load_workbook(out).active, written


# -- the header standard -----------------------------------------------------


def test_header_row_is_filled_bold_white_and_centred():
    sheet, _ = _sheet()
    for cell in sheet[1]:
        assert cell.fill.fgColor.rgb == f"00{excel_style.HEADER_COLOR}"
        assert cell.font.bold is True
        assert cell.font.color.rgb in ("00FFFFFF", "FFFFFFFF")
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"


def test_header_cells_are_bordered_and_the_row_is_tall_enough():
    sheet, _ = _sheet()
    for cell in sheet[1]:
        for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom):
            assert side.style == "thin"
    assert 24 <= sheet.row_dimensions[1].height <= 28


def test_data_rows_are_left_unstyled():
    """The standard is about the header: a data row keeps Excel's own default,
    so a portal that tints its own rows (SAM, the sweep) is not fought."""
    sheet, _ = _sheet()
    assert sheet["A2"].font.bold in (False, None)
    assert sheet["A2"].fill.fgColor.rgb in ("00000000", None)


def test_header_row_is_frozen_unless_the_caller_opts_out():
    sheet, _ = _sheet()
    assert sheet.freeze_panes == "A2"

    workbook, unfrozen = excel_style.new_workbook("Test")
    excel_style.write_table(unfrozen, HEADERS, ROWS, freeze=False)
    assert unfrozen.freeze_panes is None


def test_style_header_row_can_target_a_repeated_header():
    """RideMetro writes a header row under every agency banner."""
    workbook, sheet = excel_style.new_workbook("Test")
    sheet.append(["ignored"])
    sheet.append(HEADERS)
    excel_style.style_header_row(sheet, row=2, last_column=len(HEADERS))
    assert sheet["A2"].fill.fgColor.rgb == f"00{excel_style.HEADER_COLOR}"
    assert sheet["A1"].fill.fgColor.rgb in ("00000000", None)


# -- widths ------------------------------------------------------------------


def test_columns_fit_their_longest_value_plus_padding():
    sheet, _ = _sheet()
    longest = len("A reasonably long solicitation title")
    assert sheet.column_dimensions["B"].width == longest + excel_style.WIDTH_PADDING


def test_widths_stay_between_the_floor_and_the_cap():
    sheet, _ = _sheet(["N"], [["x"], ["y" * 500]])
    assert sheet.column_dimensions["A"].width == excel_style.MAX_WIDTH

    narrow, _ = _sheet(["N"], [["x"]])
    assert narrow.column_dimensions["A"].width == excel_style.MIN_WIDTH


def test_a_merged_banner_does_not_stretch_the_first_column():
    """The banner is as wide as the sheet, not as wide as column A."""
    workbook, sheet = excel_style.new_workbook("Test")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    sheet["A1"] = "An extremely long agency name that spans the whole sheet"
    sheet.append(HEADERS)
    sheet.append(ROWS[0])
    excel_style.autofit_columns(sheet)
    assert sheet.column_dimensions["A"].width == excel_style.MIN_WIDTH


def test_a_wrapped_multiline_value_is_measured_by_its_widest_line():
    sheet, _ = _sheet(["N"], [["a\nbbbbbbbbbbbbbbbbbbbb\nc"]])
    assert sheet.column_dimensions["A"].width == 20 + excel_style.WIDTH_PADDING


# -- cell values -------------------------------------------------------------


def test_control_characters_are_stripped_so_the_workbook_saves():
    assert excel_style.sanitize_cell("bell\x07char") == "bellchar"
    sheet, _ = _sheet(["N"], [["bell\x07char"]])
    assert sheet["A2"].value == "bellchar"


def test_an_over_long_value_is_truncated_with_a_marker():
    value = "x" * (excel_style.MAX_CELL_CHARS + 500)
    result = excel_style.sanitize_cell(value)
    assert result.endswith(excel_style.TRUNCATION_MARKER)
    assert len(result) == excel_style.MAX_CELL_CHARS + len(excel_style.TRUNCATION_MARKER)


def test_non_strings_keep_their_type():
    """Dates and numbers must reach Excel as dates and numbers, not text."""
    from datetime import date

    for value in (None, 42, 3.5, True, date(2026, 9, 1)):
        assert excel_style.sanitize_cell(value) is value


def test_write_table_returns_the_row_count_and_writes_every_row():
    sheet, written = _sheet()
    assert written == len(ROWS)
    assert sheet.max_row == len(ROWS) + 1
    assert [c.value for c in sheet[1]] == HEADERS
    assert [c.value for c in sheet[2]] == ROWS[0]


def test_a_header_only_sheet_is_still_formatted():
    sheet, written = _sheet(HEADERS, [])
    assert written == 0
    assert sheet.max_row == 1
    assert sheet["A1"].font.bold is True
