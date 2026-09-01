"""MFMP's evaluation criteria: the deterministic tiers and the LLM behind them.

Source of truth: `MFMP_Bid_Evaluation_Criteria.docx` in the repository root.
These tests are written against its sections rather than against the code, so a
rule that drifts from the document fails here.

The cases are the document's own: §3's six excluded categories, §4's seven
lanes, §5's three review routes, and — the two that decide whether this engine
is any use — §7's edge cases, where a laundry *equipment* purchase must not be
rejected as a laundry *service*, and a software-coded Agency Decision must not
be pursued as software.

    server/.venv/bin/python -m pytest server/tests/test_myflorida_evaluation.py
"""

import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.core import excel_style  # noqa: E402
from app.scrapers.myflorida import evaluation, ollama_bridge, workbook  # noqa: E402


def bid(title, ad_type="Invitation to Bid", codes="", description=""):
    return {
        "ad_number": "AD-1", "title": title, "ad_type": ad_type,
        "commodity_codes": codes, "description": description,
    }


# =============================================================================
# Tier 1 — the six auto-REJECT categories (§3)
# =============================================================================


@pytest.mark.parametrize("title, codes, category", [
    ("Mowing and herbicide application", "70141501 — Lawn care", "Agriculture"),
    ("Aerial application for invasive species", "", "Agriculture"),
    ("Substance Abuse Treatment Program", "85101500 — Health", "Health & social"),
    ("Foster care disbursement unit support", "93131600 — Child welfare", "Health & social"),
    ("Statewide towing and Road Ranger services", "78141505 — Towing", "Waste/roadside"),
    ("Solid waste and trash collection", "76121501 — Refuse", "Waste/roadside"),
    ("Moving and relocation services", "78101804 — Moving", "Waste/roadside"),
    ("Lease of property, Tallahassee", "80131500 — Lease", "Real estate"),
    ("Event sponsorship opportunity", "80141609 — Sponsorship", "Real estate"),
    ("Commercial laundry service", "91111502 — Laundry", "Textile care"),
    ("Program evaluation and administration", "80101500 — Consulting", "Consulting"),
])
def test_the_excluded_categories_are_rejected(title, codes, category):
    verdict = evaluation.evaluate(bid(title, codes=codes))

    assert verdict["decision"] == evaluation.REJECT, category
    assert verdict["rule"].startswith("Tier 1")


def test_a_rejection_says_which_category_and_what_matched():
    """The reason and evidence are what let a reader disagree with the engine.
    A red row with no explanation is the manual pass again, with extra steps."""
    verdict = evaluation.evaluate(bid("Mowing services", codes="70141501 — Lawn care"))

    assert "Agriculture" in verdict["reason"]
    assert "70141501" in verdict["evidence"]


# =============================================================================
# Tier 2 — the seven lanes (§4)
# =============================================================================


@pytest.mark.parametrize("title, codes, lane", [
    ("Agency portal application development", "81111501 — Software", "Software/Web"),
    ("Cloud hosting and SaaS subscription", "81162000 — Cloud", "Software/Web"),
    ("System maintenance and support", "", "Software/Web"),
    ("Printing of technical manuals", "82121511 — Printing", "Printing"),
    ("Industrial printing services", "73151900 — Printing", "Printing"),
    ("Graphic design and branding", "", "Graphic Design"),
    ("Digital marketing and SEO", "", "Digital Marketing"),
    ("Artificial intelligence and data analytics", "", "AI/Data"),
    ("Printed circuit board assembly", "32101500 — Components", "PCB/Electronics"),
    ("Prequalification of information technology consultants", "80101507 — IT", "IT staffing"),
])
def test_the_company_lanes_are_pursued(title, codes, lane):
    verdict = evaluation.evaluate(bid(title, codes=codes))

    assert verdict["decision"] == evaluation.PURSUE, lane
    assert verdict["rule"].startswith("Tier 2")


# =============================================================================
# Tier 3 — what only a document can decide (§5)
# =============================================================================


@pytest.mark.parametrize("title", [
    "Roof replacement, Building 4",
    "Well drilling and septic installation",
    "HVAC chiller replacement",
    "Elevator modernization",
    "Fire alarm system installation",
    "Asphalt paving and resurfacing",
    "Marine construction — seawall repair",
])
def test_construction_and_trades_go_to_review_whatever_the_code(title):
    """§5.1 — the client kept and rejected near-identical bids in this category,
    so no keyword rule decides them. Even a lane commodity code does not."""
    verdict = evaluation.evaluate(bid(title, codes="81111501 — Software"))

    assert verdict["decision"] == evaluation.MANUAL_REVIEW
    assert verdict["rule"] == "5.1 construction/trades"
    assert verdict["needs_documents"] is True


def test_the_construction_review_asks_for_what_the_document_must_answer():
    verdict = evaluation.evaluate(bid("Roof replacement, Building 4"))

    for expected in ("value", "licens", "labour"):
        assert expected in verdict["reason"].lower()


def test_a_sole_source_notice_is_never_auto_pursued():
    """§5.2 with §7's edge case: a single-source award notice for a software
    vendor carries Software/Web codes. Checking lanes first would pursue a
    contract that is not open for competition at all — which is why §5.2 runs
    before Tier 2 and not after it."""
    verdict = evaluation.evaluate(bid(
        "Intent to Award Single Source provider for Case Management Portal",
        ad_type="Agency Decision", codes="81111501 — Software",
    ))

    assert verdict["decision"] == evaluation.MANUAL_REVIEW
    assert verdict["rule"] == "5.2 sole source"


def test_a_sole_source_notice_naming_the_company_is_not_swept_aside():
    """The one sole-source posting that is not automatically uninteresting, and
    the reason §5.2 is a review rather than a rejection."""
    verdict = evaluation.evaluate(bid(
        "Intent to Award Single Source to Rizviz International Impex",
        ad_type="Agency Decision", codes="81111501 — Software",
    ))

    assert verdict["decision"] == evaluation.PURSUE


def test_a_bid_spanning_an_excluded_category_and_a_lane_goes_to_review():
    """§5.3 — mixed commodity codes. Tier 2 says "no Tier 1 exclusion also
    applies", so a bid hitting both is by definition undecided."""
    verdict = evaluation.evaluate(bid(
        "Laundry services portal and application development",
        codes="91111502 — Laundry | 81111501 — Software",
    ))

    assert verdict["decision"] == evaluation.MANUAL_REVIEW
    assert verdict["rule"] == "5.3 mixed codes"


def test_a_bid_no_rule_covers_goes_to_review_not_to_reject():
    """A bid nobody wrote a rule for is not a bid the client said no to.
    Rejecting it silently would bury exactly the postings worth a new rule."""
    verdict = evaluation.evaluate(bid("Bulk rock salt delivery", codes="11111600 — Minerals"))

    assert verdict["decision"] == evaluation.MANUAL_REVIEW
    assert verdict["rule"] == "unmatched"


# =============================================================================
# §7 — the edge cases the criteria calls out by name
# =============================================================================


def test_a_laundry_equipment_purchase_is_not_a_laundry_service():
    """§7: "keep the services-vs-goods distinction in the Tier 1 checks so
    equipment purchases aren't wrongly swept into a services exclusion"."""
    service = evaluation.evaluate(bid("Commercial laundry service", codes="91111502 — Laundry"))
    goods = evaluation.evaluate(bid("Purchase of commercial laundry equipment",
                                    codes="91111502 — Laundry"))

    assert service["decision"] == evaluation.REJECT
    assert goods["decision"] != evaluation.REJECT


def test_the_goods_override_applies_only_where_the_category_is_a_service():
    """The word "equipment" must not rescue a bid from a category that never had
    a goods/services distinction to make — mowing equipment is still agriculture."""
    verdict = evaluation.evaluate(bid("Purchase of mowing equipment", codes="70141501 — Lawn"))

    assert verdict["decision"] == evaluation.REJECT


def test_a_grant_that_funds_a_lane_deliverable_is_pursued():
    """§3: health/social grant programs are excluded "unless it funds a
    print/design/software/marketing/AI deliverable"."""
    plain = evaluation.evaluate(bid("Substance abuse grant program",
                                    ad_type="Grant Opportunities", codes="85101500 — Health"))
    funded = evaluation.evaluate(bid("Grant program — graphic design and branding outreach"))

    assert plain["decision"] == evaluation.REJECT
    assert funded["decision"] == evaluation.PURSUE


# =============================================================================
# Robustness
# =============================================================================


def test_an_empty_bid_is_reviewed_rather_than_decided():
    verdict = evaluation.evaluate({})

    assert verdict["decision"] == evaluation.MANUAL_REVIEW


def test_the_evaluator_never_raises():
    """A bid is never worth a failed run."""
    verdict = evaluation.evaluate({"title": None, "commodity_codes": 12345, "ad_type": object()})

    assert verdict["decision"] in {evaluation.PURSUE, evaluation.REJECT, evaluation.MANUAL_REVIEW}


def test_the_codes_are_read_from_the_cell_the_sheet_shows():
    """Engine and spreadsheet cannot disagree about what a bid was coded as."""
    codes = evaluation.commodity_codes(
        {"commodity_codes": "42211705 — Hearing aid | 45111701 — Assistive listening devices"}
    )

    assert codes == ["42211705", "45111701"]


def test_the_body_boilerplate_does_not_decide_a_bid():
    """A Florida advertisement body carries statutory clauses that mention
    construction and consulting in passing. Only the opening is read, and a bid
    is not routed to review by a paragraph 4,000 characters in."""
    verdict = evaluation.evaluate(bid(
        "Agency portal application development",
        codes="81111501 — Software",
        description="Scope: build the portal. " + ("filler " * 400) + " roofing construction",
    ))

    assert verdict["decision"] == evaluation.PURSUE


# =============================================================================
# The resolution layer
# =============================================================================


def test_a_low_confidence_answer_does_not_turn_a_row_red():
    """The point of this layer is to remove bids from the manual pass. A
    coin-flip landing on REJECT hides a bid behind a red fill, which is worse
    than leaving it yellow."""
    parsed = ollama_bridge.parse_response(
        "DECISION: REJECT\nREASON: probably out of scope\nCONFIDENCE: LOW"
    )

    assert parsed["decision"] == "REJECT"
    assert parsed["confidence"] == "LOW"


def test_an_unparseable_answer_leaves_the_bid_undecided():
    """Tinting a row red on the strength of text nobody could read is the one
    outcome worse than not deciding."""
    assert ollama_bridge.parse_response("I think this bid is fine, honestly.") is None
    assert ollama_bridge.parse_response("DECISION: MAYBE\nREASON: x") is None
    assert ollama_bridge.parse_response("") is None


def test_a_good_answer_becomes_a_one_line_note():
    parsed = ollama_bridge.parse_response(
        "DECISION: REJECT\n"
        "REASON: Requires Florida CGC licence and is majority on-site trade labour\n"
        "CONFIDENCE: HIGH"
    )

    assert parsed["decision"] == "REJECT"
    assert parsed["ai_notes"] == (
        "Requires Florida CGC licence and is majority on-site trade labour"
    )
    assert "\n" not in parsed["ai_notes"]
    assert len(parsed["ai_notes"]) <= 200


def test_the_prompt_carries_the_tier_specific_question():
    """A general "is this in scope" question on a construction bid reproduces
    exactly the ambiguity §5.1 exists to resolve."""
    prompt = ollama_bridge.build_prompt(
        bid("Roof replacement"),
        {"rule": "5.1 construction/trades", "reason": "construction scope"},
        "SCOPE OF WORK: replace 12,000 sq ft of roofing.",
    )

    assert "contract value" in prompt.lower()
    assert "licen" in prompt.lower()
    assert "on-site skilled-trade labour" in prompt.lower()
    assert "SCOPE OF WORK: replace 12,000 sq ft" in prompt


def test_the_sole_source_prompt_defaults_to_reject():
    """§5.2 — "Default to REJECT (the contract isn't competable)"."""
    prompt = ollama_bridge.build_prompt(
        bid("Intent to award", ad_type="Agency Decision"),
        {"rule": "5.2 sole source", "reason": "sole source"},
        "",
    )

    # The guidance is a wrapped paragraph, so the assertion is on its words
    # rather than on where the source happens to break the line.
    assert "REJECT unless the named vendor is Rizviz" in " ".join(prompt.split())


def test_documents_are_read_scope_first_and_bounded(tmp_path):
    """A 200-page specification costs minutes per bid and says nothing about
    whether the job is subcontractable; the addendum naming the value does."""
    (tmp_path / "Technical_Specification.txt").write_text("spec " * 5000)
    (tmp_path / "Scope_of_Work.txt").write_text("Estimated value $250,000. Florida CGC required.")

    text = ollama_bridge.read_documents(tmp_path, budget=1000)

    assert text.index("Scope_of_Work.txt") < text.index("Technical_Specification.txt")
    assert "Estimated value $250,000" in text
    assert len(text) < 1500


def test_a_missing_document_folder_is_no_evidence_not_an_error(tmp_path):
    assert ollama_bridge.read_documents(tmp_path / "nothing here") == ""


def test_a_disabled_ollama_leaves_the_bid_for_a_person(monkeypatch):
    monkeypatch.setattr(ollama_bridge, "OLLAMA_ENABLED", False)

    assert ollama_bridge.resolve(bid("Roof replacement"), {"rule": "5.1 construction/trades"}) is None


def test_an_unreachable_ollama_leaves_the_bid_for_a_person(monkeypatch):
    """A model that is down must never fail a run or decide a bid."""
    monkeypatch.setattr(ollama_bridge, "OLLAMA_ENABLED", True)

    def boom(*_a, **_k):
        raise ConnectionError("connection refused")

    monkeypatch.setattr(ollama_bridge.requests, "post", boom)

    assert ollama_bridge.resolve(bid("Roof replacement"), {"rule": "5.1 construction/trades"}) is None


def test_a_low_confidence_resolution_is_downgraded_to_review(monkeypatch):
    monkeypatch.setattr(ollama_bridge, "OLLAMA_ENABLED", True)

    class Response:
        def raise_for_status(self):
            pass

        def json(self):
            return {"response": "DECISION: REJECT\nREASON: unclear scope\nCONFIDENCE: LOW"}

    monkeypatch.setattr(ollama_bridge.requests, "post", lambda *_a, **_k: Response())

    resolved = ollama_bridge.resolve(bid("Roof replacement"), {"rule": "5.1 construction/trades"})

    assert resolved["decision"] == "MANUAL_REVIEW"
    assert resolved["ai_notes"].startswith("Low confidence:")


# =============================================================================
# The spreadsheet
# =============================================================================


def test_the_client_red_is_the_colour_they_use_by_hand():
    """The criteria names FFFF0000 because it is the fill the client already
    applies when they strike a bid off a sheet."""
    fill, _font = excel_style.ROW_STYLES["client_reject"]

    assert fill.fgColor.rgb == "00FFFF0000" or fill.fgColor.rgb.endswith("FFFF0000")


@pytest.mark.parametrize("decision, style", [
    ("REJECT", "client_reject"),
    ("MANUAL_REVIEW", "review"),
    ("PURSUE", None),
    ("", None),
])
def test_the_row_tint_follows_the_verdict(decision, style):
    row = ["x"] * len(workbook.RECORD_COLUMNS)
    row[workbook._DECISION_INDEX] = decision

    assert workbook._row_style(row) == style


def test_the_sheet_carries_the_verdict_and_the_note(tmp_path):
    from openpyxl import load_workbook

    path = workbook.build_from_records([
        {"ad_number": "AD-1", "title": "Mowing", "decision": "REJECT",
         "evaluation_reason": "Out of scope — Agriculture (Tier 1)", "ai_notes": ""},
        {"ad_number": "AD-2", "title": "Roofing", "decision": "MANUAL_REVIEW",
         "evaluation_reason": "Construction scope", "ai_notes": "Needs a person"},
        {"ad_number": "AD-3", "title": "Portal build", "decision": "PURSUE",
         "evaluation_reason": "In scope — Software/Web lane (Tier 2)", "ai_notes": ""},
    ], tmp_path)
    sheet = load_workbook(path).active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for expected in ("Evaluation Status", "Evaluation Reason", "AI Notes"):
        assert expected in headers
    assert rows[1][headers.index("Evaluation Status")] == "REJECT"
    assert rows[2][headers.index("AI Notes")] == "Needs a person"
    # The verdict is a column, not a filter — the PURSUE row is still here and
    # so is the REJECT one.
    assert len(rows) == 4


def test_a_rejected_bid_is_filled_with_the_clients_red(tmp_path):
    from openpyxl import load_workbook

    path = workbook.build_from_records(
        [{"ad_number": "AD-1", "title": "Mowing", "decision": "REJECT"}], tmp_path
    )
    sheet = load_workbook(path).active

    assert sheet.cell(2, 1).fill.fgColor.rgb.endswith("FFFF0000")


def test_a_pursued_bid_is_left_alone(tmp_path):
    """An in-scope row is written exactly as it is, which is what makes the
    tinted ones stand out at all."""
    from openpyxl import load_workbook

    path = workbook.build_from_records(
        [{"ad_number": "AD-1", "title": "Portal", "decision": "PURSUE"}], tmp_path
    )
    sheet = load_workbook(path).active

    assert sheet.cell(2, 1).fill.fgColor.rgb in (None, "00000000")
