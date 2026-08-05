"""SEPTA's Bid module: navigation, the title blacklist, and the two-sheet export.

Pure logic and stubbed drivers — no browser, no portal, no DB.

The Open Bids pass reuses the Quotes blacklist against a bid's *title*, so the
terms themselves are already covered by test_septa_exclusions.py. What is new
here, and what these test, is everything around that: that the Bid module is
told apart from the Quotes form it can redirect back to, that a blacklisted
title is dropped before the bid is recorded, that the two grids never leak into
each other's storage, and that the workbook carries a sheet for each.

    server/.venv/bin/python -m pytest server/tests/test_septa_open_bids.py
"""

import os
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook  # noqa: E402

from app.core import run_manager  # noqa: E402
from app.scrapers.septa import export  # noqa: E402
from app.scrapers.septa.filters import OpenDateFilter  # noqa: E402
from app.scrapers.septa.models import OPEN_BID_EXCEL_COLUMNS  # noqa: E402
from app.scrapers.septa.scraper import SEL, SeptaScraper, _bid_columns  # noqa: E402

QUOTES_URL = "https://epsadmin.septa.org/vendor/requisitions/search/"
BIDS_URL = "https://epsadmin.septa.org/vendor/bids/search/"


def _scraper():
    run = run_manager.create_run("septa", Path("/tmp"))
    return SeptaScraper(run["run_id"], OpenDateFilter())


def _bid(number, title, close="12/01/2026"):
    return {
        "bid_number": number,
        "title": title,
        "open_date": "08/01/2026",
        "close_date": close,
    }


# -- the title blacklist ----------------------------------------------------
#
# The four terms the requirement names, in their split form (GASKET and CUMMINS
# independently) — the same list the Quotes grid uses. See exclusions.py.

# Realistic Bid-module titles. Unlike the Quotes grid these are solicitation
# titles rather than parts lines, which is exactly why the terms have to survive
# the change of context.
EXCLUDED_TITLES = [
    ("GASKET CUMMINS 3974127 FILTER HEAD", "GASKET"),
    ("Procurement of CUMMINS Engine Assemblies", "CUMMINS"),
    ("GASKET MEULLER INDUSTRIES P35708", "GASKET"),
    ("BRACKET NEW FLYER 695953 ASSY-ORBSTAR", "NEW FLYER"),
    ("Overhaul of NOVA Bus Fleet", "NOVA"),
    ("FILTER NF 6321254 AIR", "NF"),
    ("RIVET USSC 9904-000018-005 (NF 6406513)", "NF"),
]

KEPT_TITLES = [
    "Rail Car Truck Overhaul Program",
    "STEP SEPTA SPEC S-3544-9 DWG D-4948-F",
    # The regression the whole-word rule exists for, in a bid title this time.
    "IGBT INFINEON BSM400GA170DLC NO",
    # INNOVATION contains NOVA as a substring but not as a word.
    "Station Innovation and Wayfinding Study",
]


def test_a_blacklisted_title_never_reaches_the_records():
    s = _scraper()
    for i, (title, _) in enumerate(EXCLUDED_TITLES):
        assert s._record_open_bid(_bid(f"B{i}", title)) is False, title
    assert s._open_bids == []
    assert s._excluded_by_title == len(EXCLUDED_TITLES)


def test_each_skip_is_attributed_to_the_term_that_fired():
    """A filter that drops rows without saying which rule fired is
    indistinguishable from a scrape that simply missed them."""
    s = _scraper()
    for i, (title, expected) in enumerate(EXCLUDED_TITLES):
        s._record_open_bid(_bid(f"B{i}", title))
        assert s._title_exclusion_reasons[expected] >= 1, (title, expected)
    assert sum(s._title_exclusion_reasons.values()) == len(EXCLUDED_TITLES)


def test_legitimate_titles_survive():
    s = _scraper()
    for i, title in enumerate(KEPT_TITLES):
        assert s._record_open_bid(_bid(f"B{i}", title)) is True, title
    assert len(s._open_bids) == len(KEPT_TITLES)
    assert s._excluded_by_title == 0


def test_matching_is_case_insensitive():
    """The requirement says case-insensitive; the grid is not all-caps."""
    s = _scraper()
    for title in ("New Flyer bus parts", "new flyer BUS PARTS", "Nova Bus Overhaul"):
        assert s._record_open_bid(_bid("B1", title)) is False, title
    assert s._open_bids == []


def test_the_grid_repeating_a_bid_across_pages_stores_it_once():
    s = _scraper()
    assert s._record_open_bid(_bid("B1", "Rail Car Truck Overhaul")) is True
    assert s._record_open_bid(_bid("B1", "Rail Car Truck Overhaul")) is False
    assert len(s._open_bids) == 1


def test_kept_bids_carry_only_the_exported_columns():
    s = _scraper()
    s._record_open_bid(_bid("B1", "Rail Car Truck Overhaul"))
    assert set(s._open_bids[0]) == {attr for attr, _ in OPEN_BID_EXCEL_COLUMNS}


def test_the_two_grids_never_leak_into_each_other():
    """Quotes and bids key on different columns and land in different tables."""
    s = _scraper()
    s._record_quote({
        "requisition_number": "A4", "summary": "STEP SEPTA SPEC S-3544-9",
        "open_date": "08/01/2026", "close_date": "12/01/2026",
    })
    s._record_open_bid(_bid("B1", "Rail Car Truck Overhaul"))

    assert len(s._records) == 1 and len(s._open_bids) == 1
    assert "bid_number" not in s._records[0]
    assert "requisition_number" not in s._open_bids[0]
    # Separate dedup maps: a bid numbered like a requisition must not collide.
    assert set(s._seen) == {"A4"}
    assert set(s._seen_open_bids) == {"B1"}


def test_the_quote_and_bid_tallies_are_counted_apart():
    s = _scraper()
    s._record_quote({
        "requisition_number": "A4", "summary": "GASKET MEULLER INDUSTRIES P35708",
        "open_date": "", "close_date": "",
    })
    s._record_open_bid(_bid("B1", "Overhaul of NOVA Bus Fleet"))

    assert (s._excluded_by_summary, s._excluded_by_title) == (1, 1)
    assert dict(s._exclusion_reasons) == {"GASKET": 1}
    assert dict(s._title_exclusion_reasons) == {"NOVA": 1}


# -- row extraction ---------------------------------------------------------


class _Cell:
    def __init__(self, text):
        self.text = text


class _Row:
    """Minimal stand-in for a Selenium <tr>."""

    def __init__(self, cells=(), text=""):
        self._cells = [_Cell(c) for c in cells]
        self.text = text

    def find_elements(self, _by, _tag):
        return self._cells


# The grid's real shape, taken from a live run (2026-08-05). The leading
# Commodity Codes column is what broke the first attempt: read positionally,
# "1313, 2151, 2153" became the bid number and every other field shifted one
# place left, putting the title into open_date — a varchar(64) — which aborted
# the insert and lost all 19 bids in the run.
LIVE_HEADERS = ["Commodity Codes", "Bid Number", "Title", "Open Date", "Close Date"]
LIVE_ROW = [
    "1313, 2151, 2153",
    "26-00210-ATXR",
    "Pantograph and Pole Catenary Interface Monitoring System (PPCIMS)",
    "08/04/2026",
    "09/04/2026",
]


def test_the_commodity_codes_column_is_skipped():
    """The reported bug, from the row that actually failed."""
    s = _scraper()
    record = s._bid_extractor(LIVE_HEADERS)(_Row(LIVE_ROW))

    assert record == {
        "bid_number": "26-00210-ATXR",
        "title": "Pantograph and Pole Catenary Interface Monitoring System (PPCIMS)",
        "open_date": "08/04/2026",
        "close_date": "09/04/2026",
    }
    # The commodity codes are not stored under any name.
    assert "1313, 2151, 2153" not in record.values()


def test_no_field_can_land_in_a_date_column():
    """The specific shape of the crash: a long title in a varchar(64)."""
    s = _scraper()
    record = s._bid_extractor(LIVE_HEADERS)(_Row(LIVE_ROW))
    for field in ("open_date", "close_date"):
        assert len(record[field]) <= 64, (field, record[field])


def test_columns_are_found_wherever_the_portal_puts_them():
    """Header-driven, so a moved or added column shifts nothing."""
    s = _scraper()
    headers = ["Bid Number", "Commodity Codes", "Close Date", "Title", "Open Date"]
    row = _Row(["26-1", "1313", "09/04/2026", "Rail Car Overhaul", "08/04/2026"])
    assert s._bid_extractor(headers)(row) == {
        "bid_number": "26-1",
        "title": "Rail Car Overhaul",
        "open_date": "08/04/2026",
        "close_date": "09/04/2026",
    }


def test_header_variants_are_recognised():
    for header in ("Bid #", "BID NO.", "Solicitation Number", "bid number"):
        assert _bid_columns([header, "Title"]).get("bid_number") == 0, header
    for header in ("Title", "Description", "Bid Title"):
        assert _bid_columns(["Bid Number", header]).get("title") == 1, header
    for header in ("Due Date", "Closing Date", "Close Date"):
        cols = _bid_columns(["Bid Number", "Title", header])
        assert cols.get("close_date") == 2, header


def test_a_partial_header_match_is_refused_outright():
    """Placing two fields correctly and guessing the rest is worse than
    falling back to a known-shape position."""
    assert _bid_columns(["Commodity Codes", "Open Date", "Close Date"]) == {}
    assert _bid_columns([]) == {}
    assert _bid_columns(["", "  "]) == {}


def test_unreadable_headers_fall_back_to_position_still_skipping_column_one():
    """No headers to map from — the layout is still known to lead with
    commodity codes, so position 0 is skipped rather than taken as the number."""
    s = _scraper()
    record = s._bid_extractor([])(_Row(LIVE_ROW))
    assert record["bid_number"] == "26-00210-ATXR"
    assert record["title"].startswith("Pantograph")


def test_a_bid_row_falls_back_to_splitting_the_row_text():
    s = _scraper()
    row = _Row(text="\n".join(LIVE_ROW))
    record = s._bid_extractor(LIVE_HEADERS)(row)
    assert record["bid_number"] == "26-00210-ATXR"
    assert record["title"].startswith("Pantograph")


def test_a_row_missing_its_trailing_columns_is_still_read():
    s = _scraper()
    record = s._bid_extractor(LIVE_HEADERS)(_Row(LIVE_ROW[:3]))
    assert record["bid_number"] == "26-00210-ATXR"
    assert record["open_date"] == "" and record["close_date"] == ""


# -- telling the Bid module apart from the Quotes form ----------------------


class _Driver:
    """Stub driver that records navigation and answers a fixed URL."""

    def __init__(self, url="", elements=None):
        self.current_url = url
        self.visited = []
        self._elements = elements or {}

    def get(self, url):
        self.visited.append(url)
        self.current_url = url

    def find_elements(self, _by, selector):
        return self._elements.get(selector, [])


def _on_form(s, url, elements=None):
    """Run _on_bids_form against `url` with the search form present."""
    s.driver = _Driver(url, elements)
    s._find = lambda *a, **k: object()   # the Search button rendered
    return s._on_bids_form()


def test_a_redirect_back_to_the_quotes_form_is_not_mistaken_for_bids():
    """Both forms carry a Search button, so the button alone cannot decide.

    Without this guard a wrong septa_bids_search_url that 302s back to Quotes
    would scrape the quotes grid a second time and file it as open bids.
    """
    s = _scraper()
    assert _on_form(s, QUOTES_URL) is False
    assert _on_form(s, "https://epsadmin.septa.org/vendor/quotes/search/") is False


def test_the_bid_module_is_recognised_by_its_url():
    s = _scraper()
    assert _on_form(s, BIDS_URL) is True


def test_an_unhelpful_url_falls_back_to_the_page_heading():
    s = _scraper()
    neutral = "https://epsadmin.septa.org/vendor/search/"
    assert _on_form(s, neutral) is False
    assert _on_form(s, neutral, {SEL["bids_heading_xpath"]: [object()]}) is True


def test_no_search_form_is_never_the_bid_module():
    s = _scraper()
    s.driver = _Driver(BIDS_URL)
    s._find = lambda *a, **k: None       # nothing rendered
    assert s._on_bids_form() is False


def test_a_wrong_configured_url_falls_back_to_the_menu_link():
    """The URL is a best guess in config; the menu link is the safety net."""
    s = _scraper()
    menu = "https://epsadmin.septa.org/vendor/solicitations/open-bids/"
    s.driver = _Driver(QUOTES_URL, {SEL["bids_link_xpath"]: [_Anchor(menu)]})
    # The configured URL lands on Quotes; only the menu link is the Bid module.
    s._on_bids_form = lambda *a: s.driver.current_url == menu

    assert s.navigate_to_open_bids() is True
    assert menu in s.driver.visited


def test_a_postback_menu_link_is_clicked_rather_than_followed():
    """Much of this ASP.NET portal navigates by __doPostBack, including the
    login submit — such an anchor has no URL to fetch, only a click."""
    s = _scraper()
    anchor = _Anchor("javascript:__doPostBack('ctl00$menu$bids','')")
    s.driver = _Driver(QUOTES_URL, {SEL["bids_link_xpath"]: [anchor]})
    clicked = []
    s._safe_click = lambda el: clicked.append(el) or True
    s._on_bids_form = lambda *a: bool(clicked)

    assert s.navigate_to_open_bids() is True
    assert clicked == [anchor]
    # The javascript: URL must never be handed to driver.get().
    assert not any(u.lower().startswith("javascript") for u in s.driver.visited)


def test_a_wrong_menu_candidate_returns_to_the_menu_before_the_next():
    """Otherwise the second candidate is looked for on whatever page the
    first one led to, where the menu may not be rendered at all."""
    s = _scraper()
    right = "https://epsadmin.septa.org/vendor/solicitations/open-bids/"
    wrong = "https://epsadmin.septa.org/vendor/profile/"
    s.driver = _Driver(QUOTES_URL, {SEL["bids_link_xpath"]: [_Anchor(wrong), _Anchor(right)]})
    s._on_bids_form = lambda *a: s.driver.current_url == right

    assert s.navigate_to_open_bids() is True
    # The configured URL is tried first and becomes the page the menu crawl
    # returns to, so: config, wrong candidate, back to config, right one.
    assert s.driver.visited == [BIDS_URL, wrong, BIDS_URL, right]


def test_an_unreachable_bid_module_fails_the_run():
    """Open Bids is now the *whole* run when it is the selected module.

    So a portal that has moved it has to fail loudly: completing with an empty
    sheet would make "0 open bids" and "we never opened the Bid module" look
    identical to whoever reads the output.
    """
    import pytest
    from selenium.common.exceptions import WebDriverException

    s = _scraper()
    s.driver = _Driver(QUOTES_URL)
    s._on_bids_form = lambda *a: False      # neither route works
    s.screenshot = lambda *a, **k: None

    assert s.navigate_to_open_bids() is False
    with pytest.raises(WebDriverException, match="could not reach the Open Bids"):
        s.scrape_open_bids()
    assert s._open_bids_reached is False


class _Anchor:
    def __init__(self, href):
        self._href = href

    def get_attribute(self, name):
        return self._href if name == "href" else None


# -- the optional date, on the Bids form ------------------------------------


def test_no_date_means_the_bids_filter_fields_are_never_touched():
    """"If no date inputs are provided, bypass the filter and just Search."""
    s = _scraper()
    touched = []
    s._fill_date = lambda *a, **k: touched.append(a) or True

    s.apply_date_filter("open bids")
    assert touched == [], "the date box was filled on a run that carried no date"


def test_a_date_is_typed_into_the_bids_form_too():
    from app.core import run_manager as rm

    run = rm.create_run("septa", Path("/tmp"))
    s = SeptaScraper(run["run_id"], OpenDateFilter(opens_from="2026-08-01"))
    filled = []
    s._fill_date = lambda xpath, value, label: filled.append((value, label)) or True

    s.apply_date_filter("open bids")
    assert len(filled) == 1
    value, label = filled[0]
    assert value == "08/01/2026", "the portal wants MM/DD/YYYY"
    assert "open bids" in label


# -- the workbook -----------------------------------------------------------


def test_the_workbook_carries_a_sheet_per_module():
    quotes = [{
        "requisition_number": "A4", "summary": "STEP SEPTA SPEC S-3544-9",
        "open_date": "08/01/2026", "close_date": "12/01/2026",
    }]
    bids = [_bid("24-00123", "Rail Car Truck Overhaul")]

    with TemporaryDirectory() as tmp:
        out = Path(tmp) / "septa.xlsx"
        export.generate_excel_from_records(quotes, out, bids)
        book = load_workbook(out)

        assert book.sheetnames == [export.QUOTES_SHEET, export.OPEN_BIDS_SHEET]
        assert [c.value for c in book[export.QUOTES_SHEET][1]] == [
            "Requisition Number", "Summary", "Open Date", "Close Date",
        ]
        assert [c.value for c in book[export.OPEN_BIDS_SHEET][1]] == [
            "Bid Number", "Bid Title", "Open Date", "Close Date",
        ]
        assert [c.value for c in book[export.OPEN_BIDS_SHEET][2]] == [
            "24-00123", "Rail Car Truck Overhaul", "08/01/2026", "12/01/2026",
        ]


def test_the_quotes_sheet_is_unchanged_when_there_are_no_open_bids():
    """The Bids pass must not alter what an existing quotes-only run produces."""
    quotes = [{
        "requisition_number": "A4", "summary": "STEP SEPTA SPEC S-3544-9",
        "open_date": "08/01/2026", "close_date": "12/01/2026",
    }]
    with TemporaryDirectory() as tmp:
        out = Path(tmp) / "septa.xlsx"
        # No open_bids argument at all — the pre-existing call shape.
        assert export.generate_excel_from_records(quotes, out) == 1
        book = load_workbook(out)

        sheet = book[export.QUOTES_SHEET]
        assert sheet.max_row == 2
        assert [c.value for c in sheet[2]] == [
            "A4", "STEP SEPTA SPEC S-3544-9", "08/01/2026", "12/01/2026",
        ]
        # Present but empty: "the Bids pass found nothing" has to be
        # distinguishable from "this workbook predates the Bids pass".
        assert book[export.OPEN_BIDS_SHEET].max_row == 1


# -- one bad field must not cost the whole run ------------------------------


def test_an_oversized_value_is_trimmed_rather_than_aborting_the_insert():
    """Postgres raises for the whole statement, and the batch is one
    transaction — so an untrimmed field loses every row in the run, which is
    how a 19-bid scrape stored nothing."""
    long_title = "Pantograph and Pole Catenary Interface Monitoring System " * 3
    assert len(long_title) > 64

    assert export._fit("run", "open_date", long_title) == long_title[:64]
    assert export._fit("run", "close_date", long_title) == long_title[:64]
    # Fields with room, and non-strings, pass through untouched.
    assert export._fit("run", "title", long_title) == long_title
    assert export._fit("run", "open_date", "08/04/2026") == "08/04/2026"
    assert export._fit("run", "open_date", None) is None


def test_a_db_failure_makes_the_export_read_the_fallback_sheet():
    """Regenerating a run whose DB save failed *succeeds* and returns an empty
    workbook, silently replacing the only copy of the rows."""
    from app.core import exports

    with TemporaryDirectory() as tmp:
        on_disk = Path(tmp) / "Septa_(all open bids).xlsx"
        export.generate_excel_from_records([], on_disk, [_bid("26-1", "Rail Car Overhaul")])

        run = {"run_id": "r1", "scraper": "septa", "excel_path": str(on_disk)}

        # Healthy run: regenerated from the DB, whatever is on disk.
        called = []
        original = exports.importlib.import_module

        # A DB-failed run must not even attempt regeneration.
        failed = {**run, "db_save_failed": True}
        exports.importlib.import_module = lambda *a, **k: called.append(a) or original(*a, **k)
        try:
            payload = exports.excel_bytes(failed)
        finally:
            exports.importlib.import_module = original

        assert called == [], "regenerated over the fallback sheet"
        assert payload is not None
        data, name = payload
        assert data == on_disk.read_bytes()
        assert name == on_disk.name


def test_a_bid_without_a_number_is_not_written():
    with TemporaryDirectory() as tmp:
        out = Path(tmp) / "septa.xlsx"
        export.generate_excel_from_records([], out, [_bid("", "No number")])
        book = load_workbook(out)
        assert book[export.OPEN_BIDS_SHEET].max_row == 1


if __name__ == "__main__":
    tests = [(name, fn) for name, fn in sorted(globals().items())
             if name.startswith("test_") and callable(fn)]
    failures = 0
    for name, fn in tests:
        try:
            fn()
            print(f"PASS  {name}")
        except AssertionError as exc:
            print(f"FAIL  {name}: {exc}")
            failures += 1
        except Exception as exc:  # noqa: BLE001 — report, don't abort the suite
            print(f"ERROR {name}: {exc.__class__.__name__}: {exc}")
            failures += 1
    print(f"\n{len(tests) - failures}/{len(tests)} passed")
    sys.exit(1 if failures else 0)
