"""The broad sweep: "Run all member agency bids", end to end without a browser.

Three things this pins down, because all three are what the mode *is* rather
than details of it:

* **Nothing is typed into the search box.** A sweep that searched a keyword
  would return a slice of the member agency list and look entirely normal doing
  it, so "no term was ever searched" is asserted directly.
* **The Member Agency Bids group is selected**, and the sidebar filters are the
  only narrowing applied.
* **One spreadsheet, named for the day, delivered without a ZIP.** The naming
  and the bare-sheet routing are the deliverable, not packaging trivia — a
  correct scrape filed under the wrong name is a failed handover.

    server/.venv/bin/python -m pytest server/tests/test_bidnet_member_agencies.py
"""

import os
import sys
from datetime import date
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.core import exports, run_manager  # noqa: E402
from app.scrapers.bidnet import member_agencies, storage  # noqa: E402
from app.scrapers.bidnet.member_agencies import (  # noqa: E402
    MAX_AGENCY_LENGTH,
    SWEEP_LABEL,
    UNKNOWN_AGENCY,
    MemberAgencySweepScraper,
)
from app.scrapers.bidnet.models import BidnetBid  # noqa: E402
from app.scrapers.bidnet.scraper import LinkHarvest  # noqa: E402

# Three pages' worth of results, with /2 repeated the way the portal repeats a
# row when the list shifts under pagination.
LINKS = [f"https://b/{n}" for n in (1, 2, 3, 4, 2, 5)]
AGENCIES = {
    "1": "City of Dutchess",
    "2": "City of Dutchess",
    "3": "Sullivan County",
    "4": "Sullivan County",
    "5": "",  # the detail page labels no agency
}


@pytest.fixture
def swept(monkeypatch, tmp_path):
    """A sweep scraper with the browser replaced but `run()`'s own flow intact."""
    folder = tmp_path / "sweep"
    folder.mkdir()
    run = run_manager.create_run(
        "bidnet",
        folder,
        {"member_agency_sweep": True, "excel_only": True, "niche_label": SWEEP_LABEL},
    )
    instance = MemberAgencySweepScraper(run["run_id"])

    calls: dict[str, list] = {"searched": [], "grouped": [], "opened": []}

    def search(keyword):
        calls["searched"].append(keyword)

    def open_filtered_session():
        # What the real one does: an empty-keyword search, then the sidebar.
        search("")

    def process_bid(link):
        calls["opened"].append(link)
        bid = link.rsplit("/", 1)[-1]
        return {
            "reference_number": f"RFP-{bid}",
            "title": f"Solicitation {bid}",
            "detail_url": link,
            "status": "OK",
        }

    for name, value in {
        "start_driver": lambda *a, **k: None,
        "login": lambda: None,
        "ensure_logged_in": lambda: None,
        "search": search,
        "open_filtered_session": open_filtered_session,
        "_ensure_filters_live": lambda: None,
        "_ensure_first_result_page": lambda: True,
        "filter_member_agency": lambda: calls["grouped"].append(True),
        "confirm_filters_active": lambda keyword="": None,
        "result_count": lambda: len(set(LINKS)),
        "collect_links": lambda: LinkHarvest(
            links=list(LINKS), rows_detected=len(LINKS), rows_parsed=len(LINKS)
        ),
        "process_bid": process_bid,
        "_extract_agency": lambda link="": AGENCIES[link.rsplit("/", 1)[-1]],
        "_save_run_row": lambda: None,
        "cleanup": lambda: None,
    }.items():
        monkeypatch.setattr(instance, name, value)

    saved: list[list[dict]] = []
    monkeypatch.setattr(
        member_agencies.export,
        "save_bids",
        lambda run, records: (saved.append(list(records)), len(records))[1],
    )
    monkeypatch.setattr(member_agencies, "archive_run", lambda run_id: None)
    monkeypatch.setattr(member_agencies, "notify_scrape_completion", lambda *a, **k: None)
    return instance, calls, saved


# -- the flow ---------------------------------------------------------------


def test_the_sweep_searches_no_keyword_at_all(swept):
    """The defining difference from a niche run. Typing anything into the box is
    what narrows the portal's list, so the way to see all of it is to search
    nothing."""
    instance, calls, _ = swept
    instance.run()

    assert calls["searched"] == [""], calls["searched"]
    assert instance.search_terms == []


def test_the_member_agency_group_is_selected(swept):
    instance, calls, _ = swept
    instance.run()
    assert calls["grouped"], "the Member Agency Bids group was never clicked"


def test_every_solicitation_is_opened_once(swept):
    """Deduplicated by solicitation id: /2 appears twice in the harvest."""
    instance, calls, _ = swept
    instance.run()

    assert calls["opened"] == ["https://b/1", "https://b/2", "https://b/3",
                               "https://b/4", "https://b/5"], calls["opened"]


def test_the_bid_cap_stops_a_runaway_sweep(swept, monkeypatch):
    """An unfiltered member agency list is thousands of detail page loads. The
    cap is a guard, and hitting it must be said out loud rather than silently
    truncating the sheet."""
    instance, calls, _ = swept
    monkeypatch.setattr(member_agencies, "MAX_SWEEP_BIDS", 2)
    instance.run()

    assert len(calls["opened"]) == 2, calls["opened"]
    warnings = run_manager.get_run(instance.run_id)["warnings"]
    assert any("MAX" in w or "more than this sweep" in w for w in warnings), warnings


# -- the agency column ------------------------------------------------------


def test_the_niche_column_carries_the_issuing_agency(swept):
    """A sweep belongs to no niche, so that column names the agency instead —
    same position in the sheet, same question answered."""
    instance, _, saved = swept
    instance.run()

    assert saved, "save_bids was never called"
    by_ref = {r["reference_number"]: r["niche"] for r in saved[0]}
    assert by_ref["RFP-1"] == "City of Dutchess"
    assert by_ref["RFP-3"] == "Sullivan County"


def test_a_bid_with_no_labelled_agency_still_gets_a_heading(swept):
    """Never blank: an empty cell reads as a broken export rather than as a
    portal that did not say."""
    instance, _, saved = swept
    instance.run()
    assert {r["reference_number"]: r["niche"] for r in saved[0]}["RFP-5"] == UNKNOWN_AGENCY


def test_a_long_agency_name_cannot_fail_the_runs_database_save(swept, monkeypatch):
    """A real sweep died here: "City and County of Denver Climate Action,
    Sustainability & Resiliency" is 68 characters, the column was 64, and
    Postgres fails the INSERT rather than truncating — which rolled back the
    **whole run's** save, not just that row. So the column is wider now *and*
    what goes into it is capped."""
    instance, _, saved = swept
    monkeypatch.setattr(instance, "_extract_agency", lambda link="": "Department of " + "Very " * 80)
    instance.run()

    written = {r["niche"] for r in saved[0]}
    assert written, "nothing was saved"
    assert all(len(name) <= MAX_AGENCY_LENGTH for name in written), written


def test_the_niche_column_can_hold_a_real_agency_name():
    """The cap is the belt; the column width is the braces. Asserted against the
    model rather than the migration so a fresh `create_all` database and a
    migrated one cannot disagree."""
    longest = "City and County of Denver Climate Action, Sustainability & Resiliency"
    width = BidnetBid.__table__.c.niche.type.length
    assert width >= len(longest), width
    assert width > MAX_AGENCY_LENGTH, "a capped name must still fit the column"


def test_no_keyword_is_credited_with_finding_the_bids(swept):
    instance, _, saved = swept
    instance.run()
    assert all(r["matched_keyword"] == "" for r in saved[0])


def test_the_run_reports_its_agency_breakdown(swept):
    """The sweep's whole point is that it spans agencies; a run reporting only a
    bid count cannot be checked against the portal by anyone."""
    instance, _, _ = swept
    instance.run()

    run = run_manager.get_run(instance.run_id)
    assert run["agency_total"] == 3
    assert run["agency_breakdown"]["City of Dutchess"] == 2
    assert run["agency_breakdown"]["Sullivan County"] == 2


# -- the deliverable --------------------------------------------------------


def test_the_sheet_is_named_for_the_day_it_swept(swept):
    instance, _, _ = swept
    instance.run()

    expected = f"bidnet_member_agencie_{date.today().isoformat()}.xlsx"
    assert instance.excel_path.name == expected
    assert instance.excel_path.is_file(), "the sheet was never written"
    assert run_manager.get_run(instance.run_id)["excel_name"] == expected


def test_the_sheet_holds_every_swept_bid(swept):
    from openpyxl import load_workbook

    instance, _, saved = swept
    instance.run()

    sheet = load_workbook(instance.excel_path).active
    assert sheet.max_row == len(saved[0]) + 1, "header plus one row per bid"


def test_the_sweep_delivers_a_bare_sheet_not_a_zip():
    """No ZIP: a ZIP exists to hold several files, and this mode produces one.

    BidNet's other flows *do* ship ZIPs, so this is a property of the run rather
    than of the portal — which is exactly what `is_excel_only` exists to say."""
    assert exports.is_excel_only({"scraper": "bidnet", "excel_only": True})
    assert not exports.is_excel_only({"scraper": "bidnet", "session_root": "/x"})


def test_the_download_keeps_the_swept_name():
    """`excel_name` is what names the file the client saves, and a correct scrape
    filed under `Bidnetdirect_(...).xlsx` is a failed handover."""
    name = storage.member_agency_excel_name()
    assert exports.excel_name(
        {"scraper": "bidnet", "download_name": name, "niche_label": SWEEP_LABEL}
    ) == name


# -- the endpoint -----------------------------------------------------------


def test_the_endpoint_queues_a_sweep_with_no_niche(monkeypatch, tmp_path):
    """The niche dropdown is irrelevant to this mode: the sidebar filters are
    the only input the request carries."""
    from fastapi.testclient import TestClient

    import main
    from app.core import jobs

    submitted: list[tuple] = []
    monkeypatch.setattr(jobs, "submit", lambda run_id, fn, *a: submitted.append((run_id, fn, a)))
    monkeypatch.setattr(
        storage, "member_agency_folder", lambda when=None: _made(tmp_path / "sweep")
    )

    with TestClient(main.app) as client:
        response = client.post("/bidnet/scrape/member-agencies", json={"filters": {}})
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["excel_name"] == storage.member_agency_excel_name()

    assert submitted, "nothing was queued"
    run_id, fn, _ = submitted[0]
    assert fn is member_agencies.execute_member_agency_sweep
    run = run_manager.get_run(run_id)
    assert run["excel_only"] is True
    assert run["member_agency_sweep"] is True
    assert run["download_name"] == storage.member_agency_excel_name()
    assert not run.get("niche"), "a sweep belongs to no niche"


def _made(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path
