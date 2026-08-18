"""The hub's one Excel look, shared by every portal's export.

Every portal writes its own sheet — different columns, different sheet counts,
some with extra per-row tinting or agency banners — but they all present the
same way: a navy header row in bold white, centred, bordered, with columns wide
enough that nothing is cut off. That styling lives here so it is described once
and cannot drift portal to portal.

Typical use, for a portal whose sheet is a plain header row plus data rows:

    from app.core import excel_style

    workbook, sheet = excel_style.new_workbook("BidNet Bids")
    excel_style.write_table(sheet, headers, rows)
    workbook.save(str(out_path))

`write_table` appends the header row and the data rows, then formats. A portal
that builds its rows in an unusual shape (EMMA's dynamic detail columns, SAM's
decision tinting, RideMetro's repeating per-agency blocks) appends its own rows
and calls `format_excel_headers` / `style_header_row` / `autofit_columns`
directly — the pieces compose, so no portal has to restate the look to get an
extra behaviour.

Nothing here touches values other than to make them writable: control characters
Excel rejects are stripped, and a cell over Excel's hard 32,767-character limit
is truncated with a marker. Everything else — dates, numbers, blanks — is passed
through as the portal produced it.
"""

from __future__ import annotations

import logging
from typing import Any, Callable, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# -- the standard ------------------------------------------------------------

#: Header fill — Office's "Blue, Accent 1, Darker 50%". One colour for every
#: portal's header row; change it here and every export follows.
HEADER_COLOR = "1F4E78"
#: Thin cell rule, light enough not to compete with the text.
BORDER_COLOR = "D5DDE5"
#: Header row height, in points.
HEADER_HEIGHT = 26

HEADER_FILL = PatternFill("solid", fgColor=HEADER_COLOR)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_EDGE = Side(style="thin", color=BORDER_COLOR)
CELL_BORDER = Border(left=_EDGE, right=_EDGE, top=_EDGE, bottom=_EDGE)

# Column auto-fit bounds. The padding keeps the widest value clear of the column
# edge; the cap stops one long description from pushing every other column off
# the screen (that cell wraps instead).
WIDTH_PADDING = 4
MIN_WIDTH = 10
MAX_WIDTH = 60

# Excel refuses a cell over 32,767 characters. Long descriptions and extracted
# document text routinely exceed it; the untruncated value is in the DB.
MAX_CELL_CHARS = 32000
TRUNCATION_MARKER = " …[truncated]"


# -- cell values -------------------------------------------------------------


def sanitize_cell(value: Any) -> Any:
    """Make `value` writable by Excel, changing nothing else.

    Strips the control characters openpyxl refuses (portal text carries them
    surprisingly often) and truncates past Excel's hard cell limit. Non-strings
    — dates, numbers, None — are returned untouched so their cell keeps its
    type.
    """
    if not isinstance(value, str):
        return value
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(value) > MAX_CELL_CHARS:
        value = value[:MAX_CELL_CHARS] + TRUNCATION_MARKER
    return value


# -- styling -----------------------------------------------------------------


def style_header_row(
    sheet: Worksheet,
    row: int = 1,
    first_column: int = 1,
    last_column: int | None = None,
) -> None:
    """Apply the standard header look to one row of `sheet`.

    Defaults to row 1 across every column that row uses — the shape of a plain
    export. `row` / `first_column` / `last_column` are for sheets that repeat
    headers further down (RideMetro writes one header row per agency block).
    """
    last_column = last_column or sheet.max_column
    for column in range(first_column, last_column + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    sheet.row_dimensions[row].height = HEADER_HEIGHT


def autofit_columns(
    sheet: Worksheet,
    min_width: int = MIN_WIDTH,
    max_width: int = MAX_WIDTH,
    padding: int = WIDTH_PADDING,
) -> None:
    """Size every column to its longest value, plus padding.

    Cells that anchor a merged range are ignored: a full-width banner merged
    across the sheet would otherwise set column A to the banner's own length and
    leave every other column at its minimum.
    """
    merged_anchors = {str(rng).split(":")[0] for rng in sheet.merged_cells.ranges}
    # Indexed rather than taken off the cells: a merged range fills its span with
    # MergedCell placeholders, which carry no column letter of their own.
    for index, column in enumerate(sheet.iter_cols(), start=1):
        longest = 0
        for cell in column:
            if cell.value is None or cell.coordinate in merged_anchors:
                continue
            # A wrapped multi-line value is only ever as wide as its widest line.
            for line in str(cell.value).splitlines() or [""]:
                longest = max(longest, len(line))
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(longest + padding, min_width), max_width
        )


def format_excel_headers(
    sheet: Worksheet,
    header_row: int = 1,
    autofit: bool = True,
    freeze: bool = True,
) -> None:
    """The one call a finished sheet needs: style its header row and fit its
    columns.

    `freeze` pins the header row so it stays visible while scrolling; pass False
    for a sheet whose headers repeat (there is no single row to pin).
    """
    if sheet.max_row >= header_row:
        style_header_row(sheet, header_row)
    if autofit:
        autofit_columns(sheet)
    if freeze:
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1).coordinate


# -- whole sheets ------------------------------------------------------------


def new_workbook(title: str) -> tuple[Workbook, Worksheet]:
    """A workbook and its first sheet, named. Saves every export the same two
    lines of boilerplate."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    return workbook, sheet


#: How a row can be tinted, by what the tint *means* rather than by its colour.
#: A caller returns one of these names and the palette is decided here, so two
#: portals cannot end up with two different reds for the same idea.
#:
#: Pastel fills with dark type: saturated colour across twenty columns makes the
#: text harder to read, and the point of the tint is to let someone skim the
#: sheet rather than to decorate it. Nothing is defined for "clean" — an
#: in-scope row is left exactly as written, which is what makes the tinted ones
#: stand out at all.
ROW_STYLES: dict[str, tuple[PatternFill, Font]] = {
    # Out of scope — the matrix rejected it, or it sits in an excluded niche.
    "reject": (PatternFill("solid", fgColor="FADBD8"), Font(color="78281F", bold=True)),
    # Undecided — a person still has to look at this one.
    "review": (PatternFill("solid", fgColor="FCF3CF"), Font(color="7D6608", bold=True)),
}


def write_table(
    sheet: Worksheet,
    headers: Sequence[Any],
    rows: Iterable[Sequence[Any]],
    freeze: bool = True,
    row_style: Callable[[Sequence[Any]], str | None] | None = None,
) -> int:
    """Write a header row and `rows` beneath it, formatted. Returns the row count.

    This is the whole body of a plain portal export: the caller decides the
    columns and produces the values, and everything about how the result looks
    is decided here.

    `row_style` is an optional function over a row's values returning a key of
    `ROW_STYLES` — or None to leave the row clean. It lets a portal tint by what
    a row *means* (out of scope, needs a person) without deciding what colour
    that is. Omitted, nothing is tinted and the output is what it was before.
    """
    sheet.append([sanitize_cell(header) for header in headers])
    written = 0
    tinted: dict[str, int] = {}
    for row in rows:
        values = [sanitize_cell(value) for value in row]
        sheet.append(values)
        written += 1
        style = row_style(values) if row_style is not None else None
        if style is None:
            continue
        fill, font = ROW_STYLES[style]
        # `sheet.max_row` rather than a counter: append() is what decides where
        # the row landed, and reading it back cannot drift from that.
        for cell in sheet[sheet.max_row]:
            cell.fill = fill
            cell.font = font
        tinted[style] = tinted.get(style, 0) + 1
    # After the body, so the header keeps its own fill: a style function that
    # matched every row would otherwise leave the header tinted too.
    format_excel_headers(sheet, freeze=freeze)
    if tinted:
        logger.info(
            "%d of %d row(s) tinted (%s)", sum(tinted.values()), written,
            ", ".join(f"{count} {name}" for name, count in sorted(tinted.items())),
        )
    return written
