"""Unison persistence: store scraped buyer requests in the DB and build the
per-run Excel from the DB (openpyxl). Same DB-first-with-fallback pattern as the
other hub portals."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.styles import PatternFill
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core import excel_style
from app.db import SessionLocal
from app.scrapers.unison.models import EXCEL_COLUMNS, UnisonRequest, UnisonRun

logger = logging.getLogger(__name__)

# Columns on UnisonRequest that a scraped record can fill directly. The
# evaluator's working-out (requirement_type, rule, location) is stored but not
# exported — see EXCEL_COLUMNS in models.py.
_BID_FIELDS = {
    "buyer_number", "bid_upload_count", "buyer_description", "buyer", "end_date",
    "detail_url",
    # General Buy Information
    "solicitation_number", "category", "subcategory", "naics", "naics_size_standard",
    "sam_contract_opportunity", "set_aside", "end_time", "seller_question_deadline",
    "delivery", "repost_reason",
    # Shipping Information
    "shipping_city", "shipping_state", "shipping_zip",
    # line items, attachments, evaluation
    "line_item_count", "seller_attachments_required", "attachment_count",
    "decision", "reason", "requirement_type", "rule", "location", "requirement_hinted",
}

# JSONB columns, which need the record's nested values made serialisable.
_JSON_FIELDS = {"line_items", "attachments", "detail_sections"}


def _parse_dt(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(k): _jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(v) for v in value]
    return str(value)


def _run_values(run: dict[str, Any]) -> dict[str, Any]:
    return {
        "run_id": run["run_id"],
        "status": run.get("status"),
        "started_at": _parse_dt(run.get("started_at")),
        "finished_at": _parse_dt(run.get("finished_at")),
        "search": run.get("search"),
        "filter_id": run.get("filter_id"),
        "filter_label": run.get("filter_label"),
        "pages_scraped": run.get("pages_scraped", 0),
        "bids_found": run.get("bids_found", 0),
        "documents_downloaded": run.get("documents_downloaded", 0),
        "folder": run.get("folder"),
        "excel_path": run.get("excel_path"),
    }


def _upsert_run(session, run: dict[str, Any]) -> None:
    values = _run_values(run)
    stmt = pg_insert(UnisonRun).values(**values)
    stmt = stmt.on_conflict_do_update(
        index_elements=[UnisonRun.run_id],
        set_={k: v for k, v in values.items() if k != "run_id"},
    )
    session.execute(stmt)


def save_run(run: dict[str, Any]) -> None:
    session = SessionLocal()
    try:
        _upsert_run(session, run)
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def save_bids(run: dict[str, Any], records: list[dict[str, Any]]) -> int:
    """Upsert a run's requests into unison_requests, deduped by buyer_number."""
    session = SessionLocal()
    try:
        _upsert_run(session, run)

        stored = 0
        seen: set[str] = set()
        for record in records:
            values: dict[str, Any] = {k: (record.get(k) or None) for k in _BID_FIELDS}
            # Counters and flags: a real 0/False, not NULL from the `or None`.
            for field in ("line_item_count", "attachment_count"):
                values[field] = int(record.get(field) or 0)
            values["requirement_hinted"] = bool(record.get("requirement_hinted"))
            for field in _JSON_FIELDS:
                values[field] = _jsonable(record.get(field) or ([] if field != "detail_sections" else {}))
            values["run_id"] = run["run_id"]
            values["raw_data"] = _jsonable(record)

            buyer_number = values.get("buyer_number")
            if buyer_number and buyer_number in seen:
                continue
            if buyer_number:
                seen.add(buyer_number)
                stmt = pg_insert(UnisonRequest).values(**values)
                stmt = stmt.on_conflict_do_update(
                    constraint="uq_unison_run_buyer",
                    set_={k: v for k, v in values.items() if k not in ("run_id", "buyer_number")},
                )
                session.execute(stmt)
            else:
                session.add(UnisonRequest(**values))
            stored += 1

        session.commit()
        logger.info("[run %s] stored %d Unison rows in DB", run.get("run_id"), stored)
        return stored
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _rows_for_run(run_id: str) -> list[UnisonRequest]:
    session = SessionLocal()
    try:
        return session.execute(
            select(UnisonRequest).where(UnisonRequest.run_id == run_id).order_by(UnisonRequest.id)
        ).scalars().all()
    finally:
        session.close()


# The evaluator's verdict in the row colour, so a long sheet sorts itself out at
# a glance:
#
#   REJECT         light red  — read as "skip"; a pastel is enough, since the
#                               point is to let the eye slide past these
#   MANUAL_REVIEW  amber      — the only rows that need a person to do
#                               something, so this is the strongest tint on the
#                               sheet. SAM uses a pale cream (#FFF2CC) here,
#                               which is all but invisible on a bright screen;
#                               a solid amber is what makes a handful of rows
#                               findable in a hundred.
#   PURSUE         none       — left plain, so the shortlist reads as the
#                               sheet's normal state rather than another colour
#
# Black text stays comfortably legible on both fills.
_REJECT_FILL = PatternFill("solid", fgColor="FFCCCC")
_REVIEW_FILL = PatternFill("solid", fgColor="FFD966")
_DECISION_FILLS = {"REJECT": _REJECT_FILL, "MANUAL_REVIEW": _REVIEW_FILL}


def _tint_by_decision(sheet) -> None:
    """Colour each data row by the verdict in its Decision column."""
    headers = [cell.value for cell in sheet[1]]
    if "Decision" not in headers:
        return
    column = headers.index("Decision") + 1
    for row_index in range(2, sheet.max_row + 1):
        fill = _DECISION_FILLS.get(str(sheet.cell(row=row_index, column=column).value or "").upper())
        if fill is not None:
            for cell in sheet[row_index]:
                cell.fill = fill


def generate_excel(run_id: str, out_path: str | Path) -> int:
    rows = _rows_for_run(run_id)
    workbook, sheet = excel_style.new_workbook("Unison Requests")
    excel_style.write_table(
        sheet,
        [header for _, header in EXCEL_COLUMNS],
        ([getattr(row, attr, None) for attr, _ in EXCEL_COLUMNS] for row in rows),
    )
    _tint_by_decision(sheet)
    workbook.save(str(out_path))
    logger.info("[run %s] wrote %d Unison rows to %s", run_id, len(rows), out_path)
    return len(rows)


def _record_cell(record: dict[str, Any], attr: str) -> Any:
    """One export cell from an in-memory record.

    `attachment_names` is a derived column — a property on the model, and here
    the same join over the record's downloaded files.
    """
    if attr == "attachment_names":
        return ", ".join(
            str(a.get("name") or "")
            for a in (record.get("attachments") or []) if a.get("name")
        )
    return record.get(attr)


def generate_excel_from_records(records: list[dict[str, Any]], out_path: str | Path) -> int:
    workbook, sheet = excel_style.new_workbook("Unison Requests")
    count = excel_style.write_table(
        sheet,
        [header for _, header in EXCEL_COLUMNS],
        (
            [_record_cell(record, attr) for attr, _ in EXCEL_COLUMNS]
            for record in records
            if record.get("buyer_number")
        ),
    )
    _tint_by_decision(sheet)
    workbook.save(str(out_path))
    return count
