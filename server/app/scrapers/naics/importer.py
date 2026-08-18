"""Reading a list of NAICS codes out of a spreadsheet somebody sent.

The file is whatever the client had to hand: a CSV, a modern `.xlsx`, or a
legacy `.xls`. What comes back is a list of six-digit codes that exist in the
catalogue, de-duplicated, in the order they appeared — plus an account of
everything that was dropped, because a run that quietly searched forty codes
when the file held forty-five is worse than one that says which five it could
not use.

**On the leading-zero rule.** Spreadsheet software does strip leading zeros, and
for many code systems padding them back is exactly right. It is wrong for NAICS.
No NAICS code begins with a zero — the catalogue's 1,012 codes all start 11
through 92 — so padding `54151` to `054151` invents a code that has never
existed, and a run filtered on it finds nothing.

`54151` is not a damaged six-digit code. It is a real *five-digit* industry
group, and what a reader means by writing it is "everything under this": 541511,
541512, 541513, 541519. So a short entry is **expanded to its children** rather
than padded, and if it matches nothing it is reported rather than guessed at.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Iterable

logger = logging.getLogger(__name__)

#: Header labels that name the column of codes, compared after normalising to
#: lowercase alphanumerics — so "NAICS Code", "naics_code" and "NAICS-Codes"
#: are one label.
HEADER_NAMES = frozenset({"naics", "naicscode", "naicscodes", "code", "codes"})

#: A NAICS code is 2–6 digits. Six is a real code; anything shorter is a level
#: above it and is expanded to the six-digit codes beneath.
_DIGITS = re.compile(r"\d+")
MIN_LENGTH, FULL_LENGTH = 2, 6

#: Rows scanned for a header before falling back to the first column. A header
#: is normally row 1; a couple of title/blank rows above it are common enough in
#: a spreadsheet somebody has formatted by hand.
HEADER_SCAN_ROWS = 5


@dataclass
class ImportResult:
    """What a file yielded, and what it cost."""

    codes: list[str] = field(default_factory=list)
    #: Entries that could not be used, as `value -> why`, in file order.
    skipped: list[tuple[str, str]] = field(default_factory=list)
    #: Short entries that were expanded, as `entry -> how many children`.
    expanded: list[tuple[str, int]] = field(default_factory=list)
    duplicates: int = 0
    #: Which column the codes were read from, for the summary line.
    source: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "codes": self.codes,
            "count": len(self.codes),
            "skipped": [{"value": value, "reason": why} for value, why in self.skipped],
            "expanded": [{"entry": entry, "codes": n} for entry, n in self.expanded],
            "duplicates": self.duplicates,
            "source": self.source,
        }


def _normalise_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


#: A whole number wearing a decimal tail: "541511.0". Excel stores codes as
#: numbers and writes them back this way, in the cell *and* in a CSV export, so
#: it has to be handled for text as well as for floats — otherwise the digits
#: run together into a seven-digit value and the code is thrown away as junk.
_TRAILING_ZEROS = re.compile(r"^(\d+)\.0+$")


def _cell_text(value: Any) -> str:
    """A cell as text, without the decimal tail a numeric cell arrives with."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    tail = _TRAILING_ZEROS.match(text)
    return tail.group(1) if tail else text


def read_rows(data: bytes, filename: str) -> list[list[Any]]:
    """The file's cells, as a list of rows. Raises ValueError on a file we cannot open."""
    suffix = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""

    if suffix == "csv" or (not suffix and b"," in data[:4096]):
        # Decoded forgivingly: these files come from Windows Excel as often as
        # not, and a stray byte in a title row should not lose the whole import.
        text = data.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [row for row in csv.reader(io.StringIO(text), dialect)]

    if suffix == "xlsx":
        from openpyxl import load_workbook

        # read_only keeps a large sheet out of memory; data_only takes the
        # cached value of a formula rather than the formula text.
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    if suffix == "xls":
        import xlrd

        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_index(0)
        return [sheet.row_values(index) for index in range(sheet.nrows)]

    raise ValueError(
        f"{filename!r} is not a spreadsheet this can read — send a .csv, .xlsx or .xls file"
    )


def find_column(rows: list[list[Any]]) -> tuple[int, str]:
    """`(column index, what it was found by)`.

    A named header wins wherever it is in the first few rows. With no header the
    first column is used, which is what a file that is just a list of codes
    looks like.
    """
    for row in rows[:HEADER_SCAN_ROWS]:
        for index, cell in enumerate(row or []):
            if _normalise_header(cell) in HEADER_NAMES:
                return index, f"the {str(cell).strip()!r} column"
    return 0, "the first column (no NAICS header found)"


def _header_row_index(rows: list[list[Any]], column: int) -> int:
    """Which row the header sat on, so its own text is not read as a code."""
    for position, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        if _normalise_header((row or [None] * (column + 1))[column]
                             if len(row or []) > column else "") in HEADER_NAMES:
            return position
    return -1


def clean_entry(raw: str) -> tuple[str, str]:
    """`(digits, reason it was rejected)` — one of the two is always empty.

    Everything non-numeric is stripped first: files arrive with "541511 —
    Custom Computer Programming", with a stray apostrophe from Excel's
    text-format trick, and with codes quoted.
    """
    text = _cell_text(raw)
    if not text:
        return "", ""            # a blank cell is not an error, just nothing
    digits = "".join(_DIGITS.findall(text))
    if not digits:
        return "", "no digits in it"
    if len(digits) > FULL_LENGTH:
        # Two codes run together, or a code with a year stuck to it. Guessing
        # which six of the digits were meant is not something to do silently.
        return "", f"{len(digits)} digits — too long to be a NAICS code"
    if len(digits) < MIN_LENGTH:
        return "", "too short to be a NAICS code"
    return digits, ""


def expand(prefix: str, catalogue: Iterable[str]) -> list[str]:
    """Every six-digit code beneath a shorter one, in catalogue order."""
    return [code for code in catalogue if code.startswith(prefix)]


def parse(
    data: bytes, filename: str, catalogue: Iterable[str] | None = None
) -> ImportResult:
    """Read a spreadsheet into a list of usable six-digit NAICS codes.

    `catalogue` is every code the reference table holds. Given it, a short entry
    is expanded to its children and a six-digit code that does not exist is
    reported rather than passed to the portal, where it would silently match
    nothing. Without it, six-digit entries are taken at face value.
    """
    rows = read_rows(data, filename)
    result = ImportResult()
    if not rows:
        result.source = "an empty file"
        return result

    column, found_by = find_column(rows)
    result.source = found_by
    header_row = _header_row_index(rows, column)

    known = list(catalogue) if catalogue is not None else []
    known_set = set(known)
    seen: set[str] = set()

    for position, row in enumerate(rows):
        if position == header_row:
            continue
        cell = (row or [])[column] if len(row or []) > column else ""
        digits, why = clean_entry(cell)
        if why:
            result.skipped.append((_cell_text(cell), why))
            continue
        if not digits:
            continue

        if len(digits) < FULL_LENGTH:
            children = expand(digits, known) if known else []
            if not children:
                result.skipped.append((
                    digits,
                    f"{len(digits)} digits, and no six-digit code in the catalogue "
                    f"begins with it",
                ))
                continue
            result.expanded.append((digits, len(children)))
            for code in children:
                if code in seen:
                    result.duplicates += 1
                    continue
                seen.add(code)
                result.codes.append(code)
            continue

        if known_set and digits not in known_set:
            result.skipped.append((digits, "not a code in the NAICS catalogue"))
            continue
        if digits in seen:
            result.duplicates += 1
            continue
        seen.add(digits)
        result.codes.append(digits)

    logger.info(
        "[NAICS IMPORT]: %s — %d code(s) from %s%s%s",
        filename, len(result.codes), result.source,
        f", {result.duplicates} duplicate(s) removed" if result.duplicates else "",
        f", {len(result.skipped)} entry/entries skipped" if result.skipped else "",
    )
    return result
