"""The agency-grouped report: one sheet, one block per agency.

    ┌──────────────────────────────────────────────────────────────┐
    │ CITY OF DALLAS                                               │  banner (merged, bold, filled)
    ├────────┬────────┬─────────┬───────────┬─────────┬─────┬──────┤
    │ Status │ Ref #  │ Project │ Department│ Closing │ Days│ URL  │  column headers
    ├────────┼────────┼─────────┼───────────┼─────────┼─────┼──────┤
    │ …data rows for that agency…                                  │
    └──────────────────────────────────────────────────────────────┘
                                                                       blank separator row
    ┌──────────────────────────────────────────────────────────────┐
    │ HOUSTON CITY COLLEGE                                         │
    …

Agencies that were visited and had nothing open still get a block, with a note
in place of the rows: "visited, nothing open" and "not in the report at all"
are different facts, and a report that only shows the former is the one a
reader can trust. Skipped (Incomplete) agencies are listed the same way, so the
sheet accounts for every row of My Network.

Styling follows the SAM / MyFlorida sweep exports so the hub's workbooks read
alike.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.scrapers.ridemetro.models import SHEET_COLUMNS

logger = logging.getLogger(__name__)

SHEET_TITLE = "Open Opportunities"

_BANNER_FILL = PatternFill("solid", fgColor="1E3A5F")
_BANNER_FONT = Font(bold=True, color="FFFFFF", size=14)
_BANNER_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)

_HEADER_FILL = PatternFill("solid", fgColor="E8EEF5")
_HEADER_FONT = Font(bold=True, color="1E3A5F", size=11)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

_DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
_URL_FONT = Font(color="1155CC", underline="single")
_NOTE_FONT = Font(italic=True, color="6B7280")
_NOTE_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)

_EDGE = Side(style="thin", color="D5DDE5")
_CELL_BORDER = Border(left=_EDGE, right=_EDGE, top=_EDGE, bottom=_EDGE)

# Per-column widths, in SHEET_COLUMNS order.
_WIDTHS = (12, 24, 58, 20, 28, 11, 46)

# Excel refuses a cell over 32,767 characters.
_MAX_CELL = 32000


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (list, tuple)):
        value = ", ".join(str(v) for v in value if v not in (None, ""))
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if len(value) > _MAX_CELL:
            value = value[:_MAX_CELL] + " …[truncated]"
    return value


def _merge_across(sheet: Worksheet, row: int) -> None:
    sheet.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=len(SHEET_COLUMNS)
    )


def _write_banner(sheet: Worksheet, row: int, agency: str) -> None:
    """The full-width agency header."""
    _merge_across(sheet, row)
    cell = sheet.cell(row=row, column=1, value=_clean(agency))
    cell.fill = _BANNER_FILL
    cell.font = _BANNER_FONT
    cell.alignment = _BANNER_ALIGN
    # The fill only follows the merge if every covered cell carries it.
    for column in range(1, len(SHEET_COLUMNS) + 1):
        sheet.cell(row=row, column=column).fill = _BANNER_FILL
    sheet.row_dimensions[row].height = 26


def _write_headers(sheet: Worksheet, row: int) -> None:
    for column, (_, header) in enumerate(SHEET_COLUMNS, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _CELL_BORDER
    sheet.row_dimensions[row].height = 20


def _write_record(sheet: Worksheet, row: int, record: dict[str, Any]) -> None:
    for column, (attr, _) in enumerate(SHEET_COLUMNS, start=1):
        value = _clean(record.get(attr))
        cell = sheet.cell(row=row, column=column, value=value)
        cell.alignment = _DATA_ALIGN
        cell.border = _CELL_BORDER
        if attr == "opportunity_url" and isinstance(value, str) and value.startswith("http"):
            cell.hyperlink = value
            cell.font = _URL_FONT


def _write_note(sheet: Worksheet, row: int, note: str) -> None:
    _merge_across(sheet, row)
    cell = sheet.cell(row=row, column=1, value=_clean(note))
    cell.font = _NOTE_FONT
    cell.alignment = _NOTE_ALIGN
    for column in range(1, len(SHEET_COLUMNS) + 1):
        sheet.cell(row=row, column=column).border = _CELL_BORDER


def build(
    groups: Sequence[tuple[str, Iterable[dict[str, Any]]]],
    out_path: str | Path,
    notes: dict[str, str] | None = None,
) -> int:
    """Write the report. Returns the number of opportunity rows written.

    `groups` is (agency name, its opportunity records) in the order the blocks
    should appear — normally the order My Network listed them. `notes` maps an
    agency name to the line shown in place of its rows when it has none.
    """
    notes = notes or {}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    for index, width in enumerate(_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    row = 1
    written = 0
    for agency, records in groups:
        _write_banner(sheet, row, agency)
        row += 1
        _write_headers(sheet, row)
        row += 1

        records = list(records)
        for record in records:
            _write_record(sheet, row, record)
            row += 1
            written += 1
        if not records:
            _write_note(sheet, row, notes.get(agency, "No open public opportunities."))
            row += 1

        row += 1  # blank separator row between agency blocks

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(out_path))
    logger.info(
        "wrote %s — %d agency block(s), %d opportunity row(s)",
        out_path.name, len(groups), written,
    )
    return written
