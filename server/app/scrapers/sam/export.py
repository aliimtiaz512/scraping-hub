"""SAM persistence: store scraped+evaluated bids in the DB and build the per-run
Excel sheet from the DB (openpyxl).

Same batched single-transaction mechanism as the other hub portals: one session
upserts the run row and every bid (de-duplicated by notice_id within the run,
complete record kept in raw_data), one commit. On DB failure the Excel is written
straight from the in-memory records.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.styles import PatternFill
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core import excel_style
from app.db import SessionLocal
from app.scrapers.sam.models import EXCEL_COLUMNS, SamBid, SamRun

logger = logging.getLogger(__name__)

# The header row, widths and cell sanitising are the hub's standard (see
# app/core/excel_style). What is SAM's own is the evaluator's verdict showing in
# the row colour: REJECT tinted red, MANUAL_REVIEW amber so it stands apart from
# a clean REJECT.
_REJECT_FILL = PatternFill("solid", fgColor="FFCCCC")
_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")


def _write_styled_sheet(rows: list[list[Any]], out_path: str | Path) -> None:
    """Write the SAM workbook: the hub's standard sheet, plus decision tinting."""
    headers = [header for _, header in EXCEL_COLUMNS]
    workbook, sheet = excel_style.new_workbook("SAM Bids")
    excel_style.write_table(sheet, headers, rows)

    if "Decision" in headers:
        decision_column = headers.index("Decision") + 1
        for row_idx in range(2, sheet.max_row + 1):
            fill = {
                "REJECT": _REJECT_FILL,
                "MANUAL_REVIEW": _REVIEW_FILL,
            }.get(sheet.cell(row=row_idx, column=decision_column).value)
            if fill is not None:
                for cell in sheet[row_idx]:
                    cell.fill = fill

    workbook.save(str(out_path))

_BID_FIELDS = {
    "notice_id", "title", "department", "subtier", "office", "description",
    "updated_date", "bid_repeat_count", "naics_code", "naics_title",
    "date_offers_due", "published_date", "decision", "reason",
    "ollama_decision", "ollama_rule", "ollama_confidence",
}


def _parse_dt(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(k): _jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(v) for v in value]
    return str(value)


def _run_values(run: dict[str, Any]) -> dict[str, Any]:
    return {
        "run_id": run["run_id"],
        "status": run.get("status"),
        "started_at": _parse_dt(run.get("started_at")),
        "finished_at": _parse_dt(run.get("finished_at")),
        "search": run.get("search"),
        "bids_found": run.get("bids_found", 0),
        "documents_downloaded": run.get("documents_downloaded", 0),
        "folder": run.get("folder"),
        "excel_path": run.get("excel_path"),
    }


def _upsert_run(session, run: dict[str, Any]) -> None:
    values = _run_values(run)
    stmt = pg_insert(SamRun).values(**values)
    stmt = stmt.on_conflict_do_update(
        index_elements=[SamRun.run_id],
        set_={k: v for k, v in values.items() if k != "run_id"},
    )
    session.execute(stmt)


def save_run(run: dict[str, Any]) -> None:
    session = SessionLocal()
    try:
        _upsert_run(session, run)
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _clean(field: str, value: Any) -> Any:
    if field == "bid_repeat_count":
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0
    return value if value not in ("", None) else None


def save_bids(run: dict[str, Any], records: list[dict[str, Any]]) -> int:
    """Upsert a run's bids into sam_bids in one transaction, deduped by notice_id."""
    session = SessionLocal()
    try:
        _upsert_run(session, run)

        stored = 0
        seen: set[str] = set()
        for record in records:
            values: dict[str, Any] = {k: _clean(k, record.get(k)) for k in _BID_FIELDS}
            values["run_id"] = run["run_id"]
            values["raw_data"] = _jsonable(record)

            notice_id = values.get("notice_id")
            if notice_id and notice_id in seen:
                continue
            if notice_id:
                seen.add(notice_id)
                stmt = pg_insert(SamBid).values(**values)
                stmt = stmt.on_conflict_do_update(
                    constraint="uq_sam_run_notice",
                    set_={k: v for k, v in values.items() if k not in ("run_id", "notice_id")},
                )
                session.execute(stmt)
            else:
                session.add(SamBid(**values))
            stored += 1

        session.commit()
        logger.info("[run %s] stored %d SAM bid rows in DB", run.get("run_id"), stored)
        return stored
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _rows_for_run(run_id: str) -> list[SamBid]:
    session = SessionLocal()
    try:
        return session.execute(
            select(SamBid).where(SamBid.run_id == run_id).order_by(SamBid.id)
        ).scalars().all()
    finally:
        session.close()


def generate_excel(run_id: str, out_path: str | Path) -> int:
    """Build this run's Excel sheet from sam_bids (sam-septa styling). Returns row count."""
    rows = _rows_for_run(run_id)
    data = [[getattr(bid, attr, None) for attr, _ in EXCEL_COLUMNS] for bid in rows]
    _write_styled_sheet(data, out_path)
    logger.info("[run %s] wrote %d SAM rows to %s", run_id, len(rows), out_path)
    return len(rows)


def generate_excel_from_records(records: list[dict[str, Any]], out_path: str | Path) -> int:
    """Build the Excel straight from in-memory records (DB-unavailable fallback), same styling."""
    data = [
        [record.get(attr) for attr, _ in EXCEL_COLUMNS]
        for record in records
        if record.get("notice_id")
    ]
    _write_styled_sheet(data, out_path)
    return len(data)
