"""Loader for `mfmp_niches.yaml` — the classifier's only source of constants.

Criteria doc §9.2: no lexicon, commodity code, weight or threshold may live in
Python. Everything the scoring and routing modules use is read from the YAML
through the objects below, so adding a seventh niche is a YAML edit and nothing
else.

Commodity-code validation (§9.6) runs at load time when a v20 workbook path is
configured: any code marked `candidate_requires_validation` that the workbook
does not contain logs a warning and demotes to Tier C. With no workbook
configured the check is skipped with one warning and declared tiers stand —
demoting every code because the reference file is absent would silently gut the
model rather than protect it.
"""

from __future__ import annotations

import logging
import os
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Any

import yaml

logger = logging.getLogger(__name__)

CONFIG_PATH = Path(__file__).resolve().parent / "mfmp_niches.yaml"

# Optional path to the "MFMP Commodity Code v20 Public" workbook used for §9.6
# validation. Unset in most deployments; see the module docstring.
V20_WORKBOOK_ENV = "MFMP_COMMODITY_V20_PATH"

TIER_A, TIER_B, TIER_C = "tier_a", "tier_b", "tier_c"
OTHER = "OTHER"

# Marks a code the criteria doc says must be checked before it is trusted.
NEEDS_VALIDATION = "candidate_requires_validation"


@dataclass(frozen=True)
class CodeEntry:
    code: str
    title: str
    source: str


@dataclass
class Niche:
    key: str
    label: str
    sheet: str
    order: int
    codes: dict[str, list[CodeEntry]]
    core_terms: list[str]
    supporting_terms: list[str]
    umbrella_terms: list[str]
    exclusion_terms: list[str]
    deliverables: list[str]
    stem_map: dict[str, list[str]]

    def tier_of(self, code: str) -> str | None:
        """Which tier `code` sits in for this niche, or None."""
        for tier in (TIER_A, TIER_B, TIER_C):
            if any(entry.code == code for entry in self.codes.get(tier, [])):
                return tier
        return None

    def tier_a_codes(self) -> list[str]:
        return [entry.code for entry in self.codes.get(TIER_A, [])]


@dataclass
class TieBreak:
    """One pair rule from criteria doc §6."""

    a: str
    b: str
    prefer_a: list[str]
    prefer_b: list[str]
    codes_force_b: list[str] = field(default_factory=list)
    contested_when_tied: bool = False


@dataclass
class Config:
    version: str
    cross_listing: bool
    scoring: dict[str, Any]
    thresholds: dict[str, Any]
    high_intent_modifiers: list[str]
    override_niche: str
    override_terms: list[str]
    tie_breaks: dict[tuple[str, str], TieBreak]
    niches: dict[str, Niche]

    def ordered_niches(self) -> list[Niche]:
        return sorted(self.niches.values(), key=lambda n: n.order)

    def tie_break_for(self, x: str, y: str) -> TieBreak | None:
        return self.tie_breaks.get((x, y)) or self.tie_breaks.get((y, x))

    def niches_owning_tier_a(self, code: str) -> list[str]:
        """Every niche key for which `code` is a Tier A code."""
        return [k for k, n in self.niches.items() if code in n.tier_a_codes()]


def _code_entries(raw: list[dict] | None) -> list[CodeEntry]:
    return [
        CodeEntry(
            code=str(item["code"]).strip(),
            title=str(item.get("title", "")).strip(),
            source=str(item.get("source", "")).strip(),
        )
        for item in (raw or [])
    ]


def _lower_list(raw: list | None) -> list[str]:
    return [str(v).strip().lower() for v in (raw or []) if str(v).strip()]


def load_config(path: Path | None = None) -> Config:
    """Parse the YAML into typed objects, then run §9.6 code validation."""
    target = path or CONFIG_PATH
    with open(target, "r", encoding="utf-8") as handle:
        raw = yaml.safe_load(handle) or {}

    niches: dict[str, Niche] = {}
    for key, block in (raw.get("niches") or {}).items():
        codes_block = block.get("codes") or {}
        niches[key] = Niche(
            key=key,
            label=block.get("label", key),
            sheet=block.get("sheet", key),
            order=int(block.get("order", 0)),
            codes={tier: _code_entries(codes_block.get(tier)) for tier in (TIER_A, TIER_B, TIER_C)},
            core_terms=_lower_list(block.get("core_terms")),
            supporting_terms=_lower_list(block.get("supporting_terms")),
            umbrella_terms=_lower_list(block.get("umbrella_terms")),
            exclusion_terms=_lower_list(block.get("exclusion_terms")),
            deliverables=_lower_list(block.get("deliverables")),
            stem_map={
                str(k).strip().lower(): _lower_list(v)
                for k, v in (block.get("stem_map") or {}).items()
            },
        )

    tie_breaks: dict[tuple[str, str], TieBreak] = {}
    for pair, rule in (raw.get("tie_breaks") or {}).items():
        a, _, b = str(pair).partition("|")
        a, b = a.strip(), b.strip()
        if not (a and b):
            logger.warning("tie_break key %r is not in 'NX|NY' form — skipped", pair)
            continue
        tie_breaks[(a, b)] = TieBreak(
            a=a,
            b=b,
            prefer_a=_lower_list(rule.get("prefer_a")),
            prefer_b=_lower_list(rule.get("prefer_b")),
            codes_force_b=[str(c).strip() for c in (rule.get("codes_force_b") or [])],
            contested_when_tied=bool(rule.get("contested_when_tied", False)),
        )

    routing = (raw.get("routing") or {}).get("hard_primary_override") or {}
    config = Config(
        version=str(raw.get("version", "")),
        cross_listing=bool(raw.get("cross_listing", False)),
        scoring=raw.get("scoring") or {},
        thresholds=raw.get("thresholds") or {},
        high_intent_modifiers=_lower_list(raw.get("high_intent_modifiers")),
        override_niche=str(routing.get("niche", "")).strip(),
        override_terms=_lower_list(routing.get("terms")),
        tie_breaks=tie_breaks,
        niches=niches,
    )
    _validate_codes(config)
    return config


# -- §9.6 commodity-code validation ------------------------------------------

def _load_v20_codes(path: Path) -> set[str]:
    """Every 8-digit code in the MFMP v20 workbook, from any column."""
    from openpyxl import load_workbook

    found: set[str] = set()
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    text = str(cell).strip() if cell is not None else ""
                    if len(text) == 8 and text.isdigit():
                        found.add(text)
    finally:
        workbook.close()
    return found


def _validate_codes(config: Config) -> None:
    """Demote unvalidated `candidate_requires_validation` codes to Tier C.

    Only runs when a workbook is configured and readable. A missing workbook is
    reported once and changes nothing — see the module docstring for why.
    """
    raw_path = os.getenv(V20_WORKBOOK_ENV, "").strip()
    if not raw_path:
        pending = sum(
            1
            for niche in config.niches.values()
            for tier in (TIER_A, TIER_B)
            for entry in niche.codes.get(tier, [])
            if entry.source == NEEDS_VALIDATION
        )
        if pending:
            logger.warning(
                "%s is not set — skipping commodity-code validation; %d Tier A/B code(s) "
                "marked %s are being trusted as declared",
                V20_WORKBOOK_ENV, pending, NEEDS_VALIDATION,
            )
        return

    path = Path(raw_path)
    if not path.is_file():
        logger.warning("%s points at %s which does not exist — validation skipped",
                       V20_WORKBOOK_ENV, path)
        return

    try:
        known = _load_v20_codes(path)
    except Exception:  # noqa: BLE001 — validation must never break a run
        logger.exception("could not read the v20 workbook at %s — validation skipped", path)
        return

    for niche in config.niches.values():
        for tier in (TIER_A, TIER_B):
            keep, demote = [], []
            for entry in niche.codes.get(tier, []):
                if entry.source == NEEDS_VALIDATION and entry.code not in known:
                    demote.append(entry)
                else:
                    keep.append(entry)
            if demote:
                niche.codes[tier] = keep
                niche.codes[TIER_C].extend(demote)
                logger.warning(
                    "%s: demoted %d unvalidated code(s) from %s to tier_c: %s",
                    niche.key, len(demote), tier, ", ".join(e.code for e in demote),
                )


@lru_cache(maxsize=1)
def get_config() -> Config:
    """The process-wide config. Cached; call `reload_config` after an edit."""
    return load_config()


def reload_config() -> Config:
    get_config.cache_clear()
    return get_config()
