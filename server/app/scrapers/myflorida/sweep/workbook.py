"""The multi-sheet output workbook — one tab per niche, plus Other.

The criteria doc stops at the §7 result object ("This spec does not define
storage"), so the sheet layout is this module's decision. Two choices worth
naming:

  * Every row carries all six niche scores. §5.2 wants a threshold change
    replayable over history without re-fetching; six columns turns that into a
    spreadsheet filter rather than a re-run.
  * Only the sheet's own niche gets its C/T/S broken out. The full 6 x C/T/S is
    18 columns and would drown the sheet — it lives in mfmp_sweep_scores, where
    replay queries actually want it.

Styling matches the SAM export so the two portals' workbooks read alike.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.core import excel_style
from app.scrapers.myflorida.sweep.config import OTHER, Config
from app.scrapers.myflorida.sweep.models import OTHER_EXTRA_COLUMNS, SHEET_COLUMNS

logger = logging.getLogger(__name__)

# Cross-listed rows are tinted so a reviewer can see at a glance that the ad is
# owned by another lane and has already been counted there. Everything else
# about how these sheets look — header row, widths, cell sanitising — is the
# hub's standard (see app/core/excel_style).
_CROSS_FILL = PatternFill("solid", fgColor="EAF2FB")

OTHER_SHEET = "Other"


def _cell(value: Any) -> Any:
    """A sweep value as a cell: lists joined, then the hub's standard cleanup."""
    if isinstance(value, (list, tuple)):
        value = ", ".join(str(v) for v in value if v not in (None, ""))
    return excel_style.sanitize_cell(value)


def _write_sheet(workbook: Workbook, title: str, columns, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(title=title)
    excel_style.write_table(
        sheet,
        [header for _, header in columns],
        ([_cell(row.get(key)) for key, _ in columns] for row in rows),
    )

    role_index = next((i for i, (key, _) in enumerate(columns) if key == "role"), None)
    if role_index is not None:
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_index, column=role_index + 1).value == "CROSS-LISTED":
                for cell in sheet[row_index]:
                    cell.fill = _CROSS_FILL


def build_workbook(config: Config, rows_by_lane: dict[str, list[dict]], out_path: Path) -> Path:
    """Write one sheet per niche in the config's declared order, then Other.

    Empty niches still get their sheet, with headers only, so the workbook's
    shape is stable run to run and a reader can tell "nothing matched N6" apart
    from "N6 was dropped".
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default sheet

    for niche in config.ordered_niches():
        _write_sheet(workbook, niche.sheet, SHEET_COLUMNS, rows_by_lane.get(niche.key, []))

    _write_sheet(
        workbook, OTHER_SHEET, SHEET_COLUMNS + OTHER_EXTRA_COLUMNS,
        rows_by_lane.get(OTHER, []),
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(out_path))
    counts = {lane: len(rows) for lane, rows in rows_by_lane.items() if rows}
    logger.info("wrote %s (%s)", out_path.name, counts or "empty")
    return out_path
