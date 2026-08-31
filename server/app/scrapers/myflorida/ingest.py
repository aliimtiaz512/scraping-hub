"""Read an MFMP Excel export and store its rows in the database.

The export's exact column headers are not known until we run against the live
portal, so mapping is best-effort: normalized header names are matched against
a set of candidates for each known field, and the complete original row is always
kept in `raw_data` so nothing is lost.
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.db import SessionLocal
from app.scrapers.myflorida.models import Bid, ScrapeRun

logger = logging.getLogger(__name__)


def _normalize(header: str) -> str:
    """lowercase, drop everything but letters/digits — 'Ad #' -> 'ad'."""
    return re.sub(r"[^a-z0-9]", "", str(header).lower())


# Map a model field -> normalized header candidates. Each is tried as an exact
# header match first and only then as a substring; see `map_row`.
#
# The first name in each tuple is our own summary sheet's header (workbook.py),
# so the sheet reads back exactly. The rest are the portal's export and the
# other spellings a procurement export arrives with, and are what the substring
# pass is for. Fields the sheet carries but the model has no column for —
# agency advertisement number, version, the detail page URL — are absent here on
# purpose: they reach the database through `raw_data`, which keeps the whole row.
FIELD_CANDIDATES: dict[str, tuple[str, ...]] = {
    "ad_number": ("advertisementnumber", "adnumber", "adid", "number", "solicitationnumber", "bidnumber"),
    "title": ("title", "adtitle", "name", "solicitationtitle"),
    # Written by runs made before the fixed summary schema; kept so their rows
    # still read back. No sheet produces this column now.
    "matched_keyword": ("matchedkeyword",),
    "agency": ("agency", "organization", "customer", "department", "buyer"),
    "ad_type": ("advertisementtype", "adtype", "type", "solicitationtype", "method"),
    "status": ("status", "state"),
    "description": ("description", "summary", "scope"),
    "commodity_codes": ("commoditycodes", "commoditycode", "commodity", "nigp", "unspsc"),
    "contact_name": ("contactperson", "contactname", "contact", "buyername"),
    "contact_email": ("contactemail", "email"),
    "contact_phone": ("contactphone", "phone", "telephone"),
    "estimated_amount": ("estimatedamount", "amount", "estimatedvalue", "value"),
    "ad_date": ("publisheddate", "addate", "advertisedate", "advertisementdate", "begindate", "startdate", "posteddate"),
    "open_date": ("opendate", "openingdate", "responsedate", "duedate", "responsedue"),
    "close_date": ("closingdate", "closedate", "enddate", "expirationdate"),
}

STRING_FIELDS = {
    "ad_number", "title", "matched_keyword", "agency", "ad_type", "status", "description",
    "commodity_codes", "contact_name", "contact_email", "contact_phone",
    "ad_date", "open_date", "close_date",
}


def _find_header_row(rows: list[list[Any]]) -> int:
    """Return the index of the row most likely to be the header.

    Some exports have a title/blank line before the header, so we pick the first
    row that has several non-empty text-ish cells.
    """
    for i, row in enumerate(rows[:10]):
        non_empty = [c for c in row if c not in (None, "")]
        if len(non_empty) >= 3:
            return i
    return 0


def parse_excel(path: str | Path) -> list[dict[str, Any]]:
    """Return the export as a list of {header: value} dicts (pure — no DB)."""
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    workbook.close()
    if not rows:
        return []

    header_idx = _find_header_row(rows)
    headers = [str(h).strip() if h is not None else f"column_{i}" for i, h in enumerate(rows[header_idx])]

    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        if all(cell in (None, "") for cell in row):
            continue
        record = {}
        for header, value in zip(headers, row):
            if isinstance(value, datetime):
                value = value.isoformat()
            record[header] = value
        records.append(record)
    return records


def map_row(raw: dict[str, Any]) -> dict[str, Any]:
    """Map a raw {header: value} row to known Bid fields (pure — no DB).

    Every candidate is tried as an **exact** header match before any is tried as
    a substring. Substring-first is what a portal export needs, where the header
    might be "Ad #" or "Advertisement No."; it is actively wrong on our own
    summary sheet, where "Agency" and "Agency Advertisement Number" are two
    different columns and the substring pass reaches the longer one first,
    filing the ad's agency number under its agency name. Exact-first settles
    that without loosening anything: a header that matches a candidate outright
    is the column that candidate names.
    """
    normalized = {_normalize(h): h for h in raw}
    mapped: dict[str, Any] = {}

    def _match(candidate: str, exact: bool) -> str | None:
        if exact:
            return normalized.get(candidate)
        return next((orig for norm, orig in normalized.items() if candidate in norm), None)

    for exact in (True, False):
        for field, candidates in FIELD_CANDIDATES.items():
            if field in mapped:
                continue
            for candidate in candidates:
                match = _match(candidate, exact)
                if match is not None and raw.get(match) not in (None, ""):
                    mapped[field] = raw[match]
                    break

    if "estimated_amount" in mapped:
        mapped["estimated_amount"] = _to_decimal(mapped["estimated_amount"])
    for field in STRING_FIELDS:
        if field in mapped and mapped[field] is not None:
            mapped[field] = str(mapped[field]).strip()

    return mapped


def _to_decimal(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


def _jsonable(raw: dict[str, Any]) -> dict[str, Any]:
    """Ensure every value in the raw row is JSON-serializable for JSONB."""
    out = {}
    for key, value in raw.items():
        if value is None or isinstance(value, (str, int, float, bool)):
            out[key] = value
        else:
            out[key] = str(value)
    return out


def ingest_excel(excel_path: str | Path, run: dict[str, Any]) -> int:
    """Parse the export and upsert its rows into mfmp_bids for a run.

    `run` is the run dict from run_manager. Returns the number of rows stored.
    """
    records = parse_excel(excel_path)
    if not records:
        logger.warning("[run %s] Excel export had no data rows", run.get("run_id"))

    session = SessionLocal()
    try:
        _upsert_run(session, run, excel_path)

        stored = 0
        seen_ad_numbers: set[str] = set()
        for raw in records:
            mapped = map_row(raw)
            ad_number = mapped.get("ad_number")
            # Skip duplicate ad numbers within the same export.
            if ad_number and ad_number in seen_ad_numbers:
                continue
            if ad_number:
                seen_ad_numbers.add(ad_number)

            values = {
                "run_id": run["run_id"],
                "category": run.get("category"),
                "priority": run.get("priority"),
                "raw_data": _jsonable(raw),
                **mapped,
            }

            if ad_number:
                stmt = pg_insert(Bid).values(**values)
                stmt = stmt.on_conflict_do_update(
                    constraint="uq_bid_run_ad",
                    set_={k: v for k, v in values.items() if k not in ("run_id", "ad_number")},
                )
                session.execute(stmt)
            else:
                session.add(Bid(**values))
            stored += 1

        session.commit()
        logger.info("[run %s] stored %d bid rows in DB", run.get("run_id"), stored)
        return stored
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _upsert_run(session, run: dict[str, Any], excel_path: str | Path) -> None:
    values = {
        "run_id": run["run_id"],
        "category": run.get("category"),
        "category_label": run.get("category_label"),
        "priority": run.get("priority"),
        "codes": run.get("codes"),
        "status": run.get("status"),
        "started_at": _parse_dt(run.get("started_at")),
        "finished_at": _parse_dt(run.get("finished_at")),
        "bids_found": run.get("bids_found", 0),
        "documents_downloaded": run.get("documents_downloaded", 0),
        "excel_path": str(excel_path),
    }
    stmt = pg_insert(ScrapeRun).values(**values)
    stmt = stmt.on_conflict_do_update(
        index_elements=[ScrapeRun.run_id],
        set_={k: v for k, v in values.items() if k != "run_id"},
    )
    session.execute(stmt)


def _parse_dt(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None
